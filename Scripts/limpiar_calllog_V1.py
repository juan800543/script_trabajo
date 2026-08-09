# -*- coding: utf-8 -*-
"""
LIMPIEZA DEL EXCEL DE REGISTRO DE LLAMADAS (CallLog)
====================================================

Qué hace el script, en orden:
  1. Lee el archivo Excel de entrada (.xlsx).
  2. Columna "Date": quita el nombre del día de la semana
     ("Thu 07/30/2026"  ->  "07/30/2026").  NO reordena día y mes.
  3. Columna "Direction": conserva SOLO las filas "Incoming".
     Elimina las "Outgoing" y las que estén vacías.
  4. Elimina por completo 5 columnas:
     Extension, Forwarded To, Result Description, Included, Purchased.
  5. Guarda el resultado en un Excel NUEVO. El original nunca se modifica.

Pensado para ejecutarse con el botón Run / F5 de Visual Studio Code,
SIN usar CMD ni terminal. No lanza comandos externos de ningún tipo.

(Al final del archivo está la nota de qué librería hay que instalar.)
"""

from datetime import date, datetime
from pathlib import Path
import re
import sys


# ============================================================
#  CONFIGURACIÓN  ->  CAMBIA AQUÍ CARPETAS Y NOMBRES DE ARCHIVO
# ============================================================
#
#  La carpeta y el nombre van por separado, para que puedas cambiar
#  solo lo que necesites. Reglas de las CARPETAS:
#
#     ""                          -> la misma carpeta donde está este .py
#     r"C:\Trabajo\Llamadas"      -> esa carpeta exacta (ruta completa)
#     r"datos"                    -> subcarpeta 'datos' dentro de la del .py
#
#  Escribe siempre la r delante de las comillas en rutas de Windows:
#  r"C:\Trabajo\..."  evita problemas con las barras invertidas.
#
#  También puedes poner la ruta completa directamente en ARCHIVO_ENTRADA
#  o ARCHIVO_SALIDA; en ese caso manda esa ruta y la variable de carpeta
#  correspondiente se ignora.

# --- ARCHIVO DE ENTRADA (el original, nunca se modifica) ---
CARPETA_ENTRADA =  r"C:\Trabajo\prueba_en"  # "" = carpeta de este script
ARCHIVO_ENTRADA = "CallLog_20260731-211238-Crudo.xlsx"           # nombre con extensión .xlsx

# --- ARCHIVO DE SALIDA (se crea nuevo; si ya existe, se sobrescribe) ---
CARPETA_SALIDA = r"C:\Trabajo\prueba_sal"  # "" = carpeta de este script
ARCHIVO_SALIDA = "CallLog_limpio.xlsx"     # nombre con extensión .xlsx

# Qué hacer si la carpeta de salida no existe:
#   True  -> el script la crea solo y avisa por pantalla.
#   False -> el script se detiene con un error (útil para cazar erratas).
CREAR_CARPETA_SALIDA = True

# Hoja del Excel de entrada que se quiere procesar.
#   None  -> usa la primera hoja del archivo (lo normal).
#   "Hoja1" -> usa la hoja con ese nombre exacto.
NOMBRE_HOJA = None


# ============================================================
#  REGLAS DE LIMPIEZA (normalmente no hace falta tocar nada)
# ============================================================

# Solo se conservan las filas cuya columna Direction sea este valor.
DIRECCION_A_CONSERVAR = "Incoming"

# Columnas que se eliminan por completo del archivo de salida.
COLUMNAS_A_ELIMINAR = [
    "Extension",
    "Forwarded To",
    "Result Description",
    "Included",
    "Purchased",
]

# Nombres de las columnas clave en el archivo original.
COLUMNA_DIRECCION = "Direction"
COLUMNA_FECHA = "Date"


# ============================================================
#  IMPORTACIÓN DE LA LIBRERÍA EXTERNA
# ============================================================
# Se importa dentro de un try para poder avisar con un mensaje claro
# si openpyxl no está instalada, en lugar de soltar un error feo.
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("=" * 60)
    print("ERROR: falta la librería 'openpyxl', que es la que lee y")
    print("       escribe archivos Excel (.xlsx).")
    print()
    print("Instálala una sola vez con:   pip install openpyxl")
    print("(mira la nota del final de este script para más detalle)")
    print("=" * 60)
    sys.exit(1)


# ============================================================
#  FUNCIONES AUXILIARES
# ============================================================

def carpeta_del_script():
    """Carpeta donde está guardado este archivo .py."""
    return Path(__file__).resolve().parent


def ruta_completa(carpeta, nombre_archivo):
    """
    Une carpeta + nombre y devuelve la ruta final, según estas reglas:

      - Si el NOMBRE ya trae una ruta completa (r"C:\\...\\archivo.xlsx"),
        manda esa ruta y la carpeta se ignora.
      - Si la CARPETA está vacía (""), se usa la carpeta de este script.
      - Si la CARPETA es una ruta completa, se usa tal cual.
      - Si la CARPETA es un nombre suelto ("datos"), se entiende como
        subcarpeta dentro de la carpeta de este script.
    """
    ruta_nombre = Path(str(nombre_archivo).strip())

    # El nombre ya incluye la ruta completa: tiene prioridad.
    if ruta_nombre.is_absolute():
        return ruta_nombre

    texto_carpeta = str(carpeta).strip()

    if texto_carpeta == "":
        base = carpeta_del_script()
    else:
        base = Path(texto_carpeta)
        if not base.is_absolute():
            # Carpeta relativa: se cuelga de la carpeta del script.
            base = carpeta_del_script() / base

    return base / ruta_nombre


def como_texto(valor):
    """Convierte cualquier celda a texto limpio ('' si está vacía)."""
    if valor is None:
        return ""
    return str(valor).strip()


def valor_de(fila, indice):
    """
    Devuelve el valor de una posición de la fila.
    Algunas filas de Excel vienen más cortas que la cabecera; en ese
    caso se devuelve None en vez de reventar con un IndexError.
    """
    if indice < len(fila):
        return fila[indice]
    return None


# Detecta el nombre del día al principio del texto: "Thu ", "Thursday, ", "mon. "...
# Acepta las 7 abreviaturas (Mon, Tue, Wed, Thu, Fri, Sat, Sun) sin importar
# mayúsculas/minúsculas, con o sin punto y con o sin coma detrás.
PATRON_DIA_SEMANA = re.compile(
    r"^\s*(?:mon|tue|wed|thu|fri|sat|sun)[a-z]*\.?[\s,]+",
    re.IGNORECASE,
)

# Detecta una fecha ya en formato mm/dd/yyyy (admitiendo 7/5/2026 sin ceros).
PATRON_FECHA = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def limpiar_fecha(valor):
    """
    Deja la fecha como 'mm/dd/yyyy' sin el nombre del día.
    IMPORTANTE: no se reordena nada, solo se quita el prefijo del día
    y se rellenan con cero el mes y el día si vienen con una sola cifra.
    """
    if valor is None:
        return None

    # Caso 1: Excel guardó la celda como fecha real, no como texto.
    # Se formatea directamente en mm/dd/yyyy.
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%m/%d/%Y")

    # Caso 2 (el habitual): la celda es texto tipo "Thu 07/30/2026".
    texto_fecha = como_texto(valor)
    texto_fecha = PATRON_DIA_SEMANA.sub("", texto_fecha).strip()

    # Si lo que queda es una fecha mm/dd/yyyy, se normaliza a dos cifras.
    coincidencia = PATRON_FECHA.match(texto_fecha)
    if coincidencia:
        mes, dia, anio = coincidencia.groups()
        return "{:02d}/{:02d}/{}".format(int(mes), int(dia), anio)

    # Si tiene otro formato inesperado, se devuelve tal cual (ya sin el día)
    # para no perder información.
    return texto_fecha


def fila_totalmente_vacia(fila):
    """True si todas las celdas de la fila están vacías."""
    return all(como_texto(celda) == "" for celda in fila)


# ============================================================
#  PROCESO PRINCIPAL
# ============================================================

def main():
    ruta_entrada = ruta_completa(CARPETA_ENTRADA, ARCHIVO_ENTRADA)
    ruta_salida = ruta_completa(CARPETA_SALIDA, ARCHIVO_SALIDA)

    # --- Comprobación 1: que el archivo de entrada exista ---
    if not ruta_entrada.exists():
        print("=" * 60)
        print("ERROR: no se encontró el archivo de entrada.")
        print()
        print("  Se buscó aquí: {}".format(ruta_entrada))
        print()
        if not ruta_entrada.parent.exists():
            # El fallo está en la carpeta, no en el nombre del archivo.
            print("  La carpeta ni siquiera existe: {}".format(ruta_entrada.parent))
            print("  Revisa CARPETA_ENTRADA arriba en la configuración.")
        else:
            print("  La carpeta sí existe, pero no hay ningún archivo con ese nombre.")
            print("  Revisa que ARCHIVO_ENTRADA sea exacto, con la extensión .xlsx.")
            # Se listan los Excel que sí hay, para ver si es una errata del nombre.
            excels = sorted(p.name for p in ruta_entrada.parent.glob("*.xls*"))
            if excels:
                print("  Archivos Excel encontrados en esa carpeta:")
                for nombre in excels[:10]:
                    print("     - {}".format(nombre))
        print("=" * 60)
        return

    # --- Comprobación 2: no sobrescribir el original por accidente ---
    if ruta_salida.resolve() == ruta_entrada.resolve():
        print("ERROR: el archivo de salida apunta al mismo sitio que el de entrada:")
        print("       {}".format(ruta_salida))
        print("       Cambia CARPETA_SALIDA o ARCHIVO_SALIDA para no tocar el original.")
        return

    # --- Comprobación 3: que exista la carpeta donde se va a guardar ---
    if not ruta_salida.parent.exists():
        if CREAR_CARPETA_SALIDA:
            ruta_salida.parent.mkdir(parents=True, exist_ok=True)
            print("Aviso: se creó la carpeta de salida: {}".format(ruta_salida.parent))
        else:
            print("ERROR: la carpeta de salida no existe: {}".format(ruta_salida.parent))
            print("       Créala a mano, corrige CARPETA_SALIDA,")
            print("       o pon CREAR_CARPETA_SALIDA = True para que se cree sola.")
            return

    print("Leyendo:  {}".format(ruta_entrada))

    # data_only=True  -> si hay fórmulas, se lee el resultado, no la fórmula.
    # read_only=True  -> lectura rápida y con poca memoria.
    libro_entrada = load_workbook(filename=ruta_entrada, data_only=True, read_only=True)

    # --- Selección de la hoja ---
    if NOMBRE_HOJA:
        if NOMBRE_HOJA not in libro_entrada.sheetnames:
            print("ERROR: no existe la hoja '{}'.".format(NOMBRE_HOJA))
            print("       Hojas disponibles: {}".format(", ".join(libro_entrada.sheetnames)))
            libro_entrada.close()
            return
        hoja_entrada = libro_entrada[NOMBRE_HOJA]
    else:
        hoja_entrada = libro_entrada[libro_entrada.sheetnames[0]]

    filas = hoja_entrada.iter_rows(values_only=True)

    # --- Cabecera ---
    try:
        cabecera_original = next(filas)
    except StopIteration:
        print("ERROR: la hoja '{}' está vacía.".format(hoja_entrada.title))
        libro_entrada.close()
        return

    encabezados = [como_texto(celda) for celda in cabecera_original]
    # Excel a veces arrastra columnas fantasma al final: se descartan.
    while encabezados and encabezados[-1] == "":
        encabezados.pop()

    # Diccionario nombre-en-minúsculas -> posición, para localizar columnas
    # aunque el Excel venga con distinta capitalización.
    posiciones = {nombre.lower(): i for i, nombre in enumerate(encabezados)}

    # --- Comprobación 3: que existan las columnas que necesitamos ---
    faltantes = [c for c in (COLUMNA_DIRECCION, COLUMNA_FECHA) if c.lower() not in posiciones]
    if faltantes:
        print("ERROR: faltan columnas obligatorias en el Excel: {}".format(", ".join(faltantes)))
        print("       Columnas encontradas: {}".format(", ".join(encabezados)))
        libro_entrada.close()
        return

    indice_direccion = posiciones[COLUMNA_DIRECCION.lower()]
    indice_fecha = posiciones[COLUMNA_FECHA.lower()]

    # --- Qué columnas se conservan (todas menos las 5 a eliminar) ---
    eliminar = {nombre.lower() for nombre in COLUMNAS_A_ELIMINAR}
    indices_conservados = [i for i, nombre in enumerate(encabezados) if nombre.lower() not in eliminar]
    encabezados_finales = [encabezados[i] for i in indices_conservados]

    # Aviso si alguna de las 5 columnas no estaba en el archivo (no es un error).
    no_encontradas = [c for c in COLUMNAS_A_ELIMINAR if c.lower() not in posiciones]
    if no_encontradas:
        print("Aviso: estas columnas a eliminar no estaban en el archivo: {}".format(
            ", ".join(no_encontradas)))

    # --- Libro de salida ---
    libro_salida = Workbook()
    hoja_salida = libro_salida.active
    hoja_salida.title = "CallLog"
    hoja_salida.append(encabezados_finales)

    # Contadores para el resumen final.
    total_leidas = 0
    conservadas = 0
    quitadas_outgoing = 0
    quitadas_vacias = 0
    quitadas_otras = 0

    # --- Recorrido de las filas de datos ---
    for fila in filas:
        # Filas completamente en blanco: se saltan sin contarlas.
        if fila_totalmente_vacia(fila):
            continue

        total_leidas += 1
        direccion = como_texto(valor_de(fila, indice_direccion))

        # Direction vacío -> fuera.
        if direccion == "":
            quitadas_vacias += 1
            continue

        # Direction distinto de "Incoming" -> fuera (Outgoing u otro valor).
        if direccion.lower() != DIRECCION_A_CONSERVAR.lower():
            if direccion.lower() == "outgoing":
                quitadas_outgoing += 1
            else:
                quitadas_otras += 1
            continue

        # La fila se conserva: se copian solo las columnas que sobreviven,
        # limpiando la fecha al pasar por la columna Date.
        fila_nueva = []
        for i in indices_conservados:
            valor = valor_de(fila, i)
            if i == indice_fecha:
                valor = limpiar_fecha(valor)
            fila_nueva.append(valor)

        hoja_salida.append(fila_nueva)
        conservadas += 1

    libro_entrada.close()

    # --- Formato de la columna Date como TEXTO ---
    # Así Excel no la reinterpreta como fecha y siempre se ve mm/dd/yyyy.
    if indice_fecha in indices_conservados:
        columna_fecha_salida = indices_conservados.index(indice_fecha) + 1  # 1 = columna A
        for celdas in hoja_salida.iter_rows(min_row=2,
                                            min_col=columna_fecha_salida,
                                            max_col=columna_fecha_salida):
            for celda in celdas:
                celda.number_format = "@"

    # --- Ancho de columnas, solo para que se lea cómodo al abrirlo ---
    for numero_columna, nombre in enumerate(encabezados_finales, start=1):
        ancho = min(32, max(12, len(nombre) + 4))
        hoja_salida.column_dimensions[get_column_letter(numero_columna)].width = ancho

    # --- Guardado ---
    try:
        libro_salida.save(ruta_salida)
    except PermissionError:
        print("ERROR: no se pudo guardar '{}'.".format(ruta_salida.name))
        print("       Suele pasar cuando el archivo está abierto en Excel.")
        print("       Ciérralo y vuelve a ejecutar el script.")
        return

    # --- Resumen ---
    print()
    print("=" * 60)
    print("LIMPIEZA TERMINADA")
    print("=" * 60)
    print("Filas leídas (sin contar la cabecera): {}".format(total_leidas))
    print("  - Conservadas (Incoming):            {}".format(conservadas))
    print("  - Eliminadas (Outgoing):             {}".format(quitadas_outgoing))
    print("  - Eliminadas (Direction vacío):      {}".format(quitadas_vacias))
    if quitadas_otras:
        print("  - Eliminadas (otro valor):           {}".format(quitadas_otras))
    print()
    print("Columnas eliminadas: {}".format(", ".join(COLUMNAS_A_ELIMINAR)))
    print("Columnas en el resultado ({}): {}".format(
        len(encabezados_finales), ", ".join(encabezados_finales)))
    print()
    print("Archivo original (intacto): {}".format(ruta_entrada))
    print("Archivo generado:           {}".format(ruta_salida))
    print("=" * 60)


# Punto de entrada: esto es lo que se ejecuta al pulsar Run / F5 en VS Code.
if __name__ == "__main__":
    main()


# ============================================================================
#  LIBRERÍAS EXTERNAS QUE HAY QUE INSTALAR  (NO instaladas por este script)
# ============================================================================
#
#  Solo hace falta UNA:
#
#      openpyxl
#
#  Comando de instalación:
#
#      pip install openpyxl
#
#  Si pip no funciona directamente, esta variante suele resolverlo porque
#  usa el mismo Python que ejecuta VS Code:
#
#      python -m pip install openpyxl
#
#  Notas para equipos con CMD/terminal bloqueado:
#   - El script en sí NO necesita terminal: se ejecuta con Run / F5.
#     La instalación de openpyxl es un paso previo que se hace UNA SOLA VEZ.
#   - Si no puedes ejecutar pip, pide a IT/soporte que instalen openpyxl
#     en el intérprete de Python que usa VS Code
#     (se ve abajo a la derecha en la barra de estado de VS Code).
#   - Todo lo demás que usa el script (pathlib, re, sys, datetime) viene
#     incluido en Python, no hay que instalar nada más.
#   - NO se usa pandas a propósito: es una instalación mucho más pesada y
#     openpyxl basta para leer y escribir .xlsx.
# ============================================================================
