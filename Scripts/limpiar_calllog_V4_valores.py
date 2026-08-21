# -*- coding: utf-8 -*-
"""
LIMPIEZA DEL EXCEL DE REGISTRO DE LLAMADAS (CallLog)
====================================================

Qué hace el script, en orden:
  1. Lee el archivo de entrada, que puede ser .xlsx o .csv.
  2. Columna "Date": quita el nombre del día de la semana
     ("Thu 07/30/2026"  ->  "07/30/2026").  NO reordena día y mes.
  3. Elimina por completo 5 columnas:
     Extension, Forwarded To, Result Description, Included, Purchased.
  4. Aplica un FILTRO de Excel al resultado, que oculta:
       - las filas cuya Direction no sea "Incoming"  (Outgoing y vacías)
       - las filas cuyo Type no sea "Voice"          (faxes y otros)
       - las filas con la columna "To" vacía
       - las filas cuyo número no está en la tabla de campañas, que son
         las que saldrían con un 0 en la columna Queue Name
     Las filas NO se borran: siguen en el archivo. Quitando el filtro en
     Excel vuelven a aparecer todas.
  5. Aplica FORMATO DE CELDA de Excel a dos columnas:
       - Date -> mm/dd/yyyy       (fecha real, no texto)
       - Time -> h:mm AM/PM       (hora real, no texto)
     Son los mismos códigos que aparecen en Excel en
     "Formato de celdas > Personalizada".
  6. Añade 5 COLUMNAS CALCULADAS al final, con el valor ya calculado en
     Python (no son fórmulas de Excel):
       - Day        -> nombre del día de la semana a partir de Date
       - Queue Name -> campaña a la que pertenece el número de To
       - 7          -> 1 si la llamada dura más del umbral (7 minutos), si no 0
       - #Calls     -> 1 si hay Action Result, si no 0
       - Missed     -> 1 si Action Result es "Missed", si no 0
     Al ser valores, el archivo se puede leer desde cualquier herramienta sin
     abrirlo antes en Excel, y abre más rápido. A cambio, no se actualizan
     solas si luego se edita una celda a mano: hay que volver a ejecutar.
     La hoja "Campañas" se sigue creando, pero ya solo como referencia de
     con qué tabla y con qué umbral se generó el archivo.
  7. Guarda el resultado en un archivo .xlsx NUEVO, con la fecha y la hora
     en el nombre para no pisar los anteriores. El original nunca se toca.

Pensado para ejecutarse con el botón Run / F5 de Visual Studio Code,
SIN usar CMD ni terminal. No lanza comandos externos de ningún tipo.

(Al final del archivo está la nota de qué librería hay que instalar.)
"""

import csv
import io
import re
import sys
from datetime import date, datetime, time, timedelta
from pathlib import Path


# ============================================================
#  CONFIGURACIÓN  ->  CAMBIA AQUÍ CARPETAS Y NOMBRES DE ARCHIVO
# ============================================================
#
#  La carpeta y el nombre van por separado, para que puedas cambiar
#  solo lo que necesites. Reglas de las CARPETAS:
#
#     ""                          -> la misma carpeta donde está este .py
#     r"C:\Trabajo\Llamadas"      -> esa carpeta exacta (ruta completa)
#     r"datos"                    -> subcarpeta 'datos' dentro de la del .py
#
#  Escribe siempre la r delante de las comillas en rutas de Windows:
#  r"C:\Trabajo\..."  evita problemas con las barras invertidas.
#
#  También puedes poner la ruta completa directamente en ARCHIVO_ENTRADA
#  o ARCHIVO_SALIDA; en ese caso manda esa ruta y la variable de carpeta
#  correspondiente se ignora.

# --- ARCHIVO DE ENTRADA (el original, nunca se modifica) ---
# Admite .xlsx y .csv: el script detecta el tipo por la extensión.
CARPETA_ENTRADA = r"C:\Users\Guillermo Valencia\Documents\py\Trabajados"   # "" = carpeta de este script
ARCHIVO_ENTRADA = "crudo.csv"

# --- ARCHIVO DE SALIDA (siempre .xlsx, siempre uno nuevo) ---
CARPETA_SALIDA = r"C:\Users\Guillermo Valencia\Documents\py\Marketing"   # "" = carpeta de este script
ARCHIVO_SALIDA = "CallLog_limpio.xlsx"      # nombre base, ver abajo

# Añadir fecha y hora al nombre del archivo generado, para que cada
# ejecución cree uno nuevo y no se pierdan los anteriores:
#
#   CallLog_limpio.xlsx  ->  CallLog_limpio_20260803-1945.xlsx
#                                          año mes día - hora minutos
#
NOMBRE_CON_FECHA_HORA = True
FORMATO_FECHA_HORA = "%Y%m%d-%H%M"

# Qué hacer si la carpeta de salida no existe:
#   True  -> el script la crea solo y avisa por pantalla.
#   False -> el script se detiene con un error (útil para cazar erratas).
CREAR_CARPETA_SALIDA = True

# Hoja que se procesa cuando la entrada es un .xlsx (en CSV no aplica).
#   None    -> usa la primera hoja del archivo (lo normal).
#   "Hoja1" -> usa la hoja con ese nombre exacto.
NOMBRE_HOJA = None


# ============================================================
#  REGLAS DE LIMPIEZA (normalmente no hace falta tocar nada)
# ============================================================

# Columnas en las que solo debe quedar VISIBLE un valor concreto. Es lo
# mismo que abrir el desplegable del filtro de esa columna y dejar marcado
# solo ese valor: todo lo demás (y los vacíos) se oculta.
#
#   "Direction": "Incoming"  -> fuera las salientes
#   "Type": "Voice"          -> fuera los faxes y cualquier otro tipo
#
# Para añadir una regla, copia una línea. Para quitarla, bórrala.
# Estas columnas son obligatorias: si alguna no está en el archivo de
# origen, el script se detiene y dice cuál, en vez de generar el archivo
# sin un filtro que sí se había pedido.
COLUMNAS_CON_VALOR_FIJO = {
    "Direction": "Incoming",
    "Type": "Voice",
}

# Columnas que no deben venir vacías. Las filas que las tengan en blanco
# se ocultan. Equivale a quitar la marca de (Blanks) en el filtro de Excel.
# Para exigir lo mismo en otra columna, añádela a la lista.
# Para desactivar la regla, deja la lista vacía:  COLUMNAS_SIN_VACIOS = []
COLUMNAS_SIN_VACIOS = ["To"]

# Ocultar también las llamadas cuyo número no está en la tabla de campañas,
# es decir las filas en las que Queue Name queda con el valor de
# CAMPANA_NO_ENCONTRADA (0). Equivale a quitar la marca del 0 en el
# desplegable del filtro de Queue Name.
#
# El resumen del final dice qué números quedaron sin campaña, que es justo
# la lista de lo que se está ocultando aquí. Conviene revisarlo: si aparece
# un número nuevo de la operación, no desaparece del reporte sin avisar.
OCULTAR_SIN_CAMPANA = True

# Qué se hace con las filas que no cumplen las reglas de arriba:
#
#   "filtrar"  -> SE CONSERVAN en el archivo, pero quedan OCULTAS por el
#                 filtro de Excel. Quitando el filtro reaparecen.
#                 No se pierde ningún dato.
#
#   "eliminar" -> desaparecen del archivo por completo.
#
MODO_FILTRADO = "filtrar"

# Flechas de filtro en la fila de cabecera del archivo generado.
ACTIVAR_FILTRO = True

# Columnas que se eliminan por completo del archivo de salida.
# Esto sí es un borrado real: una columna no se puede "ocultar con filtro".
COLUMNAS_A_ELIMINAR = [
    "Extension",
    "Forwarded To",
    "Result Description",
    "Included",
    "Purchased",
]

# Nombres de las columnas clave en el archivo original.
COLUMNA_DIRECCION = "Direction"
COLUMNA_FECHA = "Date"
COLUMNA_HORA = "Time"
COLUMNA_DURACION = "Duration"


# ============================================================
#  FORMATO DE LAS CELDAS DEL ARCHIVO GENERADO
# ============================================================
#
#  Estos son EXACTAMENTE los mismos códigos que se escriben en Excel en
#  "Formato de celdas > Número > Personalizada". El script se limita a
#  dejarlos guardados en el archivo, de modo que al abrirlo la columna ya
#  aparece con ese formato aplicado, sin tener que tocar nada a mano.
#
#  Para que un formato surta efecto, la celda tiene que contener un valor
#  REAL de fecha o de hora, no un texto que lo parezca. De eso se encargan
#  limpiar_fecha() y limpiar_hora(): convierten lo que venga en el archivo
#  de origen a una fecha o una hora auténticas.
#
#  Ventaja de hacerlo así y no guardando texto: las columnas se pueden
#  ordenar y filtrar como fechas y horas de verdad, y además se ven igual
#  en cualquier equipo, porque el formato va escrito en el archivo y no
#  depende de la configuración regional de Windows.
#
#  Para cambiar un formato, escribe aquí el mismo código que usarías en
#  Excel. Ejemplos: "mm/dd/yy", "dd/mm/yyyy", "h:mm:ss AM/PM", "hh:mm".
#
FORMATOS_DE_COLUMNA = {
    "Date": "mm/dd/yyyy",
    "Time": "h:mm AM/PM",
    "Duration": "[h]:mm:ss",
}


# ============================================================
#  COLUMNAS CALCULADAS  (Day, Queue Name, 7, #Calls, Missed)
# ============================================================
#
#  Estas 5 columnas se añaden AL FINAL, después de las columnas que
#  sobreviven del archivo original. Se calculan en Python y se guardan como
#  VALORES: el archivo llega con el número ya puesto y no depende de que
#  Excel calcule nada al abrirlo.
#
#  Lo que eso implica, para que no sorprenda: si más adelante corriges a
#  mano una duración o un Action Result dentro del Excel, estas columnas NO
#  se actualizan solas. Lo que hay escrito es la foto del momento en que se
#  generó el archivo. Para rehacerlas, se vuelve a ejecutar el script.
#
#  A cambio, el archivo se puede leer desde cualquier herramienta (otro
#  script de Python, una carga a otro sistema) sin tener que abrirlo antes
#  en Excel, y abre más rápido.
#
#  Para no añadirlas y dejar el archivo como antes, pon esto en False:
CREAR_COLUMNAS_CALCULADAS = True

# Nombres que llevarán las columnas nuevas en la fila de cabecera.
ENCABEZADO_DIA = "Day"
ENCABEZADO_CAMPANA = "Queue Name"
ENCABEZADO_LLAMADAS = "#Calls"
ENCABEZADO_PERDIDAS = "Missed"

# Nombres de los días, de lunes a domingo, en ese orden.
# Se escriben desde aquí y no desde el idioma del equipo, así el archivo
# sale igual en cualquier ordenador. Para verlos en español, cambia la
# lista por:  ["lunes", "martes", "miércoles", "jueves", "viernes",
#              "sábado", "domingo"]
NOMBRES_DIAS = ["Monday", "Tuesday", "Wednesday", "Thursday",
                "Friday", "Saturday", "Sunday"]

# Columnas del archivo original de las que se alimentan los cálculos.
# Date y Duration ya estaban declaradas arriba; estas son las que faltaban.
COLUMNA_DESTINO = "To"              # el número al que entró la llamada
COLUMNA_RESULTADO = "Action Result"  # Accepted / Missed / Voicemail / ...

# Valor de Action Result que cuenta como llamada perdida.
# Ojo: "Voicemail" NO entra aquí. Si algún día se decide que el buzón
# también es una perdida, hay que cambiar el cálculo, no esta constante.
TEXTO_PERDIDA = "Missed"

# Minutos a partir de los cuales una llamada se considera larga.
# Cambia solo este número: el criterio y el título de la columna se ajustan
# los dos a la vez, así que no pueden acabar diciendo cosas distintas.
UMBRAL_MINUTOS = 7
ENCABEZADO_UMBRAL = UMBRAL_MINUTOS   # la columna se titula con el número

# Qué escribe la columna Queue Name cuando el número de To no está en la
# tabla de campañas. Se deja 0 porque es lo que hacía el archivo anterior.
# Si prefieres verlo a simple vista, pon:  CAMPANA_NO_ENCONTRADA = "SIN CAMPAÑA"
CAMPANA_NO_ENCONTRADA = 0

# Nombre de la hoja informativa donde se deja constancia de la tabla de
# campañas y del umbral con los que se generó el archivo.
HOJA_CAMPANAS = "Campañas"

# TABLA DE CAMPAÑAS:  ("nombre de la campaña", "número marcado")
#
# Es la equivalencia entre el número al que llama el cliente (columna To)
# y la campaña a la que pertenece. Para añadir una campaña nueva, copia una
# línea y cámbiala; para quitarla, borra la línea entera.
#
# El número se puede escribir como se quiera: "(801) 877-4906",
# "801-877-4906", "8018774906" o "+1 801 877 4906". El script lo convierte
# todo al formato con el que la centralita rellena la columna To,
# "(801) 877-4906", que es el que tienen que compartir para que la
# búsqueda encuentre la campaña.
CAMPANAS = [
    ("TAEKNO Mobile",           "(801) 877-4906"),
    ("Allitech WL1",            "(808) 900-8516"),
    ("Allitech WL2",            "(808) 900-8834"),
    ("Allitech WL3",            "(808) 303-2555"),
    ("Allitech WL4",            "(808) 303-2585"),
    ("ROYI FIBRA",              "(801) 770-0728"),
    ("ITS MOB 1 - Informativa", "(801) 877-1757"),
    ("ITS MOB 2 - Generica",    "(801) 877-3762"),
    ("ITS MOB 3 - Competencia", "(801) 877-3857"),
    ("HNET",                    "(801) 666-4972"),
    ("NEXOTL 1",                "(808) 470-8757"),
    ("NEXOTL 2",                "(808) 509-2808"),
    ("Flatexco",                "(801) 758-0077"),
    ("TMETRO",                  "(808) 400-0664"),
    ("FRONTIER",                "(808) 470-8035"),
    ("META Heri WL1",           "(808) 400-8155"),
    ("Heri Mob 1",              "(801) 335-9635"),
    ("Heri Mob 2",              "(808) 400-6074"),
    ("IA FCO Internet",         "(808) 808-1527"),
    #("HR SOL Internet",         "(801) 770-0527"),
    #("LMC INTERNET",            "8018100966"),
    #("FIBERNT",                 "8018773189"),
    #("ATG MK Internet",         "8019214527"),
    #("ATG MKT Wireless",        "8017700477"),
    #("HR SOLUTION 2",           "8016917005"),
    #("MECA Internet",           "8088081628"),
]


# ============================================================
#  IMPORTACIÓN DE LA LIBRERÍA EXTERNA
# ============================================================
# Se importa dentro de un try para poder avisar con un mensaje claro
# si openpyxl no está instalada, en lugar de soltar un error feo.
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("=" * 60)
    print("ERROR: falta la librería 'openpyxl', que es la que lee y")
    print("       escribe archivos Excel (.xlsx).")
    print()
    print("Instálala una sola vez con:   pip install openpyxl")
    print("(mira la nota del final de este script para más detalle)")
    print("=" * 60)
    sys.exit(1)


# ============================================================
#  FUNCIONES AUXILIARES
# ============================================================

class ErrorDeLectura(Exception):
    """Problema al abrir o leer el archivo de entrada."""


def carpeta_del_script():
    """Carpeta donde está guardado este archivo .py."""
    return Path(__file__).resolve().parent


def ruta_completa(carpeta, nombre_archivo):
    """
    Une carpeta + nombre y devuelve la ruta final, según estas reglas:

      - Si el NOMBRE ya trae una ruta completa (r"C:\\...\\archivo.xlsx"),
        manda esa ruta y la carpeta se ignora.
      - Si la CARPETA está vacía (""), se usa la carpeta de este script.
      - Si la CARPETA es una ruta completa, se usa tal cual.
      - Si la CARPETA es un nombre suelto ("datos"), se entiende como
        subcarpeta dentro de la carpeta de este script.
    """
    ruta_nombre = Path(str(nombre_archivo).strip())

    # El nombre ya incluye la ruta completa: tiene prioridad.
    if ruta_nombre.is_absolute():
        return ruta_nombre

    texto_carpeta = str(carpeta).strip()

    if texto_carpeta == "":
        base = carpeta_del_script()
    else:
        base = Path(texto_carpeta)
        if not base.is_absolute():
            # Carpeta relativa: se cuelga de la carpeta del script.
            base = carpeta_del_script() / base

    return base / ruta_nombre


def con_fecha_y_hora(ruta):
    """
    Añade la marca de tiempo al nombre, conservando carpeta y extensión:
      CallLog_limpio.xlsx  ->  CallLog_limpio_20260803-1945.xlsx
    """
    marca = datetime.now().strftime(FORMATO_FECHA_HORA)
    return ruta.with_name("{}_{}{}".format(ruta.stem, marca, ruta.suffix))


def ruta_sin_pisar(ruta):
    """
    Devuelve una ruta que no exista todavía, para no sobrescribir nunca.
    Si el nombre está ocupado, prueba con _2, _3, _4...
    (Solo hace falta si se ejecuta dos veces dentro del mismo minuto.)
    """
    if not ruta.exists():
        return ruta
    numero = 2
    while True:
        candidata = ruta.with_name("{}_{}{}".format(ruta.stem, numero, ruta.suffix))
        if not candidata.exists():
            return candidata
        numero += 1


def como_texto(valor):
    """Convierte cualquier celda a texto limpio ('' si está vacía)."""
    if valor is None:
        return ""
    return str(valor).strip()


def valor_de(fila, indice):
    """
    Devuelve el valor de una posición de la fila.
    Algunas filas vienen más cortas que la cabecera; en ese caso se
    devuelve None en vez de reventar con un IndexError.
    """
    if indice < len(fila):
        return fila[indice]
    return None


def fila_totalmente_vacia(fila):
    """True si todas las celdas de la fila están vacías."""
    return all(como_texto(celda) == "" for celda in fila)


# Detecta el nombre del día al principio del texto: "Thu ", "Thursday, ", "mon. "...
# Acepta las 7 abreviaturas (Mon, Tue, Wed, Thu, Fri, Sat, Sun) sin importar
# mayúsculas/minúsculas, con o sin punto y con o sin coma detrás.
PATRON_DIA_SEMANA = re.compile(
    r"^\s*(?:mon|tue|wed|thu|fri|sat|sun)[a-z]*\.?[\s,]+",
    re.IGNORECASE,
)

# Detecta una fecha ya en formato mm/dd/yyyy (admitiendo 7/5/2026 sin ceros).
PATRON_FECHA = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def limpiar_fecha(valor):
    """
    Devuelve una FECHA REAL de Excel, no un texto.

    Se quita el nombre del día y se interpreta lo que queda como mm/dd/yyyy.
    IMPORTANTE: no se reordena nada. El primer número es el mes y el segundo
    el día, tal como vienen, y así se vuelven a mostrar gracias al formato
    de celda mm/dd/yyyy que se aplica luego.

    Al ser una fecha auténtica, Excel permite ordenar y filtrar por fechas.
    Si el texto no se puede interpretar, se devuelve el texto ya sin el
    nombre del día, para no perder el dato.
    """
    if valor is None:
        return None

    # Caso 1: el archivo ya guardaba la celda como fecha real.
    # (datetime va antes que date porque datetime es un tipo de date)
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    # Caso 2 (el habitual): la celda es texto tipo "Thu 07/30/2026".
    texto_fecha = como_texto(valor)
    texto_fecha = PATRON_DIA_SEMANA.sub("", texto_fecha).strip()

    coincidencia = PATRON_FECHA.match(texto_fecha)
    if coincidencia:
        mes, dia, anio = coincidencia.groups()
        try:
            return date(int(anio), int(mes), int(dia))
        except ValueError:
            # Números que no forman una fecha real (por ejemplo 13/45/2026).
            # Se deja el texto para que se vea que ahí hay algo raro.
            return texto_fecha

    # Otro formato inesperado: se devuelve tal cual (ya sin el día).
    return texto_fecha


# Horas en formato 24 h: "17:04", "17:04:00"
PATRON_HORA_24H = re.compile(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$")

# Horas en formato 12 h: "5:04 PM", "5:04:00 p.m.", "05:04PM"
PATRON_HORA_12H = re.compile(
    r"^(\d{1,2}):(\d{2})(?::(\d{2}))?\s*([ap])\.?\s*m\.?$",
    re.IGNORECASE,
)


def limpiar_hora(valor):
    """
    Devuelve una HORA REAL de Excel, no un texto, para que se le pueda
    aplicar el formato h:mm AM/PM y se pueda ordenar por hora.

    Acepta las dos formas en que suele venir la columna Time:
    24 horas ("17:04:00") y 12 horas ("5:04 PM"). Si no reconoce el
    formato, devuelve el texto tal cual para no perder el dato.
    """
    if valor is None:
        return None

    # Casos en que el archivo ya traía una hora de verdad.
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, time):
        return valor
    if isinstance(valor, timedelta):
        # Excel entrega algunas horas como "tiempo transcurrido".
        total_segundos = int(valor.total_seconds()) % 86400
        return time(total_segundos // 3600, (total_segundos % 3600) // 60,
                    total_segundos % 60)

    texto_hora = como_texto(valor)

    # Formato de 12 horas: hay que pasar la hora a 0-23.
    coincidencia = PATRON_HORA_12H.match(texto_hora)
    if coincidencia:
        horas, minutos, segundos, meridiano = coincidencia.groups()
        horas = int(horas) % 12                 # 12 AM -> 0, 12 PM -> 12
        if meridiano.lower() == "p":
            horas += 12
        try:
            return time(horas, int(minutos), int(segundos or 0))
        except ValueError:
            return texto_hora

    # Formato de 24 horas.
    coincidencia = PATRON_HORA_24H.match(texto_hora)
    if coincidencia:
        horas, minutos, segundos = coincidencia.groups()
        try:
            return time(int(horas), int(minutos), int(segundos or 0))
        except ValueError:
            return texto_hora

    return texto_hora


# Duraciones tipo "0:00:44", "12:03", "1:02:03"  (h:mm:ss o mm:ss)
PATRON_DURACION = re.compile(r"^(\d{1,3}):(\d{1,2})(?::(\d{1,2}))?$")


def limpiar_duracion(valor):
    """
    Devuelve una DURACIÓN REAL de Excel (un time), no un texto, para que
    se le pueda aplicar el formato [h]:mm:ss y se pueda ordenar y sumar.

    Acepta lo que suele traer la columna Duration: "0:00:44" (h:mm:ss) y
    "12:03" (mm:ss). Si no reconoce el formato, devuelve el texto tal cual
    para no perder el dato.
    """
    if valor is None:
        return None

    # Casos en que el archivo ya traía una hora o duración de verdad.
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, time):
        return valor
    if isinstance(valor, timedelta):
        total_segundos = int(valor.total_seconds()) % 86400
        return time(total_segundos // 3600, (total_segundos % 3600) // 60,
                    total_segundos % 60)

    texto_duracion = como_texto(valor)

    coincidencia = PATRON_DURACION.match(texto_duracion)
    if coincidencia:
        primero, segundo, tercero = coincidencia.groups()
        try:
            if tercero is None:
                # Dos grupos = mm:ss (sin horas).
                horas, minutos, segundos = 0, int(primero), int(segundo)
            else:
                # Tres grupos = h:mm:ss.
                horas, minutos, segundos = int(primero), int(segundo), int(tercero)
            if minutos > 59 or segundos > 59 or horas > 23:
                return texto_duracion
            return time(horas, minutos, segundos)
        except ValueError:
            return texto_duracion

    return texto_duracion


# ============================================================
#  COLUMNAS CALCULADAS: TELÉFONOS Y CÁLCULO DE VALORES
# ============================================================

def normalizar_telefono(valor):
    """
    Deja cualquier número en el formato con el que viene la columna To:
    "(801) 877-4906".

        "8018774906"      -> "(801) 877-4906"
        "801-877-4906"    -> "(801) 877-4906"
        "+1 801 877 4906" -> "(801) 877-4906"

    Esto importa más de lo que parece. La campaña se busca comparando el
    número de la llamada con el de la tabla: si están escritos de formas
    distintas, no coinciden y la llamada se queda sin campaña. Normalizando
    los dos lados el problema desaparece.

    Si lo que llega no tiene 10 dígitos, se devuelve tal cual: puede ser un
    número corto, una extensión o algo que no es un teléfono, y es mejor
    dejarlo visible que inventarse un formato.
    """
    texto = como_texto(valor)
    digitos = re.sub(r"\D", "", texto)          # se queda solo con los dígitos

    # Números de EE. UU. escritos con el 1 delante: se quita.
    if len(digitos) == 11 and digitos.startswith("1"):
        digitos = digitos[1:]

    if len(digitos) != 10:
        return texto

    return "({}) {}-{}".format(digitos[:3], digitos[3:6], digitos[6:])


def nombre_del_dia(valor_fecha):
    """
    Devuelve el nombre del día de la semana de una fecha ya limpia.

    El nombre sale de la lista NOMBRES_DIAS, no de la configuración regional
    del equipo ni del idioma de Excel. Eso hace que el archivo se vea igual
    en cualquier ordenador: antes, con la fórmula TEXT(...,"dddd"), un Excel
    en español ponía "lunes" y uno en inglés "Monday", y cualquier tabla o
    fórmula que dependiera del texto dejaba de cuadrar al cambiar de equipo.

    Si la fecha no se pudo interpretar (quedó como texto), se devuelve
    cadena vacía en lugar de inventar un día.
    """
    if isinstance(valor_fecha, datetime):
        valor_fecha = valor_fecha.date()
    if not isinstance(valor_fecha, date):
        return ""
    return NOMBRES_DIAS[valor_fecha.weekday()]      # 0 = lunes


def segundos_de_duracion(valor_duracion):
    """
    Pasa a segundos una duración ya limpia (un time o un timedelta).

    Devuelve None si la duración no se pudo interpretar y quedó como texto.
    Quien llama tiene que tratar ese None como "no se sabe", y NO como cero
    ni como una duración larga: una duración ilegible no debe contarse como
    llamada que pasa del umbral.
    """
    if isinstance(valor_duracion, timedelta):
        return int(valor_duracion.total_seconds())
    if isinstance(valor_duracion, datetime):
        valor_duracion = valor_duracion.time()
    if isinstance(valor_duracion, time):
        return valor_duracion.hour * 3600 + valor_duracion.minute * 60 + valor_duracion.second
    return None


def columnas_calculadas_posibles(hay_fecha, hay_destino, hay_duracion, hay_resultado):
    """
    Devuelve los encabezados de las columnas calculadas que SÍ se pueden
    construir con las columnas que trae el archivo de origen, en el orden en
    que van a aparecer.

    Si falta la columna de la que se alimenta una de ellas, esa columna no
    se crea y el script avisa por pantalla, en vez de rellenarla con ceros
    que parecerían un dato real.
    """
    columnas = []
    saltadas = []

    if hay_fecha:
        columnas.append(ENCABEZADO_DIA)
    else:
        saltadas.append((ENCABEZADO_DIA, COLUMNA_FECHA))

    if hay_destino and CAMPANAS:
        columnas.append(ENCABEZADO_CAMPANA)
    elif not CAMPANAS:
        saltadas.append((ENCABEZADO_CAMPANA, "la lista CAMPANAS está vacía"))
    else:
        saltadas.append((ENCABEZADO_CAMPANA, COLUMNA_DESTINO))

    if hay_duracion:
        columnas.append(ENCABEZADO_UMBRAL)
    else:
        saltadas.append((str(ENCABEZADO_UMBRAL), COLUMNA_DURACION))

    if hay_resultado:
        columnas.append(ENCABEZADO_LLAMADAS)
        columnas.append(ENCABEZADO_PERDIDAS)
    else:
        saltadas.append((ENCABEZADO_LLAMADAS, COLUMNA_RESULTADO))
        saltadas.append((ENCABEZADO_PERDIDAS, COLUMNA_RESULTADO))

    for encabezado, motivo in saltadas:
        print("Aviso: no se pudo crear la columna '{}' porque falta '{}'.".format(
            encabezado, motivo))

    return columnas


def calcular_columnas(fecha_limpia, campana, duracion_limpia, texto_resultado):
    """
    Calcula de una vez el valor de las 5 columnas para UNA llamada y los
    devuelve en un diccionario {encabezado: valor}.

    Se escriben como valores, no como fórmulas: el archivo llega con el
    número ya puesto, sin depender de que Excel lo calcule al abrirlo.

    Las comparaciones de texto van en minúsculas para que se comporten como
    lo hacía la fórmula de Excel, que no distingue mayúsculas: "missed",
    "Missed" y "MISSED" cuentan igual.
    """
    segundos = segundos_de_duracion(duracion_limpia)

    return {
        ENCABEZADO_DIA: nombre_del_dia(fecha_limpia),

        # Campaña encontrada, o el valor de "no encontrado" (0 por defecto).
        ENCABEZADO_CAMPANA: campana if campana is not None else CAMPANA_NO_ENCONTRADA,

        # 1 solo si la duración se pudo leer Y pasa del umbral. Con una
        # duración ilegible (segundos = None) se pone 0, no 1.
        # La comparación es estricta: una llamada de exactamente 7:00 no cuenta.
        ENCABEZADO_UMBRAL: 1 if (segundos is not None
                                 and segundos > UMBRAL_MINUTOS * 60) else 0,

        # 1 por cada llamada que traiga Action Result.
        ENCABEZADO_LLAMADAS: 0 if texto_resultado == "" else 1,

        # 1 solo si el resultado es exactamente el que cuenta como perdida.
        # Ojo: "Voicemail" NO entra aquí.
        ENCABEZADO_PERDIDAS: 1 if texto_resultado.lower() == TEXTO_PERDIDA.lower() else 0,
    }


def escribir_hoja_campanas(libro):
    """
    Crea la hoja "Campañas" dentro del archivo generado.

    Ahora es una hoja INFORMATIVA: como las columnas se calculan en Python y
    se guardan como valores, cambiar algo aquí ya no recalcula nada. Está
    para dejar constancia de con qué tabla y con qué umbral se generó este
    archivo concreto, que es justo lo que hace falta cuando alguien pregunta
    de dónde salió un número tres semanas después.

    Para cambiar la equivalencia o el umbral de verdad hay que tocar
    CAMPANAS o UMBRAL_MINUTOS arriba en el script y volver a ejecutarlo.
    """
    hoja = libro.create_sheet(HOJA_CAMPANAS)

    hoja["B1"] = "Campaña"
    hoja["C1"] = "Numero"
    for numero_fila, (nombre_campana, telefono) in enumerate(CAMPANAS, start=2):
        hoja.cell(row=numero_fila, column=2, value=nombre_campana)
        celda_telefono = hoja.cell(row=numero_fila, column=3,
                                   value=normalizar_telefono(telefono))
        celda_telefono.number_format = "@"

    hoja["E1"] = "Umbral llamada larga (minutos)"
    hoja["F1"] = UMBRAL_MINUTOS
    hoja["E2"] = "Hoja informativa: cambiar estos datos NO recalcula el archivo."
    hoja["E3"] = "Para cambiarlos de verdad, edita el script y vuelve a ejecutarlo."

    hoja.column_dimensions["B"].width = 26
    hoja.column_dimensions["C"].width = 18
    hoja.column_dimensions["E"].width = 32


# ============================================================
#  LECTURA DE LA ENTRADA (.xlsx o .csv)
# ============================================================

# Codificaciones que se prueban al abrir un CSV, en este orden.
# 'utf-8-sig' cubre UTF-8 con y sin BOM, que es lo que suele exportar
# la centralita. 'latin-1' nunca falla, por eso va la última: es la red
# de seguridad para que el script no se caiga por un acento raro.
CODIFICACIONES_CSV = ["utf-8-sig", "cp1252", "latin-1"]

# Separadores posibles. Se elige el que más veces aparezca en la cabecera.
SEPARADORES_CSV = [",", ";", "\t", "|"]


def leer_csv(ruta):
    """
    Lee un CSV y devuelve (cabecera, filas, funcion_para_cerrar).
    Detecta sola la codificación y el separador, porque cada sistema
    exporta el CSV a su manera y el usuario no tiene por qué saberlo.
    """
    texto = None
    codificacion_usada = None
    for codificacion in CODIFICACIONES_CSV:
        try:
            with open(ruta, "r", encoding=codificacion, newline="") as archivo:
                texto = archivo.read()
            codificacion_usada = codificacion
            break
        except UnicodeDecodeError:
            continue

    if texto is None:
        raise ErrorDeLectura("no se pudo leer el CSV con ninguna codificación conocida.")

    lineas = texto.splitlines()
    if not lineas:
        raise ErrorDeLectura("el CSV está vacío.")

    # El separador es el carácter que más veces aparece en la cabecera.
    separador = max(SEPARADORES_CSV, key=lambda s: lineas[0].count(s))
    if lineas[0].count(separador) == 0:
        separador = ","

    nombre_separador = {",": "coma", ";": "punto y coma", "\t": "tabulador", "|": "barra"}
    print("CSV detectado: codificación {}, separador {}".format(
        codificacion_usada, nombre_separador.get(separador, separador)))

    filas = list(csv.reader(io.StringIO(texto), delimiter=separador))
    if not filas:
        raise ErrorDeLectura("el CSV no tiene ninguna fila.")

    return filas[0], filas[1:], lambda: None


def leer_excel(ruta, nombre_hoja):
    """
    Lee un .xlsx y devuelve (cabecera, filas, funcion_para_cerrar).
      data_only=True -> si hay fórmulas, se lee el resultado, no la fórmula.
      read_only=True -> lectura rápida y con poca memoria.
    """
    libro = load_workbook(filename=ruta, data_only=True, read_only=True)

    if nombre_hoja:
        if nombre_hoja not in libro.sheetnames:
            disponibles = ", ".join(libro.sheetnames)
            libro.close()
            raise ErrorDeLectura(
                "no existe la hoja '{}'. Hojas disponibles: {}".format(nombre_hoja, disponibles))
        hoja = libro[nombre_hoja]
    else:
        hoja = libro[libro.sheetnames[0]]

    filas = hoja.iter_rows(values_only=True)
    try:
        cabecera = next(filas)
    except StopIteration:
        libro.close()
        raise ErrorDeLectura("la hoja '{}' está vacía.".format(hoja.title))

    return cabecera, filas, libro.close


def abrir_entrada(ruta, nombre_hoja):
    """Elige el lector según la extensión del archivo."""
    extension = ruta.suffix.lower()
    if extension == ".csv":
        return leer_csv(ruta)
    if extension in (".xlsx", ".xlsm"):
        return leer_excel(ruta, nombre_hoja)
    raise ErrorDeLectura(
        "extensión '{}' no admitida. Usa un archivo .xlsx o .csv.".format(extension))


# ============================================================
#  PROCESO PRINCIPAL
# ============================================================

def main():
    ruta_entrada = ruta_completa(CARPETA_ENTRADA, ARCHIVO_ENTRADA)
    ruta_salida = ruta_completa(CARPETA_SALIDA, ARCHIVO_SALIDA)

    # --- Comprobación 1: que el archivo de entrada exista ---
    if not ruta_entrada.exists():
        print("=" * 60)
        print("ERROR: no se encontró el archivo de entrada.")
        print()
        print("  Se buscó aquí: {}".format(ruta_entrada))
        print()
        if not ruta_entrada.parent.exists():
            # El fallo está en la carpeta, no en el nombre del archivo.
            print("  La carpeta ni siquiera existe: {}".format(ruta_entrada.parent))
            print("  Revisa CARPETA_ENTRADA arriba en la configuración.")
        else:
            print("  La carpeta sí existe, pero no hay ningún archivo con ese nombre.")
            print("  Revisa que ARCHIVO_ENTRADA sea exacto, con su extensión.")
            # Se listan los archivos válidos que sí hay, por si es una errata.
            candidatos = sorted(p.name for p in ruta_entrada.parent.iterdir()
                                if p.suffix.lower() in (".xlsx", ".xlsm", ".csv"))
            if candidatos:
                print("  Archivos que sí puede leer el script en esa carpeta:")
                for nombre in candidatos[:10]:
                    print("     - {}".format(nombre))
        print("=" * 60)
        return

    # --- Comprobación 2: la salida siempre es .xlsx ---
    if ruta_salida.suffix.lower() != ".xlsx":
        print("Aviso: la salida siempre se guarda en .xlsx; se corrige la extensión.")
        ruta_salida = ruta_salida.with_suffix(".xlsx")

    # --- Nombre del archivo de salida con fecha y hora ---
    if NOMBRE_CON_FECHA_HORA:
        ruta_salida = con_fecha_y_hora(ruta_salida)
    # Nunca se sobrescribe nada, ni siquiera dos ejecuciones en el mismo minuto.
    ruta_salida = ruta_sin_pisar(ruta_salida)

    # --- Comprobación 3: que exista la carpeta donde se va a guardar ---
    if not ruta_salida.parent.exists():
        if CREAR_CARPETA_SALIDA:
            ruta_salida.parent.mkdir(parents=True, exist_ok=True)
            print("Aviso: se creó la carpeta de salida: {}".format(ruta_salida.parent))
        else:
            print("ERROR: la carpeta de salida no existe: {}".format(ruta_salida.parent))
            print("       Créala a mano, corrige CARPETA_SALIDA,")
            print("       o pon CREAR_CARPETA_SALIDA = True para que se cree sola.")
            return

    print("Leyendo:  {}".format(ruta_entrada))

    # --- Apertura de la entrada (.xlsx o .csv) ---
    try:
        cabecera_original, filas, cerrar_entrada = abrir_entrada(ruta_entrada, NOMBRE_HOJA)
    except ErrorDeLectura as problema:
        print("ERROR: {}".format(problema))
        return

    encabezados = [como_texto(celda) for celda in cabecera_original]
    # Los archivos a veces arrastran columnas fantasma al final: se descartan.
    while encabezados and encabezados[-1] == "":
        encabezados.pop()

    # Diccionario nombre-en-minúsculas -> posición, para localizar columnas
    # aunque el archivo venga con distinta capitalización.
    posiciones = {nombre.lower(): i for i, nombre in enumerate(encabezados)}

    # --- Comprobación 4: que existan las columnas que necesitamos ---
    # Obligatorias: la fecha y todas las que llevan una regla de valor fijo
    # (Direction, Type...). Si falta alguna, el script se detiene: es mejor
    # eso que entregar un archivo al que le falta un filtro sin avisar.
    obligatorias = [COLUMNA_FECHA] + list(COLUMNAS_CON_VALOR_FIJO)
    faltantes = [c for c in obligatorias if c.lower() not in posiciones]
    if faltantes:
        print("ERROR: faltan columnas obligatorias: {}".format(", ".join(faltantes)))
        print("       Columnas encontradas: {}".format(", ".join(encabezados)))
        cerrar_entrada()
        return

    indice_fecha = posiciones[COLUMNA_FECHA.lower()]
    # La columna de la hora es opcional: si no está, el script sigue igual.
    indice_hora = posiciones.get(COLUMNA_HORA.lower())
    # La columna de la duración también es opcional.
    indice_duracion = posiciones.get(COLUMNA_DURACION.lower())
    # Columnas que alimentan las fórmulas calculadas (también opcionales:
    # si falta alguna, esa columna no se crea y el script avisa).
    indice_destino = posiciones.get(COLUMNA_DESTINO.lower())
    indice_resultado = posiciones.get(COLUMNA_RESULTADO.lower())

    # --- Columnas en las que solo se deja visible un valor ---
    # Cada regla es (nombre de la columna, dónde está, valor que se conserva).
    reglas_valor_fijo = [
        (nombre_columna, posiciones[nombre_columna.lower()], valor_esperado)
        for nombre_columna, valor_esperado in COLUMNAS_CON_VALOR_FIJO.items()
    ]

    # --- Columnas que no pueden venir vacías ---
    indices_sin_vacios = []
    for nombre_columna in COLUMNAS_SIN_VACIOS:
        if nombre_columna.lower() in posiciones:
            indices_sin_vacios.append((nombre_columna, posiciones[nombre_columna.lower()]))
        else:
            print("Aviso: la columna '{}' no existe, no se puede exigir que tenga dato.".format(
                nombre_columna))

    # --- Qué columnas se conservan (todas menos las que se eliminan) ---
    eliminar = {nombre.lower() for nombre in COLUMNAS_A_ELIMINAR}
    indices_conservados = [i for i, nombre in enumerate(encabezados) if nombre.lower() not in eliminar]
    encabezados_finales = [encabezados[i] for i in indices_conservados]

    no_encontradas = [c for c in COLUMNAS_A_ELIMINAR if c.lower() not in posiciones]
    if no_encontradas:
        print("Aviso: estas columnas a eliminar no estaban en el archivo: {}".format(
            ", ".join(no_encontradas)))

    # --- Libro de salida ---
    libro_salida = Workbook()
    hoja_salida = libro_salida.active
    hoja_salida.title = "CallLog"

    # --- Columnas calculadas: hoja informativa + qué columnas se pueden crear ---
    # Se decide ANTES de escribir la cabecera, porque sus nombres van en esa
    # misma fila 1, a continuación de las columnas del archivo original.
    columnas_calculadas = []
    if CREAR_COLUMNAS_CALCULADAS:
        escribir_hoja_campanas(libro_salida)
        columnas_calculadas = columnas_calculadas_posibles(
            hay_fecha=indice_fecha in indices_conservados,
            hay_destino=indice_destino is not None,
            hay_duracion=indice_duracion is not None,
            hay_resultado=indice_resultado is not None,
        )

    # Cabecera completa: columnas que sobreviven + columnas calculadas.
    hoja_salida.append(encabezados_finales + columnas_calculadas)

    # Tabla de campañas resuelta en Python, con los números ya normalizados.
    # La usan dos cosas: el aviso del resumen y, si OCULTAR_SIN_CAMPANA está
    # activo, la decisión de qué filas ocultar. Es la misma equivalencia que
    # consulta la fórmula de Queue Name, así que las dos coinciden siempre.
    campana_por_numero = {normalizar_telefono(telefono): nombre
                          for nombre, telefono in CAMPANAS}
    destinos_sin_campana = {}
    # ¿Se puede aplicar la regla del 0? Solo si hay columna calculada de
    # campaña, hay tabla y existe la columna To de la que sale el número.
    filtrar_campana = (OCULTAR_SIN_CAMPANA
                       and campana_por_numero
                       and indice_destino is not None
                       and ENCABEZADO_CAMPANA in columnas_calculadas)

    # Contadores y datos para construir el filtro.
    total_leidas = 0
    escritas = 0
    filas_a_ocultar = []        # números de fila del archivo de salida
    ocultas_por = {}            # {motivo: cuántas filas}
    eliminadas_por = {}         # solo en modo "eliminar"
    valores_visibles = {}       # {columna: valores que el filtro debe mostrar}

    # --- Recorrido de las filas de datos ---
    for fila in filas:
        # Filas completamente en blanco: se saltan sin contarlas.
        if fila_totalmente_vacia(fila):
            continue

        total_leidas += 1

        # Motivos por los que esta fila NO debe verse. Si la lista queda
        # vacía, la fila es visible.
        motivos = []

        # Regla 1: las columnas con valor fijo deben traer ese valor
        # (Direction = Incoming, Type = Voice...).
        for nombre_columna, indice_columna, valor_esperado in reglas_valor_fijo:
            valor_celda = como_texto(valor_de(fila, indice_columna))
            if valor_celda.lower() == valor_esperado.lower():
                # Se guarda el valor tal cual viene, para que el filtro de Excel
                # reconozca también variantes como "incoming" o "Incoming ".
                valores_visibles.setdefault(nombre_columna, set()).add(valor_celda)
            else:
                motivos.append("{} distinto de {}".format(nombre_columna, valor_esperado))

        # Regla 2: las columnas obligatorias no pueden estar vacías.
        for nombre_columna, indice_columna in indices_sin_vacios:
            texto_celda = como_texto(valor_de(fila, indice_columna))
            if texto_celda == "":
                motivos.append("{} vacío".format(nombre_columna))
            else:
                valores_visibles.setdefault(nombre_columna, set()).add(texto_celda)

        # Regla 3: el número marcado tiene que estar en la tabla de campañas.
        # La campaña se resuelve aquí una sola vez y se reutiliza después
        # para escribir el valor de la columna Queue Name.
        campana = None
        if indice_destino is not None and campana_por_numero:
            destino = normalizar_telefono(valor_de(fila, indice_destino))
            campana = campana_por_numero.get(destino)
            if campana is not None:
                valores_visibles.setdefault(ENCABEZADO_CAMPANA, set()).add(campana)
            else:
                # Para el aviso del resumen solo interesan las filas que
                # habrían sido visibles: las salientes o sin To ya se ocultan
                # por otro motivo y no aportan nada a esa lista.
                if not motivos and destino:
                    destinos_sin_campana[destino] = destinos_sin_campana.get(destino, 0) + 1
                if filtrar_campana:
                    motivos.append("{} sin campaña ({})".format(
                        ENCABEZADO_CAMPANA, CAMPANA_NO_ENCONTRADA))

        # En modo "eliminar" la fila se descarta aquí y no llega al archivo.
        if motivos and MODO_FILTRADO == "eliminar":
            eliminadas_por[motivos[0]] = eliminadas_por.get(motivos[0], 0) + 1
            continue

        # Fila que va a ocupar esta llamada dentro del archivo generado.
        # Se lleva con un contador propio y NO se pregunta a openpyxl con
        # hoja_salida.max_row: ese dato no está guardado, openpyxl recorre
        # todas las celdas del libro cada vez que se le pide. Preguntándolo
        # una vez por fila, el tiempo crece al cuadrado y con un archivo
        # grande el script parece colgado. Con el contador es instantáneo.
        numero_fila = escritas + 2          # +2 porque la fila 1 es la cabecera

        # Se copian solo las columnas que sobreviven, convirtiendo la fecha
        # y la hora en valores reales al pasar por sus columnas. De paso se
        # guardan la fecha y la duración ya limpias, que hacen falta para
        # calcular las columnas del final.
        fecha_limpia = None
        duracion_limpia = None
        fila_nueva = []
        for i in indices_conservados:
            valor = valor_de(fila, i)
            if i == indice_fecha:
                valor = limpiar_fecha(valor)
                fecha_limpia = valor
            elif i == indice_hora:
                valor = limpiar_hora(valor)
            elif i == indice_duracion:
                valor = limpiar_duracion(valor)
                duracion_limpia = valor
            fila_nueva.append(valor)

        # --- Valores de las columnas calculadas para ESTA fila ---
        # Se calculan en Python y se añaden al final de la misma fila. El
        # archivo queda con el número escrito, no con una fórmula.
        if columnas_calculadas:
            texto_resultado = ("" if indice_resultado is None
                               else como_texto(valor_de(fila, indice_resultado)))
            calculados = calcular_columnas(fecha_limpia, campana,
                                           duracion_limpia, texto_resultado)
            for encabezado in columnas_calculadas:
                fila_nueva.append(calculados[encabezado])

        hoja_salida.append(fila_nueva)
        escritas += 1

        # Señal de vida cada 2.000 filas, para distinguir "va lento" de
        # "se colgó" sin tener que adivinar mirando una pantalla quieta.
        if total_leidas % 2000 == 0:
            print("   ... {} filas leídas".format(total_leidas))

        # En modo "filtrar" la fila entra en el archivo pero se apunta
        # para ocultarla. Se cuenta por su primer motivo, para no sumarla
        # dos veces si incumple varias reglas a la vez.
        if motivos:
            filas_a_ocultar.append(numero_fila)
            ocultas_por[motivos[0]] = ocultas_por.get(motivos[0], 0) + 1

    cerrar_entrada()

    # --- Formato de celda de Excel para Date y Time ---
    # Es lo mismo que seleccionar la columna en Excel y elegir
    # "Formato de celdas > Personalizada", pero hecho de una vez y guardado
    # dentro del archivo, así que al abrirlo ya está aplicado.
    for nombre_columna, formato in FORMATOS_DE_COLUMNA.items():
        indice_columna = posiciones.get(nombre_columna.lower())
        if indice_columna is None:
            print("Aviso: no existe la columna '{}', no se le aplica formato.".format(
                nombre_columna))
            continue
        if indice_columna not in indices_conservados:
            continue  # es una de las columnas eliminadas

        numero_columna = indices_conservados.index(indice_columna) + 1  # 1 = columna A
        for celdas in hoja_salida.iter_rows(min_row=2,
                                            min_col=numero_columna,
                                            max_col=numero_columna):
            for celda in celdas:
                celda.number_format = formato

    # --- Ancho de columnas, solo para que se lea cómodo al abrirlo ---
    # str() porque el encabezado del umbral es un número (7), no un texto.
    todos_los_encabezados = encabezados_finales + columnas_calculadas
    for numero_columna, nombre in enumerate(todos_los_encabezados, start=1):
        ancho = min(32, max(12, len(str(nombre)) + 4))
        hoja_salida.column_dimensions[get_column_letter(numero_columna)].width = ancho

    # --- Filtro de Excel ---
    # Se escribe en el archivo lo mismo que guardaría Excel al filtrar a
    # mano: el rango de la tabla, qué valores debe mostrar cada columna
    # filtrada, y las filas ocultas.
    if ACTIVAR_FILTRO and todos_los_encabezados:
        # El filtro abarca también las columnas calculadas, para poder
        # filtrar por día o por campaña directamente desde la cabecera.
        ultima_columna = get_column_letter(len(todos_los_encabezados))
        hoja_salida.auto_filter.ref = "A1:{}{}".format(ultima_columna, escritas + 1)

        if MODO_FILTRADO == "filtrar":
            # Columnas del archivo original que llevan filtro. Su posición
            # dentro del rango se calcula con indices_conservados, porque
            # por el camino se borraron columnas.
            columnas_filtradas = [(nombre, indice) for nombre, indice, _ in reglas_valor_fijo]
            columnas_filtradas += indices_sin_vacios
            for nombre_columna, indice_columna in columnas_filtradas:
                if indice_columna not in indices_conservados:
                    continue
                # Posición de la columna dentro del rango del filtro (empieza en 0).
                columna_en_filtro = indices_conservados.index(indice_columna)
                valores = sorted(valores_visibles.get(nombre_columna, set()))
                if valores:
                    hoja_salida.auto_filter.add_filter_column(
                        columna_en_filtro, valores, blank=False)

            # Queue Name no viene del archivo original: es una columna
            # calculada, y va después de todas las demás. Se marcan como
            # visibles las campañas encontradas, con lo que el 0 queda
            # desmarcado en el desplegable, que es justo lo que se busca.
            if filtrar_campana:
                columna_en_filtro = (len(encabezados_finales)
                                     + columnas_calculadas.index(ENCABEZADO_CAMPANA))
                campanas_visibles = sorted(valores_visibles.get(ENCABEZADO_CAMPANA, set()))
                if campanas_visibles:
                    hoja_salida.auto_filter.add_filter_column(
                        columna_en_filtro, campanas_visibles, blank=False)

            # Ocultar las filas afectadas, que es lo que hace Excel al filtrar.
            for numero_fila in filas_a_ocultar:
                hoja_salida.row_dimensions[numero_fila].hidden = True

    # --- Guardado ---
    try:
        libro_salida.save(ruta_salida)
    except PermissionError:
        print("ERROR: no se pudo guardar '{}'.".format(ruta_salida.name))
        print("       Suele pasar cuando el archivo está abierto en Excel.")
        print("       Ciérralo y vuelve a ejecutar el script.")
        return

    # --- Resumen ---
    visibles = escritas - len(filas_a_ocultar)

    print()
    print("=" * 60)
    print("LIMPIEZA TERMINADA")
    print("=" * 60)
    print("Filas leídas (sin contar la cabecera): {}".format(total_leidas))
    print("Filas en el archivo generado:          {}".format(escritas))
    print("  - Visibles al abrir:                 {}".format(visibles))
    print("  - Ocultas por el filtro:             {}".format(len(filas_a_ocultar)))
    for motivo, cuantas in ocultas_por.items():
        print("      · {}: {}".format(motivo, cuantas))
    if filas_a_ocultar:
        print()
        print("  Las filas ocultas SIGUEN en el archivo.")
        print("  Quita el filtro en Excel para volver a verlas.")
    if eliminadas_por:
        print()
        print("Filas eliminadas (modo 'eliminar'):")
        for motivo, cuantas in eliminadas_por.items():
            print("      · {}: {}".format(motivo, cuantas))
    print()
    print("Columnas eliminadas: {}".format(", ".join(COLUMNAS_A_ELIMINAR)))
    print("Columnas en el resultado ({}): {}".format(
        len(encabezados_finales), ", ".join(encabezados_finales)))
    print()
    print("Formato de celda aplicado:")
    for nombre_columna, formato in FORMATOS_DE_COLUMNA.items():
        print("      · {}: {}".format(nombre_columna, formato))

    if columnas_calculadas:
        print()
        print("Columnas calculadas añadidas ({}): {}".format(
            len(columnas_calculadas),
            ", ".join(str(nombre) for nombre in columnas_calculadas)))
        print("      Son valores, no fórmulas: el archivo llega calculado.")
        print("      Umbral de llamada larga: más de {} minutos".format(UMBRAL_MINUTOS))
        print("      Tabla de campañas: {} números (copia informativa en la hoja '{}')".format(
            len(CAMPANAS), HOJA_CAMPANAS))

    if destinos_sin_campana:
        total_sin_campana = sum(destinos_sin_campana.values())
        print()
        if filtrar_campana:
            print("AVISO: {} llamadas fueron a números que NO están en la tabla".format(
                total_sin_campana))
            print("       de campañas, y quedan OCULTAS por el filtro de Queue Name.")
        else:
            print("AVISO: {} llamadas fueron a números que NO están en la tabla".format(
                total_sin_campana))
            print("       de campañas. En Queue Name saldrá '{}'.".format(
                CAMPANA_NO_ENCONTRADA))
        for destino, cuantas in sorted(destinos_sin_campana.items(),
                                       key=lambda par: -par[1])[:10]:
            print("      · {}: {} llamadas".format(destino, cuantas))
        print("       Si alguno debería contar, añádelo a la lista CAMPANAS.")

    print()
    print("Archivo original (intacto): {}".format(ruta_entrada))
    print("Archivo generado (nuevo):   {}".format(ruta_salida))
    print("=" * 60)


# Punto de entrada: esto es lo que se ejecuta al pulsar Run / F5 en VS Code.
if __name__ == "__main__":
    main()


# ============================================================================
#  LIBRERÍAS EXTERNAS QUE HAY QUE INSTALAR  (NO instaladas por este script)
# ============================================================================
#
#  Solo hace falta UNA:
#
#      openpyxl
#
#  Comando de instalación:
#
#      pip install openpyxl
#
#  Si pip no funciona directamente, esta variante suele resolverlo porque
#  usa el mismo Python que ejecuta VS Code:
#
#      python -m pip install openpyxl
#
#  Notas para equipos con CMD/terminal bloqueado:
#   - El script en sí NO necesita terminal: se ejecuta con Run / F5.
#     La instalación de openpyxl es un paso previo que se hace UNA SOLA VEZ.
#     También puedes ejecutar instalar_librerias.py con F5, que hace lo mismo.
#   - Si no puedes ejecutar pip, pide a IT/soporte que instalen openpyxl
#     en el intérprete de Python que usa VS Code
#     (se ve abajo a la derecha en la barra de estado de VS Code).
#   - Todo lo demás que usa el script (csv, io, re, sys, datetime, pathlib)
#     viene incluido en Python, no hay que instalar nada más.
#     La lectura de CSV usa el módulo 'csv', que ya trae Python: admitir
#     CSV no añadió ninguna dependencia nueva.
#   - NO se usa pandas a propósito: es una instalación mucho más pesada y
#     openpyxl basta para leer y escribir .xlsx.
# ============================================================================
