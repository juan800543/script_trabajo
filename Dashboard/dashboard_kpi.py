# -*- coding: utf-8 -*-
"""
Dashboard Marketing KPI
-----------------------
Lee el Excel mensual de KPIs de ventas y genera un dashboard HTML autocontenido.
No modifica el Excel de entrada bajo ninguna circunstancia (solo lectura).

Ejecutar con el boton Run de VS Code. No requiere argumentos ni interaccion.
"""

import base64
import glob
import json
import math
import os
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

# =============================================================================
# CONFIG - unico lugar con rutas y mapeos. Nada de rutas hardcodeadas abajo.
# =============================================================================

# Todas las rutas cuelgan de la carpeta donde esta guardado este archivo .py,
# para que la carpeta se pueda copiar a cualquier PC o unidad y siga
# funcionando sin editar nada.
CARPETA_BASE = os.path.dirname(os.path.abspath(__file__))

CARPETA_ENTRADA = os.path.join(CARPETA_BASE, "Entrada")
CARPETA_SALIDA = CARPETA_BASE
NOMBRE_SALIDA = "dashboard.html"

# Marca del tablero: sale en la cabecera, en el titulo de la pestana, en el
# subtitulo del corte y en el nombre del archivo que descarga el boton Guardar.
NOMBRE_MARCA = "Marketing KPI"

# Logo de la cabecera. Si en esta misma carpeta hay un archivo con alguno de
# estos nombres, se incrusta ese (en base64, para que el HTML siga siendo un
# unico archivo autocontenido). Si no hay ninguno, se dibuja el emblema
# vectorial de reserva que va mas abajo en _svg_marca_reserva().
ARCHIVOS_LOGO = ("logo.svg", "logo.png", "logo.webp", "logo.jpg", "logo.jpeg")

# Para fijar una ruta concreta (por ejemplo publicar en una carpeta de red),
# reemplaza la linea de arriba por una ruta absoluta con el prefijo r"...":
#     CARPETA_SALIDA = r"\\servidor\carpeta\kpi"

# Abrir el dashboard en el navegador al terminar. Ponlo en False si lo ejecutas
# de forma desatendida y no quieres que se abra una ventana cada vez.
ABRIR_AL_TERMINAR = True

# Campana en Print -> lista de colas (Queue Name) que le corresponden.
#
#   - Los nombres de cola CAMBIAN de un mes a otro (en agosto "Heri Mob 1" y
#     "Flatexco"; en julio "HERIMOBILE" y "FLATEXCO"). Por eso cada campana
#     admite VARIOS nombres: se ponen todos y el script usa el que encuentre.
#   - Una campana puede recibir llamadas de mas de una cola a la vez
#     (NEXTL2 = "NEXOTL 2" + "NEXOTL 1").
#   - La comparacion ignora mayusculas, acentos y espacios de sobra, asi que
#     "Flatexco" y "FLATEXCO" son la misma cola y basta con escribir una.
#   - Lista vacia = la campana no tiene llamadas en el raw (0 llamadas).
#
# Si aparece una campana nueva en Print, agregarla aqui con sus colas reales.
MAPEO_COLAS = {
    "TAEKNO Mobile A": ["TAEKNO Mobile"],
    "TMETRO": [],
    "ALLI WL1": ["Allitech WL1"],
    "ALLI WL2": ["Allitech WL2"],
    "ALLI WL3": ["Allitech WL3"],
    "ALLI WL4": ["Allitech WL4"],
    "ROYI FIBRA": ["ROYI FIBRA"],
    "ITSFIBERNET": [],
    "R-FRONTIER": [],
    "Heri Mob 1": ["Heri Mob 1", "HERIMOBILE"],
    "Heri Mob 2": ["Heri Mob 2"],
    "NEXTL2": ["NEXOTL 2", "NEXOTL 1"],
    "FLATEXCO": ["Flatexco"],
    "IA FCO Internet": ["IA FCO Internet"],
}

# Colas del raw que existen pero NO son campanas de llamadas de Tabla 1
# (van a leads o son ruido). Se listan para dejar constancia; el script las
# ignora automaticamente por no estar en MAPEO_COLAS.
COLAS_IGNORADAS_CONOCIDAS = {
    "META Heri WL1", "HNET", "HR SOL", "HR SOL Internet",
    "ITS MOB 1", "ITS MOB 2", "ITS MOB 3",
    "ITS MOB 2 - Generica", "ITS MOB 3 - Competencia",
}

HOJA_PRINT = "Print"
HOJA_RAW_LLAMADAS = "RawDataRingCentral"
HOJA_RAW_VENTAS = "Raw Data General"
HOJA_COSTOS = "Valores$$$"

# Ademas de la hoja principal puede haber hojas de llamadas auxiliares, con
# las de algunas campanas sueltas. NO se listan por nombre porque cambia segun
# el mes: en mayo 2026 se llamaba 'RawDataRingCentral HM2' (solo Heri Mob 2) y
# desde junio 'RawDataRingCentral HM2-IA', al incorporarse IA FCO Internet.
# Se detectan solas: cualquier hoja cuyo nombre empiece por el de la principal.
#
# Son hojas tramposas, porque su papel cambia de un mes a otro:
#   - Agosto 2026: 'Heri Mob 2' e 'IA FCO Internet' estaban en la hoja
#     principal y esta traia OTRO conjunto de datos que 'Print' no incluye.
#     Usarla habria inflado los totales cerca del 50 %.
#   - Julio y mayo 2026: la hoja principal no trae esas campanas y esta es la
#     unica fuente que existe para ellas.
# Por eso no se pueden ni usar siempre ni ignorar siempre. La regla que
# funciona en todos los casos: la hoja principal manda, y en las auxiliares
# solo se buscan las campanas que no aparecieron alli. La validacion final (la
# suma diaria tiene que cuadrar con 'Print') confirma que la eleccion fue la
# correcta.

# Cifras de control de archivos ya revisados a mano, para detectar si la
# lectura se rompe. La clave es (mes, fecha de corte), NO solo el mes: el
# mismo mes se actualiza a diario y cada corte trae cifras distintas. Con el
# mes como unica clave, cualquier corte posterior al anotado fallaba aunque
# los datos estuvieran perfectos.
#
# Rellenarlas es OPCIONAL. Un corte que no este en esta tabla se salta estas
# comprobaciones y se queda con las universales, que son las que de verdad
# protegen (que no falten llamadas, que todas las campanas esten mapeadas).
CIFRAS_DE_CONTROL_POR_CORTE = {
    (date(2026, 8, 1), date(2026, 8, 11)): {
        "llamadas_totales": 7876,
        "llamadas_perdidas": 2054,
        "ventas_cliente": 583,
        "ventas_linea": 770,
        "monto_llamadas": 2292.51,
        "pct_perdidas": 0.2068479355488419,
        "eficiencia_cliente": 0.07402234636871509,
        "eficiencia_linea": 0.09776536312849161,
        "llamadas_dia": 875.1111111111111,
        "leads_meta_heri": 56,
        "total_general": 2438.6800000000003,
    },
    (date(2026, 8, 1), date(2026, 8, 12)): {
        "llamadas_totales": 9124,
        "llamadas_perdidas": 2371,
        "ventas_cliente": 487,
        "ventas_linea": 617,
        "monto_llamadas": 2292.51,
        "pct_perdidas": 0.20626359286646367,
        "eficiencia_cliente": 0.0533757124068391,
        "eficiencia_linea": 0.06762384918895221,
        "llamadas_dia": 912.4,
        "leads_meta_heri": 56,
        "total_general": 2438.68,
    },
    (date(2026, 7, 1), date(2026, 7, 31)): {
        "llamadas_totales": 23494,
        "llamadas_perdidas": 4974,
        "ventas_cliente": 1547,
        "ventas_linea": 1984,
        "monto_llamadas": 192133.61,
        "pct_perdidas": 0.17472249543346915,
        "eficiencia_cliente": 0.0658465991316932,
        "eficiencia_linea": 0.08444709287477654,
        "llamadas_dia": 870.1481481481482,
        "leads_meta_heri": 1318,
        "total_general": 199086.51,
    },
}

# Campanas que se sabe que no tienen bloque propio en 'Raw Data General'.
# Estar en esta lista NO cambia el calculo: solo evita que salte el aviso,
# para no repetir cada dia algo ya conocido y asumido.
#
# Vacia a proposito. En el archivo de agosto con corte 11/08 hacia falta para
# ALLI WL2 y ALLI WL4, cuyas ventas eran un calco de ALLI WL3 y Heri Mob 1.
# El archivo corregido (corte 12/08) ya trae datos propios para las dos, asi
# que dejarlas aqui silenciaria el aviso si el problema volviera a aparecer.
CAMPANAS_SIN_SERIE_DIARIA_CONOCIDAS = set()

# -----------------------------------------------------------------------------
# Configuracion editable sin tocar el codigo (configuracion.json)
# -----------------------------------------------------------------------------
# Archivo OPCIONAL que vive al lado de este .py. Lo escribe la seccion
# "Configuracion" del propio dashboard, asi que nadie tiene que editar Python
# para renombrar una campana o para arreglar una cola que cambio de nombre.
#
#   {
#     "nombres_campanas": {"ALLI WL1": "Allitech Wireless 1"},
#     "mapeo_colas":      {"Heri Mob 1": ["Heri Mob 1", "HERIMOBILE"]},
#     "mostrar_notas_en_dashboard": false
#   }
#
#   - nombres_campanas: SOLO cambia como se ve el nombre en el tablero. La
#     lectura del Excel sigue usando el nombre original de 'Print', asi que
#     renombrar no puede descuadrar ninguna cifra.
#   - mapeo_colas: se fusiona sobre MAPEO_COLAS, campana por campana. Es lo que
#     hace falta cuando en el Excel cambia el nombre de una cola: en vez de
#     editar el codigo, se corrige aqui.
#   - mostrar_notas_en_dashboard: false (por defecto) deja las notas de datos
#     dentro de Configuracion; true las trae al tablero, arriba del todo.
#
# Si el archivo no existe, manda lo que diga el codigo. Nada mas.
NOMBRE_CONFIG = "configuracion.json"
RUTA_CONFIG = os.path.join(CARPETA_BASE, NOMBRE_CONFIG)

# Nombre original en 'Print' -> nombre a mostrar. Se rellena desde el JSON.
NOMBRES_MOSTRADOS = {}


class ErrorDatosExcel(Exception):
    """Error de estructura o de datos del Excel que impide generar el dashboard."""


# =============================================================================
# Utilidades
# =============================================================================

def _norm(txt):
    """Normaliza texto para comparar etiquetas: sin acentos, minusculas, sin
    espacios/saltos de linea repetidos."""
    if txt is None:
        return ""
    if not isinstance(txt, str):
        txt = str(txt)
    txt = txt.replace("\n", " ").strip()
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    txt = re.sub(r"\s+", " ", txt)
    return txt.lower().strip()


def archivo_mas_reciente(carpeta):
    candidatos = [
        f for f in glob.glob(os.path.join(carpeta, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    ]
    if not candidatos:
        raise ErrorDatosExcel(
            f"No hay ningun archivo .xlsx en la carpeta:\n         {carpeta}\n"
            f"         Copia ahi el Excel del mes y vuelve a ejecutar."
        )
    return max(candidatos, key=os.path.getmtime)


def a_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except ValueError:
                continue
        raise ErrorDatosExcel(f"No se pudo interpretar la fecha {valor!r}")
    raise ErrorDatosExcel(f"Valor de fecha inesperado: {valor!r} ({type(valor)})")


class _Celda:
    """Celda minima con la misma interfaz que usa openpyxl (value/row/column)."""
    __slots__ = ("row", "column", "value")

    def __init__(self, row, column, value):
        self.row = row
        self.column = column
        self.value = value


class HojaEnMemoria:
    """Copia en memoria de una hoja pequeña.

    El libro se abre siempre en modo read_only porque contiene hojas enormes
    (una declara mas de un millon de filas) que en modo normal se cargarian
    completas aunque no se usen. Las hojas chicas se vuelcan aqui para poder
    consultarlas por coordenadas con comodidad."""

    def __init__(self, titulo, filas):
        self.title = titulo
        self._filas = filas
        self.max_row = len(filas)
        self.max_column = max((len(f) for f in filas), default=0)

    def cell(self, row, column):
        valor = None
        if 1 <= row <= len(self._filas):
            fila = self._filas[row - 1]
            if 1 <= column <= len(fila):
                valor = fila[column - 1]
        return _Celda(row, column, valor)

    def iter_rows(self, min_row=1, max_row=None):
        tope = min(max_row or self.max_row, self.max_row)
        for r in range(min_row, tope + 1):
            yield tuple(self.cell(r, c) for c in range(1, self.max_column + 1))


def hoja_a_memoria(wb, nombre_hoja, limite_filas=2000):
    if nombre_hoja not in wb.sheetnames:
        raise ErrorDatosExcel(
            f"El Excel no tiene la hoja '{nombre_hoja}'. Comprueba que has copiado "
            f"el archivo correcto en la carpeta Entrada."
        )
    ws = wb[nombre_hoja]
    filas = []
    vacias_seguidas = 0
    for fila in ws.iter_rows(values_only=True):
        if fila is None or all(v is None for v in fila):
            vacias_seguidas += 1
            if vacias_seguidas > 50:
                break
        else:
            vacias_seguidas = 0
        filas.append(fila if fila is not None else ())
        if len(filas) >= limite_filas:
            break
    while filas and all(v is None for v in filas[-1]):
        filas.pop()
    return HojaEnMemoria(nombre_hoja, filas)


def colas_de_campana(nombre_campana):
    """Nombres de cola que le corresponden a una campana de 'Print'.
    Acepta tanto una lista como un unico nombre suelto."""
    valor = MAPEO_COLAS.get(nombre_campana)
    if not valor:
        return []
    if isinstance(valor, str):
        return [valor]
    return list(valor)


def campana_por_cola():
    """dict: nombre de cola normalizado -> campana a la que pertenece."""
    return {
        _norm(cola): campana
        for campana in MAPEO_COLAS
        for cola in colas_de_campana(campana)
    }


def nombre_visible(clave):
    """Nombre con el que se muestra una campana en el tablero.

    Es SOLO cosmetico: el resto del script sigue trabajando con el nombre tal
    cual viene de 'Print', de modo que renombrar nunca puede mover una cifra."""
    return NOMBRES_MOSTRADOS.get(clave, clave) or clave


def cargar_configuracion():
    """Lee configuracion.json (si existe) y lo aplica sobre MAPEO_COLAS y sobre
    los nombres a mostrar. Devuelve la configuracion ya limpia.

    Un JSON roto detiene el proceso en vez de ignorarse en silencio: si alguien
    corrigio ahi el nombre de una cola y el archivo no se lee, el dashboard
    saldria con llamadas de menos y sin avisar."""
    cfg = {"nombres_campanas": {}, "mapeo_colas": {}, "mostrar_notas_en_dashboard": False}
    if not os.path.isfile(RUTA_CONFIG):
        return cfg

    try:
        with open(RUTA_CONFIG, "r", encoding="utf-8-sig") as f:
            datos = json.load(f)
    except ValueError as e:
        raise ErrorDatosExcel(
            f"El archivo '{NOMBRE_CONFIG}' no es un JSON valido ({e}).\n"
            f"    Vuelve a guardarlo desde la seccion Configuracion del dashboard, "
            f"o borralo para volver a los valores del codigo."
        )
    if not isinstance(datos, dict):
        raise ErrorDatosExcel(
            f"El archivo '{NOMBRE_CONFIG}' tiene que contener un objeto JSON."
        )

    nombres = datos.get("nombres_campanas") or {}
    if not isinstance(nombres, dict):
        raise ErrorDatosExcel(
            f"'nombres_campanas' en '{NOMBRE_CONFIG}' tiene que ser un objeto "
            f'del tipo {{"nombre en Print": "nombre a mostrar"}}.'
        )
    cfg["nombres_campanas"] = {
        str(k): str(v).strip()
        for k, v in nombres.items()
        if v is not None and str(v).strip() and str(v).strip() != str(k)
    }

    mapeo = datos.get("mapeo_colas") or {}
    if not isinstance(mapeo, dict):
        raise ErrorDatosExcel(
            f"'mapeo_colas' en '{NOMBRE_CONFIG}' tiene que ser un objeto del tipo "
            f'{{"nombre en Print": ["cola 1", "cola 2"]}}.'
        )
    limpio = {}
    for campana, colas in mapeo.items():
        if isinstance(colas, str):
            colas = colas.split(",")
        if colas is None:
            colas = []
        if not isinstance(colas, (list, tuple)):
            raise ErrorDatosExcel(
                f"Las colas de '{campana}' en '{NOMBRE_CONFIG}' tienen que ser una lista "
                f'de nombres, por ejemplo ["NEXOTL 1", "NEXOTL 2"], o [] si no tiene llamadas.'
            )
        limpio[str(campana)] = [str(c).strip() for c in colas if str(c).strip()]
    cfg["mapeo_colas"] = limpio
    cfg["mostrar_notas_en_dashboard"] = bool(datos.get("mostrar_notas_en_dashboard", False))

    MAPEO_COLAS.update(cfg["mapeo_colas"])
    NOMBRES_MOSTRADOS.clear()
    NOMBRES_MOSTRADOS.update(cfg["nombres_campanas"])
    return cfg


def verificar_no_error(valor, contexto):
    if isinstance(valor, str) and valor.strip() in ("#REF!", "#VALUE!", "#N/A", "#DIV/0!"):
        raise ErrorDatosExcel(f"Celda con error {valor!r} en {contexto}. Revisa el Excel de origen.")
    return valor


# =============================================================================
# Fase 1 - Extraccion: hoja Print
# =============================================================================

def leer_encabezado_corte(ws):
    """Busca la etiqueta 'Corte' y toma la fecha (celda inmediata a la
    derecha) y los dias habiles (siguiente celda a la derecha). Busca 'Mes'
    para el nombre del mes."""
    fecha_corte = None
    dias_habiles = None
    mes = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if _norm(cell.value) == "corte":
                r, c = cell.row, cell.column
                fecha_corte = verificar_no_error(ws.cell(row=r, column=c + 1).value, "Print (Corte, fecha)")
                dias_habiles = verificar_no_error(ws.cell(row=r, column=c + 2).value, "Print (Corte, dias habiles)")
            elif _norm(cell.value) == "mes":
                r, c = cell.row, cell.column
                mes = verificar_no_error(ws.cell(row=r, column=c + 1).value, "Print (Mes)")

    if fecha_corte is None or dias_habiles is None:
        raise ErrorDatosExcel(
            "No se encontro la etiqueta 'Corte' (o sus valores) en la hoja 'Print'."
        )
    if not isinstance(dias_habiles, (int, float)):
        raise ErrorDatosExcel(f"El valor de dias habiles junto a 'Corte' no es numerico: {dias_habiles!r}")

    return a_fecha(mes) if mes else a_fecha(fecha_corte).replace(day=1), a_fecha(fecha_corte), dias_habiles


def _fila_a_dict(ws, fila, col_map):
    out = {}
    for nombre_col, idx in col_map.items():
        v = ws.cell(row=fila, column=idx).value
        verificar_no_error(v, f"Print!fila{fila} columna {nombre_col!r}")
        out[nombre_col] = v
    return out


def leer_tabla(ws, etiquetas_clave, nombre_tabla, max_filas_busqueda=300):
    """Localiza una tabla por sus etiquetas de cabecera (busca una fila que
    contenga TODAS las etiquetas de etiquetas_clave), lee las filas de datos
    (aquellas con la primera columna no vacia) y se detiene en la fila de
    totales (primera fila con primera columna vacia pero con datos
    numericos en el resto)."""
    fila_cab = None
    col_map = {}

    for row in ws.iter_rows(min_row=1, max_row=max_filas_busqueda):
        valores_norm = {_norm(c.value): c for c in row if c.value is not None}
        if all(_norm(et) in valores_norm for et in etiquetas_clave):
            fila_cab = row[0].row
            for c in row:
                if isinstance(c.value, str) and c.value.strip():
                    col_map[c.value.strip()] = c.column
            break

    if fila_cab is None:
        raise ErrorDatosExcel(
            f"No se encontro la cabecera de {nombre_tabla} (se buscaron las etiquetas {etiquetas_clave}) "
            f"en la hoja 'Print'."
        )

    primera_col_nombre = next(iter(col_map))  # 'Campaña' siempre es la primera columna leida
    primera_col_idx = col_map[primera_col_nombre]

    filas = []
    fila_totales = None
    r = fila_cab + 1
    vacios_seguidos = 0
    while r <= ws.max_row and vacios_seguidos < 3:
        val_primera_col = ws.cell(row=r, column=primera_col_idx).value
        fila_dict = _fila_a_dict(ws, r, col_map)
        todo_vacio = all(v is None for v in fila_dict.values())

        if todo_vacio:
            vacios_seguidos += 1
            r += 1
            continue
        vacios_seguidos = 0

        if val_primera_col is None:
            # primera fila sin nombre de campana pero con datos = fila de totales
            fila_totales = fila_dict
            break

        if isinstance(val_primera_col, str) and _norm(val_primera_col) == _norm(primera_col_nombre):
            # nos topamos con la cabecera de la SIGUIENTE tabla (mismo texto repetido) -> no es dato
            break

        filas.append(fila_dict)
        r += 1

    if fila_totales is None:
        raise ErrorDatosExcel(f"No se encontro la fila de totales de {nombre_tabla} en 'Print'.")

    return col_map, filas, fila_totales


def leer_total_general(ws):
    """El 'Total general' (llamadas + leads) no vive en la fila de totales de
    ninguna de las dos tablas: esta en una fila propia etiquetada 'Total',
    debajo del bloque de leads, con el importe en la celda de al lado."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if _norm(cell.value) == "total":
                valor = ws.cell(row=cell.row, column=cell.column + 1).value
                verificar_no_error(valor, "Print (Total general)")
                if isinstance(valor, (int, float)):
                    return valor
    raise ErrorDatosExcel(
        "No se encontro la fila 'Total general' (etiqueta 'Total' con importe a la derecha) "
        "en la hoja 'Print'."
    )


def extraer_print(ws):
    mes, fecha_corte, dias_habiles = leer_encabezado_corte(ws)

    _, campanas, totales_llamadas = leer_tabla(
        ws, ["Campaña", "Monto", "Numero de llamadas", "#Perdidas", "Costo Llamada"], "Tabla 1 (Llamadas)"
    )
    _, leads, totales_leads = leer_tabla(
        ws, ["Campaña", "Numero de leads", "Costo Lead"], "Tabla 2 (Leads)"
    )

    return {
        "mes": mes,
        "fecha_corte": fecha_corte,
        "dias_habiles": dias_habiles,
        "campanas": campanas,
        "totales_llamadas": totales_llamadas,
        "leads": leads,
        "totales_leads": totales_leads,
        "total_general": leer_total_general(ws),
    }


# =============================================================================
# Fase 1 - Extraccion: RawDataRingCentral (llamadas por dia)
# =============================================================================

def _leer_hoja_llamadas(wb, nombre_hoja):
    """Lee una hoja de llamadas en streaming: declaran decenas de miles de
    filas y solo unas pocas tienen datos, asi que se corta al detectar filas
    vacias consecutivas. Agrupa #Calls y Missed por (fecha, cola). No
    recalcula logica de negocio: esas dos columnas ya vienen calculadas por
    formula en el propio Excel."""
    ws = wb[nombre_hoja]

    it = ws.iter_rows(values_only=True)
    cabecera = next(it)
    idx = {str(v).strip(): i for i, v in enumerate(cabecera) if v is not None}

    requeridas = ["Date", "Queue Name", "#Calls", "Missed"]
    faltantes = [c for c in requeridas if c not in idx]
    if faltantes:
        raise ErrorDatosExcel(
            f"Faltan las columnas {faltantes} en la hoja '{nombre_hoja}'. "
            f"Encontradas: {list(idx)}"
        )

    i_date, i_queue, i_calls, i_missed = idx["Date"], idx["Queue Name"], idx["#Calls"], idx["Missed"]

    agregado = defaultdict(lambda: [0, 0])  # (fecha, cola) -> [calls, missed]
    vacios_seguidos = 0

    for row in it:
        if row is None or all(v is None for v in row):
            vacios_seguidos += 1
            if vacios_seguidos > 500:
                break
            continue
        vacios_seguidos = 0

        cola = row[i_queue] if i_queue < len(row) else None
        fecha_val = row[i_date] if i_date < len(row) else None
        if cola is None or fecha_val is None:
            continue

        for v in (row[i_calls] if i_calls < len(row) else None, row[i_missed] if i_missed < len(row) else None):
            verificar_no_error(v, f"hoja '{nombre_hoja}', fila de la cola '{cola}'")

        fecha = a_fecha(fecha_val)
        calls = row[i_calls] or 0
        missed = row[i_missed] or 0
        clave = (fecha, str(cola).strip())
        agregado[clave][0] += calls
        agregado[clave][1] += missed

    return agregado  # dict[(date, queue_name)] = [calls, missed]


def hojas_llamadas_auxiliares(wb):
    """Hojas de llamadas distintas de la principal, detectadas por prefijo.

    El nombre cambia de un mes a otro ('RawDataRingCentral HM2' en mayo,
    'RawDataRingCentral HM2-IA' desde junio), asi que buscarlas por nombre
    exacto dejaba fuera meses enteros sin que nada lo delatara salvo el
    descuadre de totales."""
    prefijo = _norm(HOJA_RAW_LLAMADAS)
    return [
        nombre for nombre in wb.sheetnames
        if _norm(nombre).startswith(prefijo) and _norm(nombre) != prefijo
    ]


def extraer_llamadas_diarias(wb, campanas_print):
    """Devuelve las llamadas por (fecha, cola) y de que hoja auxiliar se
    rescato cada campana, como lista de (campana, hoja).

    La hoja principal manda. Si una campana declara llamadas en 'Print' pero
    ninguna de sus colas aparece alli, se la busca en las hojas auxiliares.
    Asi funciona tanto en los meses en que esas hojas sobran como en los que
    son la unica fuente, sin contar nada dos veces."""
    if HOJA_RAW_LLAMADAS not in wb.sheetnames:
        raise ErrorDatosExcel(f"El Excel no tiene la hoja '{HOJA_RAW_LLAMADAS}'.")

    agregado = _leer_hoja_llamadas(wb, HOJA_RAW_LLAMADAS)
    colas_presentes = {_norm(cola) for (_fecha, cola) in agregado}

    faltantes = []
    for camp in campanas_print:
        nombre = camp["Campaña"]
        alias = colas_de_campana(nombre)
        if not alias or not (camp.get("Numero de llamadas") or 0):
            continue
        if not any(_norm(a) in colas_presentes for a in alias):
            faltantes.append(nombre)

    rescatadas = []
    pendientes = list(faltantes)
    for hoja in hojas_llamadas_auxiliares(wb):
        if not pendientes:
            break
        alias_buscados = {_norm(a) for n in pendientes for a in colas_de_campana(n)}
        extra = _leer_hoja_llamadas(wb, hoja)
        encontradas = set()
        for (fecha, cola), (calls, missed) in extra.items():
            if _norm(cola) in alias_buscados:
                agregado[(fecha, cola)][0] += calls
                agregado[(fecha, cola)][1] += missed
                encontradas.add(_norm(cola))
        recuperadas_aqui = [
            n for n in pendientes
            if any(_norm(a) in encontradas for a in colas_de_campana(n))
        ]
        rescatadas.extend((n, hoja) for n in recuperadas_aqui)
        pendientes = [n for n in pendientes if n not in recuperadas_aqui]

    return agregado, rescatadas


# =============================================================================
# Fase 1 - Extraccion: Raw Data General (ventas por dia y por campana)
# =============================================================================

def _titulo_arriba(ws, fila, col):
    for rr in range(max(1, fila - 3), fila):
        v = ws.cell(row=rr, column=col).value
        if v not in (None, ""):
            return str(v).strip()
    return None


def _leer_bloque(ws, fila_cab, col_inicio, max_filas=60):
    """Lee un bloque de tabla dinamica de 'Raw Data General'.

    Las columnas se localizan por el texto de su cabecera y NUNCA por
    posicion: el orden de 'Count of WO Sale Date' y 'Sum of V-linea' se
    invierte de un mes a otro (en agosto 2026 iba primero el conteo de
    clientes, en julio 2026 la suma de lineas). Leerlas por posicion
    intercambiaria clientes con lineas sin que salte ningun error."""
    cabeceras = [ws.cell(row=fila_cab, column=col_inicio + i).value for i in range(7)]

    desplazamiento = {}
    for i, valor in enumerate(cabeceras):
        etiqueta = _norm(valor)
        if etiqueta == "count of wo sale date" and "cliente" not in desplazamiento:
            desplazamiento["cliente"] = i
        elif etiqueta == "sum of v-linea" and "linea" not in desplazamiento:
            desplazamiento["linea"] = i
        elif etiqueta == "costos" and "costos" not in desplazamiento:
            desplazamiento["costos"] = i

    titulo = _titulo_arriba(ws, fila_cab, col_inicio)
    base = {
        "col": col_inicio,
        "fila": fila_cab,
        "titulo": titulo,
        "cabeceras": cabeceras,
        "diario": {},
        "costos": {},
        "grand_total": None,
    }

    # Sin las dos columnas de ventas el bloque no sirve para nada; se devuelve
    # vacio y el emparejamiento lo descartara solo.
    if "cliente" not in desplazamiento or "linea" not in desplazamiento:
        return base

    i_cli, i_lin = desplazamiento["cliente"], desplazamiento["linea"]
    i_cos = desplazamiento.get("costos")

    filas_diarias = {}  # date -> (ventas_cliente, ventas_linea)
    costos_por_dia = {}
    grand_total = None
    r = fila_cab + 1
    while r <= fila_cab + max_filas:
        v0 = ws.cell(row=r, column=col_inicio).value
        if v0 is None:
            break
        if isinstance(v0, str) and _norm(v0) == "grand total":
            grand_total = (
                ws.cell(row=r, column=col_inicio + i_cli).value,
                ws.cell(row=r, column=col_inicio + i_lin).value,
            )
            break
        if isinstance(v0, (datetime, date)):
            fecha = a_fecha(v0)
            cliente = ws.cell(row=r, column=col_inicio + i_cli).value or 0
            linea = ws.cell(row=r, column=col_inicio + i_lin).value or 0
            filas_diarias[fecha] = (cliente, linea)
            if i_cos is not None:
                costo = ws.cell(row=r, column=col_inicio + i_cos).value
                if costo:
                    costos_por_dia[fecha] = costo
        r += 1

    base["diario"] = filas_diarias
    base["costos"] = costos_por_dia
    base["grand_total"] = grand_total
    return base


def extraer_bloques_raw_data_general(ws):
    if _norm(ws.title) != _norm(HOJA_RAW_VENTAS):
        pass  # solo por claridad; ws ya viene de la hoja correcta

    posiciones = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if _norm(cell.value) == "row labels":
                posiciones.append((cell.row, cell.column))

    if not posiciones:
        raise ErrorDatosExcel(f"No se encontro ningun bloque 'Row Labels' en {HOJA_RAW_VENTAS!r}")

    return [_leer_bloque(ws, r, c) for r, c in posiciones]


def deduplicar_bloques(bloques):
    """La hoja trae bloques repetidos con exactamente los mismos datos. Sumarlos
    todos duplicaria las ventas, asi que se conserva uno solo de cada grupo
    identico. Entre duplicados se prefiere el que tiene titulo, porque el titulo
    permite identificar la campana sin ambiguedad."""
    vistos = {}
    for b in bloques:
        clave = (
            tuple(sorted(b["diario"].items())),
            b["grand_total"],
            tuple(sorted(b["costos"].items())),
        )
        anterior = vistos.get(clave)
        if anterior is None:
            vistos[clave] = b
        elif not anterior["titulo"] and b["titulo"]:
            vistos[clave] = b  # nos quedamos con la version que si trae titulo
    return list(vistos.values())


def _bloque_compatible(bloque, ventas_cliente, ventas_linea, monto=None):
    """Un bloque puede pertenecer a una campana si su Grand Total coincide con
    las ventas de 'Print'. El Monto NO se usa aqui: los costos del bloque unas
    veces vienen en un solo dia y otras repartidos, y a veces no suman el total
    de 'Print'. Usarlo para descartar rechazaba bloques correctos; solo sirve
    para desempatar (ver _costo_respalda)."""
    return bloque["grand_total"] == (ventas_cliente, ventas_linea)


def _costo_respalda(bloque, monto):
    """True si los costos del bloque apoyan que pertenezca a esa campana, ya
    sea porque suman su Monto o porque algun dia suelto lo iguala. Solo se
    consulta para deshacer empates entre campanas con las mismas ventas."""
    costos = [c for c in bloque["costos"].values() if c]
    if not costos or not monto:
        return False
    tolerancia = max(0.02, abs(monto) * 0.001)
    if abs(sum(costos) - monto) <= tolerancia:
        return True
    return any(abs(c - monto) <= tolerancia for c in costos)


def asignar_bloques_a_campanas(bloques, campanas_objetivo):
    """campanas_objetivo: dict nombre_campana -> (ventas_cliente, ventas_linea, monto)
    Devuelve (asignaciones: dict nombre -> bloque, avisos, campanas sin bloque)."""
    asignaciones = {}
    avisos = []
    restantes = deduplicar_bloques(bloques)

    # Paso 1: por titulo del bloque (coincidencia de prefijo, sin acentos ni mayusculas)
    for camp in list(campanas_objetivo):
        for b in list(restantes):
            if not b["titulo"]:
                continue
            t, c = _norm(b["titulo"]), _norm(camp)
            if t.startswith(c) or c.startswith(t):
                if camp not in asignaciones:
                    asignaciones[camp] = b
                    restantes.remove(b)
                break

    # Paso 2: emparejar por Grand Total, aceptando solo correspondencias 1 a 1.
    # Cuando varias campanas se disputan el mismo bloque (mismas ventas por
    # casualidad, o por una referencia cruzada del propio Excel) se recurre al
    # Monto para deshacer el empate. Si aun asi queda ambiguo no se adivina:
    # esa campana se queda sin serie diaria y se reporta.
    pendientes = {
        camp: datos for camp, datos in campanas_objetivo.items()
        if camp not in asignaciones and (datos[0] or datos[1])
    }
    compatibles = {
        camp: [b for b in restantes if _bloque_compatible(b, vc, vl)]
        for camp, (vc, vl, _monto) in pendientes.items()
    }

    for camp, candidatos in compatibles.items():
        if len(candidatos) != 1:
            continue
        bloque = candidatos[0]
        if bloque not in restantes:
            continue  # ya se lo llevo otra campana
        rivales = [otra for otra, cs in compatibles.items() if otra != camp and bloque in cs]

        if rivales:
            # Desempate por Monto: se lo queda la unica campana cuyo importe
            # respalden los costos del bloque.
            aspirantes = [
                c for c in [camp] + rivales
                if _costo_respalda(bloque, campanas_objetivo[c][2])
            ]
            if len(aspirantes) != 1:
                avisos.append(
                    f"AVISO: el bloque de '{HOJA_RAW_VENTAS}' con ventas "
                    f"{bloque['grand_total'][0]}/{bloque['grand_total'][1]} encaja con varias "
                    f"campañas ({', '.join([camp] + rivales)}) y el Monto no permite decidir. "
                    f"No se asigna a ninguna para no atribuir ventas al azar."
                )
                continue
            if aspirantes[0] != camp:
                continue  # le corresponde a otra, ya le tocara su turno

        asignaciones[camp] = bloque
        restantes.remove(bloque)

    sin_bloque = [
        camp for camp, (vc, vl, _m) in campanas_objetivo.items()
        if camp not in asignaciones and (vc or vl)
    ]
    for camp in sin_bloque:
        if camp not in CAMPANAS_SIN_SERIE_DIARIA_CONOCIDAS:
            avisos.append(
                f"AVISO: la campaña '{camp}' tiene ventas en 'Print' pero no se encontro su bloque "
                f"en '{HOJA_RAW_VENTAS}'. Se excluye de la serie diaria de ventas; sus totales del "
                f"corte se siguen mostrando."
            )

    return asignaciones, avisos, sin_bloque


# =============================================================================
# Fase 1 - Extraccion: Valores$$$ (costo del corte)
# =============================================================================

def extraer_costo_corte(ws):
    """La hoja tiene una fila por dia y una columna 'Total' con el gasto de
    ese dia. Este mes solo el dia 1 tiene importe; el resto viene en cero
    porque se carga a mano y con retraso. Se reporta como dato puntual del
    corte, nunca como serie diaria completa (para no sugerir que el gasto
    se desplomo)."""
    cabecera = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    idx_total = None
    for i, v in enumerate(cabecera, start=1):
        if _norm(v) == "total":
            idx_total = i
            break
    if idx_total is None:
        return None  # no bloquea el dashboard, es informacion complementaria

    dias_con_costo = []
    for r in range(2, ws.max_row + 1):
        fecha_val = ws.cell(row=r, column=1).value
        if not isinstance(fecha_val, (datetime, date)):
            continue
        total = ws.cell(row=r, column=idx_total).value or 0
        if total:
            dias_con_costo.append((a_fecha(fecha_val), total))

    return dias_con_costo


# =============================================================================
# Construccion de la serie diaria combinada (llamadas + ventas)
# =============================================================================

def construir_serie_diaria(datos_print, llamadas_agregadas, bloques_asignados, mes, fecha_corte):
    campanas = datos_print["campanas"]

    # El eje del timeline abarca el mes completo, no solo hasta el corte.
    dias_del_mes = []
    d = mes.replace(day=1)
    while d.month == mes.month:
        dias_del_mes.append(d)
        d += timedelta(days=1)

    serie = {d: {"llamadas": 0, "perdidas": 0, "ventas_cliente": 0, "ventas_linea": 0,
                 "llamadas_efic": 0, "tiene_datos": False}
             for d in dias_del_mes}

    # Colas de las campanas que SI aportan ventas diarias. La eficiencia diaria
    # tiene que dividir ventas y llamadas del mismo conjunto de campanas: si una
    # campana no aporta su numerador, tampoco puede aportar su denominador, o la
    # curva se hunde por una diferencia de cobertura y no por rendimiento real.
    de_cola_a_campana = campana_por_cola()
    colas_con_ventas = {
        _norm(cola)
        for camp in campanas
        if bloques_asignados.get(camp["Campaña"])
        for cola in colas_de_campana(camp["Campaña"])
    }

    # Llamadas: sumar todas las colas mapeadas, por fecha
    for (fecha, cola), (calls, missed) in llamadas_agregadas.items():
        cola_norm = _norm(cola)
        if cola_norm not in de_cola_a_campana:
            continue
        if fecha not in serie:
            continue  # fecha fuera del mes del corte, se ignora
        serie[fecha]["llamadas"] += calls
        serie[fecha]["perdidas"] += missed
        if cola_norm in colas_con_ventas:
            serie[fecha]["llamadas_efic"] += calls
        serie[fecha]["tiene_datos"] = True

    # Ventas: solo de las campanas con bloque asignado (ver aviso de exclusion)
    for camp in campanas:
        nombre = camp["Campaña"]
        bloque = bloques_asignados.get(nombre)
        if not bloque:
            continue
        for fecha, (count, suma) in bloque["diario"].items():
            if fecha not in serie:
                continue
            serie[fecha]["ventas_cliente"] += count
            serie[fecha]["ventas_linea"] += suma

    return dias_del_mes, serie, fecha_corte


# =============================================================================
# Fase 2 - Validacion (antes de renderizar)
# =============================================================================

def validar(datos_print, serie_diaria_dict, avisos, colas_del_raw=()):
    errores = []
    tot = datos_print["totales_llamadas"]
    tot_leads = datos_print["totales_leads"]

    # ---------------------------------------------------------------------
    # Validaciones que corren SIEMPRE, sea el mes que sea. Son las que
    # protegen a los meses futuros: sin ellas, una campana nueva sin mapear
    # perderia sus llamadas en silencio.
    # ---------------------------------------------------------------------

    # 1) Toda campana de 'Print' tiene que estar en MAPEO_COLAS.
    sin_mapear = [
        c["Campaña"] for c in datos_print["campanas"]
        if c["Campaña"] not in MAPEO_COLAS
    ]
    if sin_mapear:
        errores.append(
            "Hay campañas en la hoja 'Print' que no están en MAPEO_COLAS: "
            + ", ".join(f"'{c}'" for c in sin_mapear)
            + ".\n    Lo más rápido: abre el dashboard anterior, entra en Configuración y añade "
              "ahí sus colas (se guarda en " + NOMBRE_CONFIG + ", sin tocar código).\n"
            + "    O bien abre dashboard_kpi.py, busca MAPEO_COLAS y añade una línea por cada una:\n"
            + "\n".join(
                f'        "{c}": ["nombre de su cola en RawDataRingCentral"],  # o [] si no tiene llamadas'
                for c in sin_mapear
            )
        )

    # 2) La serie diaria de llamadas tiene que sumar el total del corte.
    #    Si no cuadra, algo se esta quedando fuera del conteo.
    suma_serie = sum(v["llamadas"] for v in serie_diaria_dict.values())
    total_print = tot.get("Numero de llamadas")
    if total_print and suma_serie != total_print:
        diferencia = total_print - suma_serie
        errores.append(
            f"La suma de llamadas día a día ({fmt_num(suma_serie)}) no coincide con el total de "
            f"'Print' ({fmt_num(total_print)}). Faltan {fmt_num(diferencia)} llamadas.\n"
            f"    Suele significar que una cola de RawDataRingCentral no está asignada "
            f"a ninguna campaña en MAPEO_COLAS."
        )

    # 3) Colas presentes en el raw que no conocemos: aviso, no error. Puede ser
    #    una campana nueva que todavia no llego a 'Print'.
    colas_conocidas = set(campana_por_cola()) | {_norm(c) for c in COLAS_IGNORADAS_CONOCIDAS}
    desconocidas = sorted({c for c in colas_del_raw if c and _norm(c) not in colas_conocidas})
    if desconocidas:
        avisos.append(
            "AVISO: hay colas en '" + HOJA_RAW_LLAMADAS + "' que no están en MAPEO_COLAS y "
            "por eso no se cuentan: " + ", ".join(f"'{c}'" for c in desconocidas)
            + ". Si alguna es una campaña nueva, añádela al MAPEO_COLAS."
        )

    # ---------------------------------------------------------------------
    # Cifras exactas del mes, solo si ese mes esta en la tabla de control.
    # ---------------------------------------------------------------------
    control = CIFRAS_DE_CONTROL_POR_CORTE.get(
        (datos_print["mes"], datos_print["fecha_corte"])
    )
    if control:
        llamadas = tot.get("Numero de llamadas")
        chequeos = [
            ("Llamadas totales", llamadas, control["llamadas_totales"]),
            ("Llamadas perdidas", tot.get("#Perdidas"), control["llamadas_perdidas"]),
            ("Ventas (Cliente)", tot.get("Ventas (Cliente)"), control["ventas_cliente"]),
            ("Ventas (Linea)", tot.get("Ventas (Linea)"), control["ventas_linea"]),
            ("Monto llamadas", tot.get("Monto"), control["monto_llamadas"]),
            ("% perdidas", tot.get("% Llamadas Perdidas"), control["pct_perdidas"]),
            ("Total general", datos_print.get("total_general"), control["total_general"]),
            ("Llamadas / dia", (llamadas / datos_print["dias_habiles"]) if llamadas and
             datos_print["dias_habiles"] else None, control["llamadas_dia"]),
            ("Eficiencia /Cliente", (tot["Ventas (Cliente)"] / llamadas) if llamadas else None,
             control["eficiencia_cliente"]),
            ("Eficiencia /Linea", (tot["Ventas (Linea)"] / llamadas) if llamadas else None,
             control["eficiencia_linea"]),
            ("Suma de la serie diaria de llamadas",
             sum(v["llamadas"] for v in serie_diaria_dict.values()), control["llamadas_totales"]),
        ]
        for nombre, real, esperado in chequeos:
            if real is None:
                errores.append(f"{nombre}: no se pudo leer (esperado {esperado})")
            elif abs(real - esperado) > max(0.01, abs(esperado) * 0.001):
                errores.append(f"{nombre}: esperado {esperado}, se leyo {real}")

        leads_heri = next(
            (f for f in datos_print["leads"] if _norm(f["Campaña"]) == _norm("META HERI")), None
        )
        if leads_heri is None:
            errores.append("No se encontro la campaña 'META HERI' en la tabla de leads")
        elif leads_heri.get("Numero de leads") != control["leads_meta_heri"]:
            errores.append(
                f"Leads META HERI: esperado {control['leads_meta_heri']}, "
                f"se leyo {leads_heri.get('Numero de leads')}"
            )

    if errores:
        mensaje = "Validacion fallida, NO se genera el dashboard:\n  - " + "\n  - ".join(errores)
        raise ErrorDatosExcel(mensaje)

    return avisos


# =============================================================================
# Formato
# =============================================================================

MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}
MESES_ES_CORTO = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
    7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
}


def _a_formato_es(texto):
    """Convierte 1,234.56 (formato de Python) a 1.234,56 (formato espanol)."""
    return texto.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def fmt_num(v, decimales=0):
    if v is None:
        return "—"
    try:
        return _a_formato_es(f"{v:,.{decimales}f}")
    except (TypeError, ValueError):
        return "—"


def fmt_pct(v, decimales=1):
    if v is None:
        return "—"
    try:
        return _a_formato_es(f"{v * 100:,.{decimales}f}") + " %"
    except (TypeError, ValueError):
        return "—"


def fmt_usd(v, decimales=2):
    if v is None:
        return "—"
    try:
        return "$" + _a_formato_es(f"{v:,.{decimales}f}")
    except (TypeError, ValueError):
        return "—"


def fmt_fecha_es(f):
    return f"{f.day:02d} {MESES_ES_CORTO[f.month]}"


def fmt_mes_es(f):
    return f"{MESES_ES[f.month].capitalize()} {f.year}"


# =============================================================================
# Fase 3 - Render: sistema de diseno
# =============================================================================
#
# El HTML se abre sin internet: no hay fuentes web ni CSS externo. Todo el
# diseno vive en las dos constantes de abajo y en los pequenos ayudantes que
# arman los componentes (columnas, barras tramadas, matriz de puntos).

C_AZUL = "#2F6FED"
C_AZUL_OSC = "#1E4FD8"
C_VERDE = "#1FA84A"
C_ROSA = "#EC2E7B"
C_AMBAR = "#E8963C"
C_VIOLETA = "#7C5CE6"
C_TINTA = "#141412"
C_SUAVE = "#8A8A80"
C_TENUE = "#B7B7AE"
C_LINEA = "#E7E7E2"

FUENTE = ('-apple-system, BlinkMacSystemFont, "Segoe UI Variable Display", '
          '"Segoe UI", Inter, Roboto, "Helvetica Neue", Arial, sans-serif')

# Alto en px de la cabecera y de la zona de barras del bloque de columnas.
# Estan aqui porque el eje Y de la izquierda tiene que alinearse con ellas.
H_CAB_COL = 78
H_ZONA_COL = 200

CSS_DASHBOARD = """
:root{
  --bg:#F0F0EE; --tarjeta:#FAFAF9; --blanco:#FFFFFF;
  --linea:#E7E7E2; --linea-fuerte:#DADAD3;
  --tinta:#141412; --tinta-2:#3E3E39; --suave:#8A8A80; --tenue:#B7B7AE;
  --azul:#2F6FED; --azul-osc:#1E4FD8; --verde:#1FA84A; --rosa:#EC2E7B;
  --ambar:#E8963C; --violeta:#7C5CE6; --rojo:#DC2626;
  --r-tarjeta:20px;
  --sombra:0 1px 2px rgba(20,20,15,.04), 0 10px 28px -14px rgba(20,20,15,.16);
  --h-cab-col:78px; --h-zona-col:200px;
}
*,*::before,*::after{box-sizing:border-box}
/* Sin scroll-padding en html: el desfase del ancla lo pone scroll-margin-top de
   cada seccion. Estando los dos, se sumaban (96+96) y al pulsar un enlace la
   seccion aterrizaba 192 px mas abajo de donde el menu la daba por activa, que
   es por lo que se quedaba iluminada la anterior. --h-barra la mide el JS. */
html{-webkit-text-size-adjust:100%; scroll-behavior:smooth}
body{
  margin:0; background:var(--bg); color:var(--tinta);
  font-family:-apple-system,BlinkMacSystemFont,"Segoe UI Variable Display","Segoe UI",Inter,Roboto,"Helvetica Neue",Arial,sans-serif;
  font-size:14px; line-height:1.45; -webkit-font-smoothing:antialiased;
}
.num{font-variant-numeric:tabular-nums}

/* ---------- barra superior ---------- */
.barra-sup{
  position:sticky; top:0; z-index:60; background:rgba(240,240,238,.86);
  backdrop-filter:saturate(180%) blur(14px); border-bottom:1px solid var(--linea);
}
.barra-sup .interior{
  max-width:1400px; margin:0 auto; padding:12px 24px;
  display:flex; align-items:center; gap:28px;
}
.marca{display:flex; align-items:center; gap:11px; font-size:19px; font-weight:600; letter-spacing:-.02em}
/* El logo puede venir de un archivo (un PNG con el emblema y el texto de la
   empresa, mas ancho que alto) o del emblema vectorial de reserva, que es
   cuadrado. El alto manda y el ancho se ajusta solo, con tope para que un
   logo apaisado no empuje la navegacion. */
.marca .logo{
  height:38px; width:auto; max-width:150px; object-fit:contain; display:block;
  flex:0 0 auto; border-radius:7px; /* si el archivo trae fondo blanco, que no corte en seco */
}
.nav{display:flex; gap:2px; flex-wrap:wrap}
.nav a{
  text-decoration:none; color:var(--tinta-2); font-size:14px; font-weight:500;
  padding:9px 16px; border-radius:999px; white-space:nowrap;
  transition:background .15s, color .15s;
}
.nav a:hover{background:rgba(20,20,15,.05)}
.nav a.activo{background:#1D1D1A; color:#fff}

/* La configuracion no es una seccion mas del tablero: vive aparte, detras de
   este boton, para que no se mezcle con los datos del corte. */
.acciones-sup{margin-left:auto; display:flex; align-items:center; gap:8px}
.btn-icono{
  position:relative; display:inline-flex; align-items:center; gap:8px; cursor:pointer;
  background:var(--blanco); border:1px solid var(--linea); border-radius:999px;
  padding:8px 15px; font:inherit; font-size:13.5px; font-weight:500; color:var(--tinta-2);
  box-shadow:0 1px 2px rgba(20,20,15,.05); white-space:nowrap;
  transition:background .15s, color .15s, border-color .15s;
}
.btn-icono:hover{background:#F3F3EF; color:var(--tinta)}
.btn-icono svg{flex:0 0 auto}
.btn-icono .globo{
  position:absolute; top:-3px; right:-3px; min-width:17px; height:17px; padding:0 4px;
  border-radius:999px; background:var(--ambar); color:#fff; font-size:10.5px; font-weight:600;
  display:flex; align-items:center; justify-content:center; box-shadow:0 0 0 2px var(--bg);
}

/* ---------- cabecera de pagina ---------- */
.envoltura{max-width:1400px; margin:0 auto; padding:26px 24px 72px}
.titulo-fila{display:flex; align-items:flex-end; justify-content:space-between; gap:24px; flex-wrap:wrap; margin-bottom:22px}
.titulo-fila h1{margin:0; font-size:44px; font-weight:600; letter-spacing:-.035em; line-height:1.05}
.titulo-fila .sub{margin-top:8px; color:var(--suave); font-size:13.5px}
.herramientas{display:flex; align-items:center; gap:8px; flex-wrap:wrap}
.chip{
  display:inline-flex; align-items:center; gap:9px; background:var(--blanco);
  border:1px solid var(--linea); border-radius:13px; padding:9px 14px;
  font-size:13.5px; color:var(--tinta); white-space:nowrap;
  box-shadow:0 1px 2px rgba(20,20,15,.05);
}
.chip svg{flex:0 0 auto; color:var(--suave)}
.chip.plano{background:transparent; border-color:transparent; box-shadow:none; color:var(--tenue); padding:9px 4px}

/* ---------- rejilla y tarjetas ---------- */
.rejilla{display:grid; grid-template-columns:repeat(12,1fr); gap:14px; margin-bottom:14px}
.s3{grid-column:span 3} .s4{grid-column:span 4} .s6{grid-column:span 6}
.s8{grid-column:span 8} .s12{grid-column:span 12}
.pila{display:flex; flex-direction:column; gap:14px; height:100%}
.pila > *{flex:1}

.tarjeta{
  background:var(--tarjeta); border:1px solid var(--linea); border-radius:var(--r-tarjeta);
  padding:22px 24px; box-shadow:var(--sombra);
}
/* Las tarjetas de una misma fila se igualan en alto, como en el original. */
.rejilla > div > .tarjeta, .rejilla > div > .pila{height:100%}
.tarjeta-cab{display:flex; align-items:flex-start; justify-content:space-between; gap:16px; margin-bottom:18px}
.tarjeta-cab h2{margin:0; font-size:20px; font-weight:600; letter-spacing:-.022em}
.tarjeta-cab .pie-cab{margin-top:5px; font-size:12.5px; color:var(--tenue)}
.cab-der{display:flex; align-items:center; gap:10px; flex:0 0 auto}
.rotulo{font-size:11.5px; text-transform:uppercase; letter-spacing:.09em; color:var(--tenue); font-weight:600; margin:32px 0 12px}
section{scroll-margin-top:calc(var(--h-barra, 58px) + 24px)}

/* boton redondo "..." : abre la nota de que mide la tarjeta */
.ayuda{position:relative; flex:0 0 auto}
.ayuda summary{
  list-style:none; width:36px; height:36px; border-radius:50%;
  border:1px solid var(--linea-fuerte); color:var(--suave); cursor:pointer;
  display:flex; align-items:center; justify-content:center;
  font-size:17px; line-height:1; letter-spacing:1px; user-select:none;
  transition:background .15s, color .15s;
}
.ayuda summary::-webkit-details-marker{display:none}
.ayuda summary:hover{background:rgba(20,20,15,.05); color:var(--tinta)}
.ayuda[open] summary{background:#1D1D1A; color:#fff; border-color:#1D1D1A}
.ayuda-panel{
  position:absolute; right:0; top:44px; width:290px; z-index:40;
  background:var(--blanco); border:1px solid var(--linea); border-radius:14px;
  padding:13px 15px; font-size:12.5px; line-height:1.5; color:var(--tinta-2);
  box-shadow:0 18px 42px -14px rgba(20,20,15,.34);
}

/* ---------- cifras ---------- */
.cifra{font-size:56px; font-weight:600; letter-spacing:-.042em; line-height:1; font-variant-numeric:tabular-nums}
.cifra-md{font-size:38px; font-weight:600; letter-spacing:-.035em; line-height:1; font-variant-numeric:tabular-nums}
.insignia{
  display:inline-flex; align-items:center; gap:6px; background:var(--blanco);
  border:1px solid var(--linea); border-radius:999px; padding:6px 13px;
  font-size:13px; font-weight:500; box-shadow:0 2px 8px -4px rgba(20,20,15,.25);
}
.insignia.verde{color:#15803D} .insignia.rojo{color:#B91C1C}
.punto{width:8px; height:8px; border-radius:50%; flex:0 0 auto}

/* ---------- tarjetas KPI ---------- */
.kpi{background:var(--tarjeta); border:1px solid var(--linea); border-radius:16px; padding:17px 19px; box-shadow:var(--sombra)}
.kpi-label{font-size:12.5px; color:var(--suave); font-weight:500}
.kpi-valor{font-size:29px; font-weight:600; letter-spacing:-.035em; margin-top:9px; font-variant-numeric:tabular-nums; line-height:1.1}
.kpi-sub{font-size:12px; color:var(--tenue); margin-top:7px; font-variant-numeric:tabular-nums}

/* ---------- bloque de columnas (llamadas por campana) ----------
   Las alturas son elasticas a proposito. Esta tarjeta comparte fila con
   'Monto del corte', cuya lista crece con el numero de campanas: con la zona
   de barras fija, todo ese alto de mas se quedaba como un hueco vacio debajo
   del grafico. Ahora las barras se reparten el alto que sobre, sea el que sea,
   y --h-zona-col pasa a ser el minimo en vez del alto exacto. */
.cols{display:flex; min-height:calc(var(--h-cab-col) + var(--h-zona-col))}
.tarjeta.crece .cols{flex:1 1 auto}
.cols-eje{flex:0 0 52px; display:flex; flex-direction:column}
.cols-eje .hueco{height:var(--h-cab-col); flex:0 0 auto}
.cols-eje .escala{position:relative; flex:1 1 auto; min-height:var(--h-zona-col)}
.cols-eje .escala span{position:absolute; right:10px; transform:translateY(50%); font-size:11px; color:var(--tenue); font-variant-numeric:tabular-nums}
.cols-cuerpo{flex:1; display:flex; position:relative; min-width:0; border-bottom:1px solid var(--linea-fuerte)}
.cols-lineas{position:absolute; left:0; right:0; top:var(--h-cab-col); bottom:0; pointer-events:none}
.cols-lineas i{position:absolute; left:0; right:0; height:1px; background:var(--linea)}
.col{
  flex:1; min-width:0; position:relative; display:flex; flex-direction:column;
  border-left:1px solid var(--linea); border-radius:12px 12px 0 0;
}
.col:first-child{border-left:0}
.col.foco{background:linear-gradient(180deg, rgba(47,111,237,.11) 0%, rgba(47,111,237,0) 72%)}
.col-cab{height:var(--h-cab-col); flex:0 0 auto; padding:0 14px}
.col-label{font-size:12.5px; color:var(--tenue); white-space:nowrap; overflow:hidden; text-overflow:ellipsis}
.col-valor{font-size:29px; font-weight:600; letter-spacing:-.035em; color:var(--tenue); margin-top:7px; font-variant-numeric:tabular-nums}
.col.foco .col-label{color:var(--tinta-2)}
.col.foco .col-valor{color:var(--tinta)}
.col-zona{position:relative; flex:1 1 auto; min-height:var(--h-zona-col)}
.col-barra{
  position:absolute; left:13px; right:13px; bottom:0; min-height:3px;
  border-radius:7px 7px 0 0; border-top:2.5px solid var(--azul);
  background-color:rgba(47,111,237,.07);
  background-image:repeating-linear-gradient(-45deg, rgba(47,111,237,.55) 0 2.5px, rgba(47,111,237,0) 2.5px 7px);
}
.col.foco .col-barra{
  background-color:transparent; background-image:none;
  background:linear-gradient(180deg,#4681F7 0%,#1A46C9 100%);
  border-top:0; box-shadow:0 10px 26px -10px rgba(30,79,216,.75);
}
.col-marca{position:absolute; left:50%; transform:translateX(-50%); width:24px; height:5px; border-radius:3px; background:var(--azul); opacity:.32}
.col.foco .col-marca{display:none}
.col-tip{
  position:absolute; left:50%; top:10px; transform:translateX(-50%);
  background:var(--blanco); border:1px solid var(--linea); border-radius:999px;
  padding:8px 15px; font-size:12px; white-space:nowrap; color:var(--tinta-2);
  box-shadow:0 14px 30px -12px rgba(20,20,15,.45);
  opacity:0; pointer-events:none; transition:opacity .15s ease; z-index:12;
}
.col-tip b{color:var(--tinta); font-weight:600}
.col:hover .col-tip{opacity:1}

/* ---------- barras de reparto con tramado ---------- */
/* El tope de alto evita que un mes con muchas campanas estire toda la fila y
   deje el grafico de al lado desproporcionado: a partir de ahi, la lista
   scrollea dentro de la tarjeta. */
.reparto{
  border-top:1px solid var(--linea); margin-top:20px; padding-top:6px;
  max-height:430px; overflow-y:auto; scrollbar-width:thin; scrollbar-gutter:stable;
}
.reparto::-webkit-scrollbar{width:8px}
.reparto::-webkit-scrollbar-thumb{background:#DCDCD5; border-radius:999px}
.reparto::-webkit-scrollbar-track{background:transparent}
.fila-barra{margin-top:16px}
.fila-barra-cab{display:flex; align-items:baseline; justify-content:space-between; gap:12px}
.fila-barra-cab span{font-size:13.5px; color:var(--tinta-2); overflow:hidden; text-overflow:ellipsis; white-space:nowrap}
.fila-barra-cab b{font-size:13.5px; font-weight:600; font-variant-numeric:tabular-nums; white-space:nowrap}
.pista{margin-top:8px; height:13px; border-radius:999px; background:#EBEBE5; overflow:hidden}
.relleno{
  height:100%; border-radius:999px; min-width:6px;
  background-image:repeating-linear-gradient(-45deg, rgba(255,255,255,.5) 0 2px, rgba(255,255,255,0) 2px 5px);
}

/* ---------- mini grafico de area escalonada ---------- */
.tarjeta.crece{display:flex; flex-direction:column}
.tarjeta.crece .mini-chart{margin-top:auto; flex:1; display:flex; flex-direction:column; justify-content:flex-end}
.tarjeta.crece .mini-chart svg{flex:1 1 auto; height:auto; min-height:150px}
.mini-chart{position:relative; margin-top:6px}
.mini-chart svg{display:block; width:100%}
.pico-flotante{
  position:absolute; top:0; transform:translateX(-50%); z-index:2;
  background:var(--blanco); border:1px solid var(--linea); border-radius:999px;
  padding:4px 11px; font-size:11.5px; font-weight:600; font-variant-numeric:tabular-nums;
  box-shadow:0 8px 20px -10px rgba(20,20,15,.4); white-space:nowrap;
  transition:opacity .12s ease;
}
.eje-x{display:flex; justify-content:space-between; margin-top:10px; font-size:11.5px; color:var(--tenue)}

/* Rejilla de referencia del area: dos o tres cifras redondas por debajo del
   maximo. Con un dia pico que aplasta al resto, sin estas lineas no hay forma
   de saber si la parte baja de la serie son 500 llamadas o 2.000. */
.area-reja{position:absolute; inset:0; pointer-events:none}
.area-reja i{position:absolute; left:0; right:0; height:1px; background:var(--linea)}
.area-reja i b{
  position:absolute; right:0; bottom:2px; padding:0 4px; border-radius:5px;
  background:var(--blanco); font-size:10.5px; font-weight:500; color:var(--tenue);
  font-variant-numeric:tabular-nums; font-style:normal; /* la cifra cuelga de un <i> */
}

/* Capa de dias sobre el area: una columna invisible por dia, del mismo ancho
   que su escalon. Al pasar el raton marca el dia y saca su globo, que es lo
   que ya hacen las matrices de puntos de contestadas y perdidas. */
.area-dias{position:absolute; inset:0; display:flex}
.ah-col{position:relative; flex:1 1 0; min-width:0}
.ah-guia{
  position:absolute; left:50%; top:0; bottom:0; width:1px; margin-left:-.5px;
  background:var(--c); opacity:0; transition:opacity .12s ease;
}
.ah-punto{
  position:absolute; left:50%; bottom:var(--h); width:9px; height:9px;
  margin:0 0 -4.5px -4.5px; border-radius:50%; background:var(--c);
  border:2px solid var(--blanco); box-shadow:0 3px 8px -2px rgba(20,20,15,.55);
  opacity:0; transform:scale(.4); transition:opacity .12s ease, transform .12s ease;
}
.ah-col:hover .ah-guia{opacity:.5}
.ah-col:hover .ah-punto{opacity:1; transform:scale(1)}
/* El globo va justo encima del punto del dia, no en un sitio fijo: asi la
   cifra queda al lado de lo que se esta senalando. En los dias altos el min()
   lo frena antes de que se salga por arriba de la tarjeta. */
.ah-tip{
  position:absolute; left:50%; bottom:min(calc(var(--h) + 15px), calc(100% - 30px));
  transform:translateX(-50%);
  background:var(--blanco); border:1px solid var(--linea); border-radius:999px;
  padding:6px 13px; font-size:11.5px; white-space:nowrap; color:var(--tinta-2);
  box-shadow:0 14px 30px -12px rgba(20,20,15,.45);
  opacity:0; pointer-events:none; transition:opacity .12s ease; z-index:14;
}
.ah-tip b{color:var(--tinta); font-weight:600; font-variant-numeric:tabular-nums}
.ah-col:hover .ah-tip{opacity:1}
/* El globo es mucho mas ancho que una columna: en los primeros y los ultimos
   dias se alinea al borde para que no se salga de la tarjeta. */
.ah-col:nth-child(-n+5) .ah-tip{left:0; transform:none}
.ah-col:nth-last-child(-n+5) .ah-tip{left:auto; right:0; transform:none}
/* El globo del dia y la etiqueta del pico ocupan el mismo sitio arriba:
   mientras se recorre la serie, la del pico se aparta. */
.mini-chart:hover .pico-flotante{opacity:0}

/* ---------- matriz de puntos ----------
   La matriz ocupa el ancho entero de la tarjeta, no una columna estrecha entre
   la cifra y el porcentaje: con 31 dias en 180 px los puntos salian como
   rayitas verticales de 3 px. Cada punto es redondo y su diametro sale del
   ancho disponible, con tope, asi que un mes de 31 dias y uno de 12 se ven
   igual de limpios. */
.pila .tarjeta{display:flex; flex-direction:column}
.pila .tarjeta-punto{flex:1}
.tarjeta-punto{display:flex; flex-direction:column; gap:16px}
.punto-cifras{display:flex; align-items:flex-end; justify-content:space-between; gap:18px; flex-wrap:wrap}
.punto-cifras .lado{flex:0 0 auto}
.tarjeta-punto .centro{margin-top:auto}
.etiqueta-pico{
  display:inline-flex; align-items:center; gap:6px; background:var(--blanco);
  border:1px solid var(--linea); border-radius:999px; padding:4px 11px;
  font-size:11.5px; color:var(--tinta-2); box-shadow:0 6px 16px -10px rgba(20,20,15,.4);
  margin-bottom:10px;
}
.etiqueta-pico b{font-weight:600}
.puntos{display:flex; align-items:flex-end; gap:3px; min-height:56px}
.pt-col{
  position:relative; display:flex; flex-direction:column-reverse; align-items:center;
  gap:3px; flex:1 1 0; min-width:0; padding-top:26px;
}
.pt-col i{
  display:block; width:100%; max-width:9px; min-height:4px; aspect-ratio:1;
  border-radius:50%; background:var(--c-tenue);
  transition:background .12s ease, transform .12s ease;
}
.pt-col.pico i{background:var(--c)}
.pt-col:hover i{background:var(--c); transform:scale(1.18)}
.pt-tip{
  position:absolute; left:50%; bottom:calc(100% - 20px); transform:translateX(-50%);
  background:var(--blanco); border:1px solid var(--linea); border-radius:999px;
  padding:5px 11px; font-size:11.5px; white-space:nowrap; color:var(--tinta-2);
  box-shadow:0 14px 28px -14px rgba(20,20,15,.5);
  opacity:0; pointer-events:none; transition:opacity .12s ease; z-index:14;
}
.pt-tip b{color:var(--tinta); font-weight:600; font-variant-numeric:tabular-nums}
.pt-col:hover .pt-tip{opacity:1}
/* El globo mide bastante mas que la columna (10 px): en los primeros y los
   ultimos dias se alinea al borde para que no se salga de la tarjeta. */
.pt-col:nth-child(-n+3) .pt-tip{left:0; transform:none}
.pt-col:nth-last-child(-n+3) .pt-tip{left:auto; right:0; transform:none}
.lado-der{text-align:right; font-size:12px; color:var(--tenue); margin-left:auto}
.lado-der b{display:block; font-size:19px; font-weight:600; color:var(--tinta); letter-spacing:-.02em; font-variant-numeric:tabular-nums; margin-top:3px}

/* ---------- tarjeta de conclusion (degradado) ---------- */
.insight{
  position:relative; overflow:hidden; border-radius:var(--r-tarjeta); padding:24px;
  color:#fff; display:flex; flex-direction:column; justify-content:space-between; min-height:270px;
  border:1px solid rgba(20,20,15,.06); box-shadow:var(--sombra);
  background:
    radial-gradient(72% 62% at 88% 6%, #F7DAAC 0%, rgba(247,218,172,0) 58%),
    radial-gradient(80% 72% at 102% 42%, #EE9A63 0%, rgba(238,154,99,0) 62%),
    radial-gradient(92% 80% at 4% 98%, #CFE0A6 0%, rgba(207,224,166,0) 60%),
    radial-gradient(92% 92% at -4% 8%, #7FA7E6 0%, rgba(127,167,230,0) 62%),
    linear-gradient(135deg,#6E97DE 0%,#A9B9D2 46%,#EBA972 100%);
}
.insight::after{
  content:""; position:absolute; inset:0; pointer-events:none;
  background:radial-gradient(58% 46% at 76% 20%, rgba(255,255,255,.5), rgba(255,255,255,0) 62%);
}
.insight > *{position:relative; z-index:1}
.insight .cifra{font-size:62px; text-shadow:0 2px 10px rgba(30,40,60,.22)}
.insight h3{margin:14px 0 0; font-size:19px; font-weight:600; letter-spacing:-.02em; line-height:1.3; text-shadow:0 1px 6px rgba(30,40,60,.28)}
.insight p{margin:10px 0 0; font-size:13px; line-height:1.55; color:rgba(255,255,255,.92); text-shadow:0 1px 6px rgba(30,40,60,.3)}
.insight-pill{
  display:inline-flex; align-items:center; gap:7px; align-self:flex-start;
  background:rgba(255,255,255,.72); color:#33342F; border-radius:999px;
  padding:5px 13px; font-size:12px; font-weight:500; backdrop-filter:blur(6px);
}
.insight-avance{margin-top:22px}
.insight-avance .pista-b{height:5px; border-radius:999px; background:rgba(255,255,255,.38); overflow:hidden}
.insight-avance .pista-b i{display:block; height:100%; border-radius:999px; background:#fff}
.insight-avance .txt{margin-top:9px; font-size:11.5px; color:rgba(255,255,255,.9)}

/* ---------- graficos plotly ---------- */
.plot{margin:0 -10px -10px}
.js-plotly-plot .plotly .modebar{display:none !important}

/* ---------- controles de reproduccion del grafico diario ---------- */
.repro{display:flex; align-items:center; gap:16px; margin:14px 8px 2px; flex-wrap:wrap}
.btn-repro{
  display:inline-flex; align-items:center; gap:8px; cursor:pointer;
  background:var(--blanco); border:1px solid var(--linea); border-radius:999px;
  padding:9px 18px; font:inherit; font-size:13.5px; font-weight:500; color:var(--tinta);
  box-shadow:0 1px 2px rgba(20,20,15,.06); white-space:nowrap;
  transition:background .15s, color .15s, border-color .15s;
}
.btn-repro:hover{background:#F3F3EF}
.btn-repro.activo{background:#1D1D1A; color:#fff; border-color:#1D1D1A}
.rango-dia{
  flex:1 1 220px; min-width:160px; height:22px; cursor:pointer;
  -webkit-appearance:none; appearance:none; background:transparent;
}
.rango-dia::-webkit-slider-runnable-track{height:5px; border-radius:999px; background:#E4E4DE}
.rango-dia::-webkit-slider-thumb{
  -webkit-appearance:none; appearance:none; margin-top:-7px;
  width:19px; height:19px; border-radius:50%; background:var(--blanco);
  border:1px solid var(--linea-fuerte); box-shadow:0 2px 7px rgba(20,20,15,.28);
}
.rango-dia::-moz-range-track{height:5px; border-radius:999px; background:#E4E4DE}
.rango-dia::-moz-range-thumb{
  width:19px; height:19px; border:1px solid var(--linea-fuerte); border-radius:50%;
  background:var(--blanco); box-shadow:0 2px 7px rgba(20,20,15,.28);
}
.rango-etiqueta{font-size:12.5px; color:var(--suave); min-width:96px; text-align:right; font-variant-numeric:tabular-nums}
.rango-etiqueta b{color:var(--tinta); font-weight:600}

/* ---------- tablas ---------- */
.tabla-wrap{border:1px solid var(--linea); border-radius:16px; overflow:auto; background:var(--blanco); max-height:720px}
table.tabla-detalle{border-collapse:separate; border-spacing:0; width:100%; font-size:13px}
table.tabla-detalle th{
  position:sticky; top:0; z-index:3; background:var(--blanco); text-align:right;
  font-size:11px; font-weight:600; letter-spacing:.05em; text-transform:uppercase;
  color:var(--suave); padding:13px 14px; border-bottom:1px solid var(--linea); white-space:nowrap;
}
table.tabla-detalle td{
  padding:11px 14px; text-align:right; white-space:nowrap;
  border-bottom:1px solid #F0F0EB; font-variant-numeric:tabular-nums; color:var(--tinta-2);
}
table.tabla-detalle th:first-child, table.tabla-detalle td:first-child{
  text-align:left; position:sticky; left:0; z-index:2; background:var(--blanco);
  font-weight:500; color:var(--tinta);
}
table.tabla-detalle th:first-child{z-index:4}
table.tabla-detalle tbody tr:hover td{background:#F6F8FE}
table.tabla-detalle tbody tr:hover td:first-child{background:#F6F8FE}
tr.fila-total td{background:#F3F5FB; font-weight:600; color:var(--tinta); border-top:1px solid var(--linea-fuerte)}
tr.fila-total td:first-child{background:#F3F5FB}

/* ---------- notas ---------- */
.notas{background:#FFFBF2; border:1px solid #F0E0BE; border-radius:var(--r-tarjeta); padding:20px 22px; box-shadow:var(--sombra)}
.notas-cab{display:flex; align-items:center; gap:9px; font-size:13.5px; font-weight:600; color:#8A5A12}
.notas ul{margin:12px 0 0; padding-left:20px; color:#6B4E16; font-size:13px; line-height:1.6}
.notas li + li{margin-top:8px}

.pie{margin-top:26px; text-align:center; font-size:12px; color:var(--tenue)}

/* ---------- panel de configuracion ----------
   Vive fuera del flujo del tablero: se abre por encima y no comparte pantalla
   con las cifras del corte. Aqui dentro van los nombres de campana y las notas
   de datos, que son informacion de mantenimiento, no del negocio. */
.capa-config{
  position:fixed; inset:0; z-index:200; overflow-y:auto;
  display:flex; align-items:flex-start; justify-content:center; padding:34px 18px 60px;
  background:rgba(20,20,15,.36); backdrop-filter:blur(3px);
}
.capa-config[hidden]{display:none}
body.config-abierta{overflow:hidden}
.caja-config{
  width:100%; max-width:1080px; background:var(--bg);
  border:1px solid var(--linea); border-radius:24px;
  box-shadow:0 44px 90px -34px rgba(20,20,15,.55);
}
.config-cab{
  display:flex; align-items:center; gap:18px; padding:20px 26px;
  border-bottom:1px solid var(--linea); background:var(--blanco);
  border-radius:24px 24px 0 0; position:sticky; top:0; z-index:6;
}
.config-cab h2{margin:0; font-size:23px; font-weight:600; letter-spacing:-.028em}
.config-cab .sub{margin-top:4px; font-size:12.5px; color:var(--suave)}
.config-cab .btn{margin-left:auto}
.config-cuerpo{padding:22px 26px 28px; display:flex; flex-direction:column; gap:16px}
.config-bloque{background:var(--tarjeta); border:1px solid var(--linea); border-radius:18px; padding:20px 22px}
.config-bloque h3{margin:0; font-size:16.5px; font-weight:600; letter-spacing:-.02em}
.config-bloque .desc{margin-top:6px; font-size:12.5px; color:var(--suave); line-height:1.6}
.config-bloque .desc code{
  background:#EFEFE9; border-radius:5px; padding:1px 5px; font-size:11.5px;
  font-family:ui-monospace,Consolas,"Courier New",monospace;
}
.tabla-config{width:100%; border-collapse:separate; border-spacing:0; font-size:13px; margin-top:14px}
.tabla-config th{
  text-align:left; font-size:11px; text-transform:uppercase; letter-spacing:.05em;
  color:var(--suave); font-weight:600; padding:0 12px 9px 0; white-space:nowrap;
}
.tabla-config td{padding:5px 12px 5px 0; vertical-align:middle; border-top:1px solid #EFEFEA}
.tabla-config td:last-child, .tabla-config th:last-child{padding-right:0}
.tabla-config td.clave{color:var(--tinta-2); white-space:nowrap; max-width:210px; overflow:hidden; text-overflow:ellipsis}
.campo{
  width:100%; font:inherit; font-size:13px; color:var(--tinta); background:var(--blanco);
  border:1px solid var(--linea-fuerte); border-radius:10px; padding:8px 11px;
}
.campo:focus{outline:2px solid rgba(47,111,237,.32); outline-offset:1px; border-color:var(--azul)}
.campo::placeholder{color:var(--tenue)}
.btn{
  display:inline-flex; align-items:center; gap:8px; cursor:pointer; font:inherit;
  font-size:13.5px; font-weight:500; border-radius:999px; padding:10px 18px;
  border:1px solid var(--linea-fuerte); background:var(--blanco); color:var(--tinta);
  transition:background .15s, color .15s, border-color .15s;
}
.btn:hover{background:#F3F3EF}
.btn.principal{background:#1D1D1A; color:#fff; border-color:#1D1D1A}
.btn.principal:hover{background:#333330}
.acciones-config{display:flex; align-items:center; gap:10px; flex-wrap:wrap; margin-top:16px}
.marca-guardado{font-size:12.5px; color:#15803D; opacity:0; transition:opacity .2s ease}
.marca-guardado.visible{opacity:1}
.interruptor{display:inline-flex; align-items:center; gap:10px; font-size:13.5px; cursor:pointer; margin-top:14px}
.interruptor input{width:17px; height:17px; accent-color:#1D1D1A; cursor:pointer}
.aviso-local{
  margin-top:14px; background:#FFFBF2; border:1px solid #F0E0BE; border-radius:12px;
  padding:11px 14px; font-size:12.5px; line-height:1.55; color:#6B4E16;
}
.aviso-local[hidden]{display:none}
.config-bloque .notas{box-shadow:none; margin-top:14px}
/* Dentro del panel el bloque ya se titula 'Notas de datos'; la cabecera propia
   de la tarjeta solo hace falta cuando el interruptor la manda al tablero. */
.config-bloque .notas .notas-cab{display:none}
.config-bloque .notas ul{margin-top:0}
.sin-notas{margin-top:12px; font-size:13px; color:var(--suave)}

@media (max-width:1180px){
  .s8,.s6{grid-column:span 12} .s4{grid-column:span 6} .s3{grid-column:span 6}
}
@media (max-width:760px){
  .envoltura{padding:20px 14px 56px}
  .titulo-fila h1{font-size:32px}
  .s3,.s4{grid-column:span 12}
  .cifra{font-size:44px}
  .barra-sup .interior{gap:14px; flex-wrap:wrap}
}
"""

JS_DASHBOARD = """
(function(){
  // Reproduccion dia a dia del grafico de eficiencia. Se recorta la serie con
  // Plotly.restyle en vez de usar frames + Plotly.animate: animate deja la
  // traza sin linea ni relleno (solo los puntos sueltos).
  function reproductor(){
    var sec = document.getElementById('eficiencia');
    if (!sec || !window.Plotly) return;
    var gd = sec.querySelector('.js-plotly-plot');
    var rango = sec.querySelector('.rango-dia');
    var boton = sec.querySelector('.btn-repro');
    var etiqueta = sec.querySelector('.rango-etiqueta');
    if (!gd || !gd.data || !rango || !boton) return;

    var dias = (sec.getAttribute('data-dias') || '').split('|');
    var xs = gd.data[0].x.slice();
    var ys = gd.data[0].y.slice();
    var ultimo = parseInt(rango.max, 10);
    var timer = null;
    var ICONO_PLAY = '<svg width="11" height="12" viewBox="0 0 11 12" aria-hidden="true">' +
      '<path d="M1 1.2 10 6 1 10.8Z" fill="currentColor"/></svg>';
    var ICONO_PAUSA = '<svg width="11" height="12" viewBox="0 0 11 12" aria-hidden="true">' +
      '<rect x="1.2" y="1" width="3" height="10" rx="1" fill="currentColor"/>' +
      '<rect x="6.8" y="1" width="3" height="10" rx="1" fill="currentColor"/></svg>';

    function pintar(n){
      Plotly.restyle(gd, {x: [xs.slice(0, n + 1)], y: [ys.slice(0, n + 1)]}, [0]);
      if (etiqueta) etiqueta.innerHTML = 'Día <b>' + (dias[n] || '') + '</b>';
    }
    function parar(){
      if (timer) { clearInterval(timer); timer = null; }
      boton.classList.remove('activo');
      boton.innerHTML = ICONO_PLAY + ' Reproducir';
    }
    rango.addEventListener('input', function(){ parar(); pintar(+rango.value); });
    boton.addEventListener('click', function(){
      if (timer) { parar(); return; }
      var n = (+rango.value >= ultimo) ? 0 : +rango.value;
      boton.classList.add('activo');
      boton.innerHTML = ICONO_PAUSA + ' Pausa';
      rango.value = n; pintar(n);
      timer = setInterval(function(){
        n++;
        if (n > ultimo) { parar(); return; }
        rango.value = n; pintar(n);
      }, 320);
    });
    parar();
  }
  // ---------------------------------------------------------------------
  // Menu de arriba: ilumina la seccion en la que se esta.
  //
  // La version anterior comparaba offsetTop contra un umbral fijo de 150 px
  // que no tenia nada que ver con donde aterriza el ancla, asi que al pulsar
  // un enlace quedaba iluminada la seccion ANTERIOR. Ahora la linea de corte
  // se calcula desde el alto real de la barra fija, que es tambien de donde
  // sale el scroll-margin de las secciones, y ademas el enlace pulsado manda
  // durante el desplazamiento suave.
  // ---------------------------------------------------------------------
  var barra = document.querySelector('.barra-sup');
  var enlaces = [].slice.call(document.querySelectorAll('.nav a'));
  var secciones = enlaces.map(function(a){ return document.querySelector(a.getAttribute('href')); });
  var indiceFijado = -1, fijadoHasta = 0, pedido = false;

  function altoBarra(){
    return barra ? Math.round(barra.getBoundingClientRect().height) : 58;
  }
  function medirBarra(){
    // Lo lee el CSS en scroll-margin-top: el desfase del ancla y la linea de
    // corte del menu salen los dos de la misma medida y no pueden desalinearse.
    document.documentElement.style.setProperty('--h-barra', altoBarra() + 'px');
  }
  function marcar(i){
    enlaces.forEach(function(a, j){ a.classList.toggle('activo', j === i); });
  }
  function actualizar(){
    pedido = false;
    if (indiceFijado >= 0 && Date.now() < fijadoHasta) { marcar(indiceFijado); return; }
    indiceFijado = -1;
    var linea = altoBarra() + 30;
    var activo = 0;
    for (var i = 0; i < secciones.length; i++){
      if (secciones[i] && secciones[i].getBoundingClientRect().top - linea <= 1) activo = i;
    }
    // Al final de la pagina manda siempre la ultima seccion: si es corta nunca
    // llega a cruzar la linea de corte y el menu se quedaria en la anterior.
    var doc = document.documentElement;
    if (window.pageYOffset + window.innerHeight >= doc.scrollHeight - 2){
      for (var k = secciones.length - 1; k >= 0; k--){
        if (secciones[k]) { activo = k; break; }
      }
    }
    marcar(activo);
  }
  function pedirActualizar(){
    if (pedido) return;
    pedido = true;
    // El temporizador no sobra: requestAnimationFrame solo corre cuando el
    // navegador dibuja un fotograma, y si no llega ninguno el menu se queda
    // clavado en la seccion en la que estaba. actualizar() es idempotente, asi
    // que no molesta que lleguen los dos.
    if (window.requestAnimationFrame) window.requestAnimationFrame(actualizar);
    setTimeout(actualizar, 60);
  }
  enlaces.forEach(function(a, i){
    a.addEventListener('click', function(){
      // Mientras dura el desplazamiento suave manda el enlace pulsado: sin
      // esto el menu parpadea con las secciones que va cruzando por el camino.
      indiceFijado = i; fijadoHasta = Date.now() + 900; marcar(i);
    });
  });
  ['wheel', 'touchmove', 'keydown'].forEach(function(ev){
    window.addEventListener(ev, function(){ indiceFijado = -1; }, {passive:true});
  });
  window.addEventListener('scroll', pedirActualizar, {passive:true});
  window.addEventListener('resize', function(){ medirBarra(); pedirActualizar(); });
  medirBarra();
  actualizar();
  window.addEventListener('load', function(){ medirBarra(); actualizar(); });

  // Cerrar la nota de una tarjeta al abrir otra o al pulsar fuera.
  document.addEventListener('click', function(ev){
    document.querySelectorAll('details.ayuda[open]').forEach(function(d){
      if (!d.contains(ev.target)) d.removeAttribute('open');
    });
  });

  // ---------------------------------------------------------------------
  // Configuracion: nombres de campana, colas y donde van las notas de datos.
  //
  // Todo sale del propio DOM (cada campo lleva su valor generado en
  // data-inicial), asi que el panel no depende de ningun dato incrustado
  // aparte. Lo que se escribe aqui se ve al momento y se recuerda en este
  // navegador; el archivo configuracion.json es lo que lo hace permanente
  // para todos, porque lo lee dashboard_kpi.py en la siguiente generacion.
  // ---------------------------------------------------------------------
  function iniciarConfig(){
    var capa = document.getElementById('configuracion');
    if (!capa) return;
    var campos = [].slice.call(capa.querySelectorAll('.campo[data-clave]'));
    var chkNotas = document.getElementById('chk-notas');
    var aviso = document.getElementById('marca-guardado');
    var LS = 'kpi-sales-config';

    function leerLS(){
      try { return JSON.parse(window.localStorage.getItem(LS) || 'null'); }
      catch (e) { return null; }
    }
    function escribirLS(obj){
      try { window.localStorage.setItem(LS, JSON.stringify(obj)); return true; }
      catch (e) { return false; }
    }

    function recolectar(){
      var nombres = {}, colas = {};
      campos.forEach(function(inp){
        var clave = inp.getAttribute('data-clave');
        var val = (inp.value || '').trim();
        if (inp.getAttribute('data-tipo') === 'nombre'){
          if (val && val !== clave) nombres[clave] = val;
        } else {
          colas[clave] = val ? val.split(',').map(function(s){ return s.trim(); })
                                  .filter(function(s){ return s.length; }) : [];
        }
      });
      return {
        nombres_campanas: nombres,
        mapeo_colas: colas,
        mostrar_notas_en_dashboard: !!(chkNotas && chkNotas.checked)
      };
    }
    function aplicarNombres(nombres){
      document.querySelectorAll('[data-campana]').forEach(function(el){
        var k = el.getAttribute('data-campana');
        el.textContent = nombres[k] || k;
      });
      if (!window.Plotly) return;
      document.querySelectorAll('.plot[data-campanas]').forEach(function(cont){
        var gd = cont.querySelector('.js-plotly-plot');
        if (!gd || !gd.data || !gd.data.length) return;
        var vis = cont.getAttribute('data-campanas').split('|').map(function(k){
          return nombres[k] || k;
        });
        Plotly.restyle(gd, {x: gd.data.map(function(){ return vis.slice(); })});
      });
    }
    function moverNotas(alTablero){
      var notas = document.getElementById('bloque-notas');
      var destino = document.getElementById(alTablero ? 'notas-en-tablero' : 'notas-en-config');
      if (notas && destino && notas.parentNode !== destino) destino.appendChild(notas);
    }
    function decir(txt){
      if (!aviso) return;
      aviso.textContent = txt;
      aviso.classList.add('visible');
      clearTimeout(aviso._t);
      aviso._t = setTimeout(function(){ aviso.classList.remove('visible'); }, 6000);
    }
    function avisoLocal(visible){
      var el = document.getElementById('aviso-local');
      if (el) el.hidden = !visible;
    }
    function aplicar(recordar){
      var estado = recolectar();
      aplicarNombres(estado.nombres_campanas);
      moverNotas(estado.mostrar_notas_en_dashboard);
      if (!recordar) return;
      avisoLocal(true);
      decir(escribirLS(estado)
        ? 'Aplicado y recordado en este navegador. Guarda el archivo para que valga también al regenerar.'
        : 'Aplicado en esta pantalla. Este navegador no deja recordarlo: guarda el archivo.');
    }
    function texto(){ return JSON.stringify(recolectar(), null, 2); }

    function bajarArchivo(contenido, nombre, tipo){
      var blob = new Blob([contenido], {type: tipo});
      var url = URL.createObjectURL(blob);
      var a = document.createElement('a');
      a.href = url; a.download = nombre;
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
      setTimeout(function(){ URL.revokeObjectURL(url); }, 4000);
    }
    function descargar(contenido){
      bajarArchivo(contenido, 'configuracion.json', 'application/json');
      decir('Descargado configuracion.json: muévelo a la carpeta del dashboard, junto a dashboard_kpi.py.');
    }
    function guardarArchivo(){
      var contenido = texto();
      // El dialogo de guardar deja escribir directamente en la carpeta del
      // dashboard. Donde no exista (o si se cancela), se cae a la descarga.
      if (window.showSaveFilePicker){
        window.showSaveFilePicker({
          suggestedName: 'configuracion.json',
          types: [{description: 'Configuración del dashboard', accept: {'application/json': ['.json']}}]
        }).then(function(h){
          return h.createWritable().then(function(w){
            return w.write(contenido).then(function(){ return w.close(); });
          });
        }).then(function(){
          decir('Guardado. Vuelve a generar el dashboard y los nombres quedan fijos para todos.');
        }).catch(function(err){
          if (err && (err.name === 'AbortError' || err.name === 'NotAllowedError')) return;
          descargar(contenido);
        });
        return;
      }
      descargar(contenido);
    }
    function copiar(){
      var contenido = texto();
      function aMano(){
        var ta = document.createElement('textarea');
        ta.value = contenido; ta.setAttribute('readonly', '');
        ta.style.position = 'fixed'; ta.style.opacity = '0';
        document.body.appendChild(ta); ta.select();
        try { document.execCommand('copy'); decir('JSON copiado al portapapeles.'); }
        catch (e) { decir('No se pudo copiar automáticamente.'); }
        document.body.removeChild(ta);
      }
      if (navigator.clipboard && navigator.clipboard.writeText){
        navigator.clipboard.writeText(contenido).then(function(){
          decir('JSON copiado al portapapeles.');
        }, aMano);
      } else { aMano(); }
    }
    function restablecer(){
      try { window.localStorage.removeItem(LS); } catch (e) {}
      campos.forEach(function(inp){ inp.value = inp.getAttribute('data-inicial') || ''; });
      if (chkNotas) chkNotas.checked = chkNotas.getAttribute('data-inicial') === '1';
      aplicar(false);
      avisoLocal(false);
      decir('Restablecido a lo que trae el dashboard generado.');
    }

    // -------------------------------------------------------------------
    // Copia para compartir: el mismo tablero, sin la parte editable.
    //
    // Se arma clonando la pagina viva, asi que sale exactamente lo que se ve
    // (incluidos los nombres cambiados). Los graficos se vacian antes de
    // serializar: el <script> que los dibuja sigue dentro y los vuelve a
    // pintar al abrir la copia, en vez de guardar el SVG ya dibujado.
    // -------------------------------------------------------------------
    var CONFIG_PLOTLY = {displaylogo: false, responsive: true, displayModeBar: false};

    function copiaParaCompartir(conNotas){
      var clon = document.documentElement.cloneNode(true);
      function uno(sel){ return clon.querySelector(sel); }
      function varios(sel){ return [].slice.call(clon.querySelectorAll(sel)); }
      function json(o){ return JSON.stringify(o).replace(/<\\//g, '<\\\\/'); }

      var nombres = recolectar().nombres_campanas;
      var hayRenombres = Object.keys(nombres).length > 0;

      varios('.js-plotly-plot').forEach(function(div){
        div.innerHTML = '';
        // Si se renombro alguna campana, el <script> original todavia lleva
        // los nombres viejos en los ejes. Se reescribe con los datos que
        // tiene el grafico vivo, que ya estan renombrados. Al de eficiencia
        // diaria no se le toca: no lleva nombres de campana y el reproductor
        // recorta su serie, asi que copiarla podria dejarla a medias.
        if (!hayRenombres || !div.closest('.plot[data-campanas]')) return;
        var vivo = document.getElementById(div.id);
        var guion = varios('script').filter(function(s){
          return s.textContent.indexOf('"' + div.id + '"') >= 0;
        })[0];
        if (!vivo || !vivo.data || !guion) return;
        guion.textContent = 'window.PLOTLYENV=window.PLOTLYENV||{};Plotly.newPlot("' +
          div.id + '",' + json(vivo.data) + ',' + json(vivo.layout) + ',' + json(CONFIG_PLOTLY) + ');';
      });

      // La copia no se puede renombrar: fuera los enganches del panel.
      varios('[data-campana]').forEach(function(el){ el.removeAttribute('data-campana'); });
      varios('[data-campanas]').forEach(function(el){ el.removeAttribute('data-campanas'); });

      // La copia se hace con el panel abierto (el boton esta dentro), y el
      // estado abierto se clona con el resto: hay que cerrarlo a mano o la
      // copia arrancaria con el panel encima y sin poder desplazarse.
      var panel = uno('#configuracion');
      if (panel) panel.hidden = true;
      var cuerpo = uno('body');
      if (cuerpo) cuerpo.classList.remove('config-abierta');

      var boton = uno('#btn-config');
      var notas = document.getElementById('bloque-notas');
      var notasEnTablero = !!notas && notas.parentNode.id === 'notas-en-tablero';
      var puedeQuedarseNotas = conNotas && !!uno('#bloque-notas') && !notasEnTablero;

      if (panel && boton && puedeQuedarseNotas){
        // Del panel solo sobrevive el bloque de notas, ya sin nada editable.
        varios('#configuracion .config-bloque').forEach(function(b){
          if (!b.querySelector('#bloque-notas')) b.remove();
        });
        var interruptor = uno('#chk-notas');
        if (interruptor && interruptor.parentNode) interruptor.parentNode.remove();
        var titulo = uno('#configuracion h2');
        if (titulo) titulo.textContent = 'Notas de datos';
        var sub = uno('#configuracion .sub');
        if (sub) sub.textContent = 'Rarezas del Excel de este corte. No cambian ninguna cifra: se cuentan para que nadie las lea al revés.';
        var cabBloque = uno('#configuracion .config-bloque h3');
        if (cabBloque) cabBloque.remove();
        var descBloque = uno('#configuracion .config-bloque .desc');
        if (descBloque) descBloque.remove();
        var icono = document.getElementById('icono-notas');
        var globo = boton.querySelector('.globo');
        boton.innerHTML = (icono ? icono.innerHTML : '') + '<span>Notas de datos</span>' +
                          (globo ? globo.outerHTML : '');
        boton.setAttribute('title', 'Rarezas del Excel de este corte');
      } else if (panel && boton){
        panel.remove();
        boton.remove();
      }

      return '<!doctype html>\\n' + clon.outerHTML;
    }

    function descargarCopia(){
      var chk = document.getElementById('chk-notas-copia');
      var aviso = document.getElementById('marca-copia');
      var html = copiaParaCompartir(!chk || chk.checked);
      var nombre = (document.title || 'Marketing KPI').replace(/[\\\\/:*?"<>|]/g, '-') + '.html';
      bajarArchivo(html, nombre, 'text/html;charset=utf-8');
      if (aviso){
        aviso.textContent = 'Descargado «' + nombre + '» (' +
          (Math.round(html.length / 104857.6) / 10) + ' MB). Está en tu carpeta de Descargas: ' +
          'mándalo por correo y se abre con doble clic.';
        aviso.classList.add('visible');
        clearTimeout(aviso._t);
        aviso._t = setTimeout(function(){ aviso.classList.remove('visible'); }, 9000);
      }
    }

    function abrirPanel(){
      capa.hidden = false;
      document.body.classList.add('config-abierta');
    }
    function cerrarPanel(){
      capa.hidden = true;
      document.body.classList.remove('config-abierta');
      pedirActualizar();
    }
    var btnAbrir = document.getElementById('btn-config');
    if (btnAbrir) btnAbrir.addEventListener('click', abrirPanel);
    capa.querySelectorAll('[data-cerrar]').forEach(function(b){
      b.addEventListener('click', cerrarPanel);
    });
    capa.addEventListener('click', function(ev){ if (ev.target === capa) cerrarPanel(); });
    document.addEventListener('keydown', function(ev){
      if (ev.key === 'Escape' && !capa.hidden) cerrarPanel();
    });

    var btn;
    if ((btn = document.getElementById('btn-aplicar'))) btn.addEventListener('click', function(){ aplicar(true); });
    if ((btn = document.getElementById('btn-archivo'))) btn.addEventListener('click', guardarArchivo);
    if ((btn = document.getElementById('btn-copiar'))) btn.addEventListener('click', copiar);
    if ((btn = document.getElementById('btn-restablecer'))) btn.addEventListener('click', restablecer);
    if ((btn = document.getElementById('btn-compartir'))) btn.addEventListener('click', descargarCopia);
    if (chkNotas) chkNotas.addEventListener('change', function(){ moverNotas(chkNotas.checked); });

    // Lo guardado en este navegador se vuelve a pintar al abrir el tablero
    // recien generado: si no, cada regeneracion borraria los cambios de nombre
    // hasta que alguien se acordara de mover el archivo a su sitio.
    var guardado = leerLS();
    if (guardado){
      avisoLocal(true);
      campos.forEach(function(inp){
        var clave = inp.getAttribute('data-clave');
        if (inp.getAttribute('data-tipo') === 'nombre'){
          var n = guardado.nombres_campanas && guardado.nombres_campanas[clave];
          if (n) inp.value = n;
        } else {
          var c = guardado.mapeo_colas && guardado.mapeo_colas[clave];
          if (c) inp.value = c.join(', ');
        }
      });
      if (chkNotas && typeof guardado.mostrar_notas_en_dashboard === 'boolean'){
        chkNotas.checked = guardado.mostrar_notas_en_dashboard;
      }
      aplicar(false);
    }
    if (location.hash === '#configuracion') abrirPanel();
  }

  if (document.readyState === 'complete') { reproductor(); iniciarConfig(); }
  else window.addEventListener('load', function(){ reproductor(); iniciarConfig(); });
})();
"""


def _esc(txt):
    """Escapa texto para meterlo en un atributo HTML. Los nombres de campana
    salen del Excel y acaban dentro de value="..." y data-clave="..."."""
    return (str(txt).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _mezclar(hex_color, hex_fondo, t):
    """Mezcla dos colores hex. t=0 devuelve el fondo, t=1 el color."""
    def _c(h):
        h = h.lstrip("#")
        return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))
    a, b = _c(hex_color), _c(hex_fondo)
    return "#%02x%02x%02x" % tuple(round(b[i] + (a[i] - b[i]) * t) for i in range(3))


def _escala_bonita(maximo, divisiones=4):
    """Devuelve (tope, valores) para un eje Y con numeros redondos.
    El tope siempre queda por encima del maximo real, nunca lo recorta."""
    if not maximo or maximo <= 0:
        return 1, []
    bruto = maximo / divisiones
    base = 10 ** math.floor(math.log10(bruto))
    paso = base
    for m in (1, 2, 2.5, 5, 10):
        paso = m * base
        if paso >= bruto:
            break
    tope = paso * divisiones
    while tope < maximo:
        tope += paso
    n = int(round(tope / paso))
    valores = [paso * i for i in range(1, n + 1)]
    return tope, valores


_MIME_LOGO = {
    ".svg": "image/svg+xml", ".png": "image/png", ".webp": "image/webp",
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
}


def _svg_marca():
    """Logo de la cabecera.

    Si hay un archivo de logo en la carpeta base se usa ese, incrustado en
    base64: asi el dashboard se puede seguir mandando por correo o dejar en una
    carpeta de red como un unico .html, sin imagenes sueltas al lado. Si no hay
    ninguno, se dibuja el emblema vectorial de reserva."""
    for nombre in ARCHIVOS_LOGO:
        ruta = os.path.join(CARPETA_BASE, nombre)
        if not os.path.isfile(ruta):
            continue
        try:
            with open(ruta, "rb") as fh:
                datos = fh.read()
        except OSError:
            continue
        mime = _MIME_LOGO[os.path.splitext(nombre)[1].lower()]
        b64 = base64.b64encode(datos).decode("ascii")
        return f'<img class="logo" src="data:{mime};base64,{b64}" alt="" aria-hidden="true">'
    return _svg_marca_reserva()


def _svg_marca_reserva():
    """Emblema de Alliance Technology Group dibujado en vectores: el anillo de
    nodos naranjas alrededor de un nucleo. Es el que se usa mientras no haya un
    archivo de logo en la carpeta."""
    # (cx, cy, color) de los seis nodos del anillo, empezando arriba y girando
    # en el sentido de las agujas del reloj. El tono va del ambar al rojo.
    nodos = (
        (16.0, 6.4, "#F4A31B"), (24.3, 11.2, "#EF8A1D"), (24.3, 20.8, "#E4671F"),
        (16.0, 25.6, "#D6471E"), (7.7, 20.8, "#C6371F"), (7.7, 11.2, "#E07A20"),
    )
    brazos = "".join(
        f'<line x1="16" y1="16" x2="{cx}" y2="{cy}" stroke="{color}" '
        f'stroke-width="2.6" stroke-linecap="round" opacity=".85"/>'
        for cx, cy, color in nodos
    )
    anillo = "".join(
        f'<circle cx="{cx}" cy="{cy}" r="4" fill="{color}"/>'
        for cx, cy, color in nodos
    )
    return (
        '<svg class="logo" width="32" height="32" viewBox="0 0 32 32" fill="none" '
        'aria-hidden="true">'
        + brazos + anillo
        + '<circle cx="16" cy="16" r="4.4" fill="#fff"/>'
        '<circle cx="16" cy="16" r="3.2" fill="#E4671F"/>'
        '</svg>'
    )


def _icono_calendario():
    return ('<svg width="15" height="15" viewBox="0 0 16 16" fill="none" stroke="currentColor" '
            'stroke-width="1.4" aria-hidden="true"><rect x="2" y="3.2" width="12" height="10.5" rx="2.2"/>'
            '<path d="M2 6.6h12M5.4 1.9v2.3M10.6 1.9v2.3" stroke-linecap="round"/></svg>')


def _icono_reloj():
    return ('<svg width="15" height="15" viewBox="0 0 16 16" fill="none" stroke="currentColor" '
            'stroke-width="1.4" aria-hidden="true"><circle cx="8" cy="8" r="6.2"/>'
            '<path d="M8 4.6V8l2.4 1.5" stroke-linecap="round"/></svg>')


def _icono_hoja():
    return ('<svg width="15" height="15" viewBox="0 0 16 16" fill="none" stroke="currentColor" '
            'stroke-width="1.4" aria-hidden="true"><path d="M4 1.9h5l3 3v9.2H4z" stroke-linejoin="round"/>'
            '<path d="M9 1.9v3h3" stroke-linejoin="round"/></svg>')


def _icono_engranaje():
    return ('<svg width="15" height="15" viewBox="0 0 16 16" fill="none" stroke="currentColor" '
            'stroke-width="1.4" aria-hidden="true"><circle cx="8" cy="8" r="2.3"/>'
            '<path d="M8 1.6v1.7M8 12.7v1.7M14.4 8h-1.7M3.3 8H1.6M12.5 3.5l-1.2 1.2M4.7 11.3'
            'l-1.2 1.2M12.5 12.5l-1.2-1.2M4.7 4.7 3.5 3.5" stroke-linecap="round"/></svg>')


def _icono_bombilla():
    return ('<svg width="13" height="13" viewBox="0 0 16 16" fill="none" stroke="currentColor" '
            'stroke-width="1.5" aria-hidden="true"><path d="M6.2 12.6h3.6M6.6 14.4h2.8" stroke-linecap="round"/>'
            '<path d="M8 1.8a4.3 4.3 0 0 0-2.5 7.8v1.2h5V9.6A4.3 4.3 0 0 0 8 1.8Z" stroke-linejoin="round"/></svg>')


def _cabecera_tarjeta(titulo, ayuda, pie="", extra=""):
    """Cabecera de tarjeta: titulo + boton redondo que abre la nota de la
    tarjeta (que mide y de donde sale el dato). 'extra' se coloca a la
    izquierda del boton (por ejemplo, una insignia de tendencia)."""
    sub = f'<div class="pie-cab">{pie}</div>' if pie else ""
    return (
        '<div class="tarjeta-cab"><div><h2>' + titulo + '</h2>' + sub + '</div>'
        '<div class="cab-der">' + extra +
        '<details class="ayuda"><summary title="Sobre esta tarjeta">···</summary>'
        '<div class="ayuda-panel">' + ayuda + '</div></details></div></div>'
    )


def _bloque_columnas(items, idx_foco, ticks):
    """Columnas al estilo del panel de referencia: etiqueta + cifra grande
    arriba, barra tramada abajo y una columna destacada en solido.

    items: lista de dicts con clave, label, valor_fmt, pct (0-100) y tip.
    ticks: lista de (pct_desde_abajo, etiqueta) para el eje Y.
    """
    escala = "".join(f'<span style="bottom:{p}%">{t}</span>' for p, t in ticks)
    lineas = "".join(f'<i style="bottom:{p}%"></i>' for p, _t in ticks)

    cols = []
    for i, it in enumerate(items):
        foco = " foco" if i == idx_foco else ""
        pct = max(1.0, min(100.0, it["pct"]))
        marca = "" if i == idx_foco else (
            f'<div class="col-marca" style="bottom:calc({pct}% + 9px)"></div>'
        )
        cols.append(
            f'<div class="col{foco}">'
            f'<div class="col-cab"><div class="col-label" data-campana="{it["clave"]}">{it["label"]}</div>'
            f'<div class="col-valor">{it["valor_fmt"]}</div></div>'
            f'<div class="col-zona">'
            f'<div class="col-tip">{it["tip"]}</div>{marca}'
            f'<div class="col-barra" style="height:{pct}%"></div>'
            f'</div></div>'
        )

    return (
        '<div class="cols">'
        f'<div class="cols-eje"><div class="hueco"></div><div class="escala">{escala}</div></div>'
        f'<div class="cols-cuerpo"><div class="cols-lineas">{lineas}</div>{"".join(cols)}</div>'
        '</div>'
    )


def _barras_reparto(filas):
    """Lista de barras horizontales con tramado diagonal.
    filas: lista de (clave, etiqueta, valor_fmt, pct 0-100, color hex)."""
    if not filas:
        return ""
    out = []
    for clave, etiqueta, valor_fmt, pct, color in filas:
        ancho = max(2.0, min(100.0, pct))
        out.append(
            '<div class="fila-barra">'
            f'<div class="fila-barra-cab"><span data-campana="{clave}">{etiqueta}</span><b>{valor_fmt}</b></div>'
            f'<div class="pista"><div class="relleno" style="width:{ancho}%;background-color:{color}"></div></div>'
            '</div>'
        )
    return '<div class="reparto">' + "".join(out) + '</div>'


def _area_escalonada(valores, color=C_ROSA, alto=150, etiqueta_valor="llamadas"):
    """Serie diaria dibujada como area escalonada con relleno de rayas
    verticales (el mismo gesto grafico que la tarjeta de retencion de la
    referencia).

    El trazo solo cuenta la forma del mes, y con un dia pico que triplica al
    resto la forma sale casi plana: se ve donde esta el pico, pero no cuanto
    vale ningun dia. Por eso encima del SVG van dos capas mas, las dos en
    porcentaje del alto (el SVG se estira con preserveAspectRatio="none" y su
    alto real depende de lo que crezca la tarjeta):

      - una rejilla con dos o tres cifras redondas, para tener referencia;
      - una columna invisible por dia que, al pasar el raton, marca el dia con
        una guia vertical y un punto sobre la linea y saca el globo con la
        fecha y la cifra, igual que las tarjetas de contestadas y perdidas.

    Devuelve el bloque .mini-chart entero. valores: lista de (fecha, valor)."""
    if not valores:
        return ""
    ancho = 320.0
    n = len(valores)
    w = ancho / n
    maximo = max(v for _f, v in valores)
    tope = maximo or 1
    util = alto - 14

    def _y(v):
        return alto - (v / tope) * util

    def _pct(v):
        """Alto del valor contado desde abajo, en % del alto del bloque."""
        return (v / tope) * util / alto * 100

    pts = []
    for i, (_f, v) in enumerate(valores):
        pts.append((i * w, _y(v), (i + 1) * w))

    d_linea = f"M{pts[0][0]:.2f},{pts[0][1]:.2f}"
    for x0, y, x1 in pts:
        d_linea += f" L{x0:.2f},{y:.2f} L{x1:.2f},{y:.2f}"
    d_area = d_linea + f" L{ancho:.2f},{alto} L0,{alto} Z"

    idx_pico = max(range(n), key=lambda i: valores[i][1])
    x_pico = (idx_pico + 0.5) * w / ancho * 100

    # Cifras de la rejilla: las del eje redondo que quedan por debajo del
    # maximo real. El trazo sigue llegando arriba del todo, la rejilla no
    # recorta nada.
    _t, escalones = _escala_bonita(maximo)
    reja = "".join(
        f'<i style="bottom:{_pct(t):.2f}%"><b>{fmt_num(t)}</b></i>'
        for t in escalones if t < maximo * 0.94
    )

    dias = "".join(
        f'<div class="ah-col" style="--h:{_pct(v):.2f}%">'
        f'<span class="ah-tip">{fmt_fecha_es(f)} &nbsp;·&nbsp; '
        f'<b>{fmt_num(v)}</b> {etiqueta_valor}</span>'
        f'<i class="ah-guia"></i><i class="ah-punto"></i></div>'
        for i, (f, v) in enumerate(valores)
    )

    return (
        f'<div class="mini-chart" style="--c:{color}">'
        f'<div class="pico-flotante" style="left:{x_pico:.1f}%">{fmt_num(maximo)}</div>'
        f'<svg viewBox="0 0 {ancho:.0f} {alto}" height="{alto}" preserveAspectRatio="none" role="img">'
        f'<defs><pattern id="ray" width="7" height="7" patternUnits="userSpaceOnUse">'
        f'<line x1="1" y1="0" x2="1" y2="7" stroke="{color}" stroke-width="1" opacity=".3"/>'
        f'</pattern></defs>'
        f'<path d="{d_area}" fill="url(#ray)"/>'
        f'<path d="{d_linea}" fill="none" stroke="{color}" stroke-width="2.4" '
        f'stroke-linejoin="round" vector-effect="non-scaling-stroke"/>'
        f'</svg>'
        f'<div class="area-reja">{reja}</div>'
        f'<div class="area-dias">{dias}</div>'
        f'</div>'
    )


def _matriz_puntos(valores, color, filas_max=5, etiqueta_valor="llamadas"):
    """Distribucion diaria como matriz de puntos. El dia pico va en color
    pleno y el resto en un tono claro del mismo color.

    Cada columna es un dia y lleva su propio globo con la fecha y la cifra: la
    matriz sola dice la forma del mes, pero no el dato del dia concreto que se
    esta mirando."""
    if not valores:
        return ""
    tope = max(v for _f, v in valores) or 1
    idx_pico = max(range(len(valores)), key=lambda i: valores[i][1])
    tenue = _mezclar(color, "#E9E9E3", 0.30)

    cols = []
    for i, (f, v) in enumerate(valores):
        n = 0 if v <= 0 else max(1, int(round(v / tope * filas_max)))
        pico = " pico" if i == idx_pico else ""
        tip = (f'<span class="pt-tip">{fmt_fecha_es(f)} &nbsp;·&nbsp; '
               f'<b>{fmt_num(v)}</b> {etiqueta_valor}</span>')
        cols.append(f'<div class="pt-col{pico}">{tip}' + ('<i></i>' * n) + '</div>')
    return (f'<div class="puntos" style="--c:{color};--c-tenue:{tenue}">'
            + "".join(cols) + '</div>')


def _fila_config(clave, colas=None):
    """Fila de la tabla de configuracion: la clave del Excel (fija), el nombre
    a mostrar y, si la campana tiene llamadas, sus colas.

    data-inicial guarda lo que trae el dashboard recien generado, que es a lo
    que vuelve el boton Restablecer."""
    k = _esc(clave)
    nombre = _esc(nombre_visible(clave))
    fila = (
        f'<tr><td class="clave" title="{k}">{k}</td>'
        f'<td><input class="campo" type="text" data-tipo="nombre" data-clave="{k}" '
        f'data-inicial="{nombre}" value="{nombre}" placeholder="{k}" '
        f'aria-label="Nombre a mostrar de {k}"></td>'
    )
    if colas is not None:
        txt = _esc(", ".join(colas))
        fila += (
            f'<td><input class="campo" type="text" data-tipo="colas" data-clave="{k}" '
            f'data-inicial="{txt}" value="{txt}" placeholder="sin llamadas" '
            f'aria-label="Colas de {k}"></td>'
        )
    return fila + "</tr>"


def _panel_configuracion(claves_llamadas, claves_leads, bloque_notas, notas_en_tablero):
    """Panel de Configuracion: se abre encima del tablero y nunca comparte
    pantalla con los datos del corte.

    Es lo unico del HTML que se puede editar. Los cambios se ven al momento y
    se recuerdan en el navegador, pero lo que los hace permanentes (y comunes a
    todos) es el archivo configuracion.json que se guarda desde aqui: es el que
    lee dashboard_kpi.py en la siguiente generacion."""
    filas_llamadas = "".join(
        _fila_config(clave, colas_de_campana(clave)) for clave in claves_llamadas
    )
    filas_leads = "".join(_fila_config(clave) for clave in claves_leads)

    bloque_leads = ""
    if filas_leads:
        bloque_leads = (
            '<div class="config-bloque">'
            '<h3>Campañas de leads</h3>'
            '<div class="desc">Las de la tabla de leads (META). Solo cambian de nombre: '
            'no tienen colas de llamadas asociadas.</div>'
            '<table class="tabla-config">'
            '<thead><tr><th>Campaña en el Excel</th><th>Nombre a mostrar</th></tr></thead>'
            f'<tbody>{filas_leads}</tbody></table></div>'
        )

    dentro = " checked" if notas_en_tablero else ""
    inicial = "1" if notas_en_tablero else "0"
    contenido_notas = "" if notas_en_tablero else bloque_notas
    sin_notas = "" if bloque_notas else (
        '<div class="sin-notas">Este corte no tiene notas: el Excel vino limpio.</div>'
    )

    return f"""
<div class="capa-config" id="configuracion" hidden>
  <div class="caja-config">
    <div class="config-cab">
      <div>
        <h2>Configuración</h2>
        <div class="sub">Nombres, colas y notas de datos. Nada de lo que hay aquí cambia una cifra del corte.</div>
      </div>
      <button type="button" class="btn" data-cerrar>Volver al tablero</button>
    </div>

    <div class="config-cuerpo">

      <div class="config-bloque">
        <h3>Nombres de las campañas</h3>
        <div class="desc">
          La primera columna es el nombre tal cual viene en la hoja '{HOJA_PRINT}'. Es la clave con
          la que se leen los datos, así que no se toca.<br>
          <b>Nombre a mostrar</b>: lo único que cambia en el tablero. Renombrar es seguro: la lectura
          del Excel sigue usando la clave.<br>
          <b>Colas en {HOJA_RAW_LLAMADAS}</b>: los nombres de cola que suman las llamadas de esa
          campaña, separados por comas. Cuando en el Excel cambia el nombre de una cola, se corrige
          aquí y deja de hacer falta editar el código. Vacío significa que la campaña no tiene llamadas.
          Puedes dejar el nombre viejo y añadir el nuevo: se admiten varios y valen los dos.
        </div>
        <table class="tabla-config">
          <thead><tr><th>Campaña en el Excel</th><th>Nombre a mostrar</th>
          <th>Colas en {HOJA_RAW_LLAMADAS}</th></tr></thead>
          <tbody>{filas_llamadas}</tbody>
        </table>
      </div>

      {bloque_leads}

      <div class="config-bloque">
        <h3>Notas de datos</h3>
        <div class="desc">
          Rarezas del Excel de este corte que conviene conocer antes de sacar conclusiones. Viven
          aquí para no mezclarlas con los números del tablero.
        </div>
        <div id="notas-en-config">{contenido_notas}</div>
        {sin_notas}
        <label class="interruptor">
          <input type="checkbox" id="chk-notas" data-inicial="{inicial}"{dentro}>
          Mostrar también las notas en el tablero, arriba del todo
        </label>
      </div>

      <div class="config-bloque">
        <h3>Guardar los cambios</h3>
        <div class="desc">
          <b>Aplicar ahora</b> los deja a la vista al momento y los recuerda en este navegador,
          pero solo aquí: quien abra el archivo en otro sitio los seguirá viendo como estaban.<br>
          <b>Guardar {NOMBRE_CONFIG}</b> es lo que los hace definitivos. Deja el archivo en la
          carpeta del dashboard, al lado de <code>dashboard_kpi.py</code>: a partir de la siguiente
          generación, los nombres y las colas salen de ahí para todo el mundo.
        </div>
        <div class="acciones-config">
          <button type="button" class="btn principal" id="btn-aplicar">Aplicar ahora</button>
          <button type="button" class="btn" id="btn-archivo">Guardar {NOMBRE_CONFIG}</button>
          <button type="button" class="btn" id="btn-copiar">Copiar JSON</button>
          <button type="button" class="btn" id="btn-restablecer">Restablecer</button>
          <span class="marca-guardado" id="marca-guardado"></span>
        </div>
        <div class="aviso-local" id="aviso-local" hidden>
          Ahora mismo estás viendo cambios guardados <b>solo en este navegador</b>, no en el archivo
          generado: tus compañeros ven los nombres originales. Guarda {NOMBRE_CONFIG} para que valgan
          para todos, o pulsa Restablecer para volver a lo que trae el dashboard.
        </div>
      </div>

      <div class="config-bloque">
        <h3>Enviar el tablero a otras personas</h3>
        <div class="desc">
          Descarga una copia del tablero <b>tal como lo estás viendo ahora</b>, pero sin esta
          sección de Configuración: quien la reciba solo ve los datos, y no puede tocar nombres ni
          colas sin querer. Es un único archivo que se abre con doble clic — sin Python, sin
          librerías, sin internet y sin instalar nada.
        </div>
        <label class="interruptor">
          <input type="checkbox" id="chk-notas-copia" checked>
          Incluir las notas de datos, en un botón «Notas de datos» aparte
        </label>
        <div class="acciones-config">
          <button type="button" class="btn principal" id="btn-compartir">Descargar copia para compartir</button>
          <span class="marca-guardado" id="marca-copia"></span>
        </div>
      </div>

      <span id="icono-notas" hidden>{_icono_bombilla()}</span>

    </div>
  </div>
</div>
"""


def _estilo_plotly(fig, altura, margenes=None):
    """Deja los graficos de Plotly con la misma tipografia, colores y
    limpieza que el resto del tablero: sin fondo propio, sin titulo (lo pone
    la cabecera de la tarjeta) y con rejilla apenas visible."""
    fig.update_layout(
        height=altura,
        title=None,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family=FUENTE, size=12.5, color=C_SUAVE),
        margin=margenes or dict(l=52, r=18, t=18, b=52),
        hoverlabel=dict(bgcolor="#FFFFFF", bordercolor=C_LINEA,
                        font=dict(family=FUENTE, size=12.5, color=C_TINTA)),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
                    font=dict(size=12.5, color=C_SUAVE), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(showgrid=False, zeroline=False, linecolor=C_LINEA,
                   ticks="outside", tickcolor=C_LINEA, ticklen=4,
                   tickfont=dict(size=11.5, color=C_TENUE), title=None),
        yaxis=dict(showgrid=True, gridcolor="#EBEBE5", zeroline=False, showline=False,
                   tickfont=dict(size=11.5, color=C_TENUE), title=None),
    )
    return fig


def construir_html(contexto):
    import plotly.graph_objects as go
    from plotly.offline import plot as plotly_plot

    datos_print = contexto["datos_print"]
    dias_del_mes = contexto["dias_del_mes"]
    serie = contexto["serie"]
    fecha_corte = contexto["fecha_corte"]
    ruta_excel = contexto["ruta_excel"]
    hora_generacion = contexto["hora_generacion"]
    avisos = contexto["avisos"]
    campanas_excluidas_serie = contexto["campanas_excluidas_serie"]
    costo_corte = contexto["costo_corte"]

    tot = datos_print["totales_llamadas"]
    tot_leads = datos_print["totales_leads"]
    campanas = datos_print["campanas"]
    leads = datos_print["leads"]

    # ---- Timeline animado de eficiencia diaria -----------------------------
    dias_con_datos = sorted(d for d, v in serie.items() if v["tiene_datos"])
    x_dias = dias_del_mes
    y_eficiencia = []
    for d in x_dias:
        v = serie[d]
        # Los dias posteriores al corte quedan como None (hueco), nunca como 0:
        # un cero dibujaria una caida a plomo que no ocurrio.
        if v["tiene_datos"] and v["llamadas_efic"]:
            y_eficiencia.append(v["ventas_linea"] / v["llamadas_efic"] * 100)
        else:
            y_eficiencia.append(None)

    ultimo_dia_con_datos = max(dias_con_datos) if dias_con_datos else x_dias[0]
    tope_eficiencia = max([y for y in y_eficiencia if y is not None] or [1]) * 1.18

    # Tendencia: regresion lineal simple sobre los dias con dato, para que se
    # vea de un vistazo si la eficiencia sube o baja.
    puntos = [(i, y) for i, y in enumerate(y_eficiencia) if y is not None]
    y_tendencia = None
    pendiente = 0.0
    if len(puntos) >= 2:
        n = len(puntos)
        sx = sum(p[0] for p in puntos)
        sy = sum(p[1] for p in puntos)
        sxy = sum(p[0] * p[1] for p in puntos)
        sxx = sum(p[0] * p[0] for p in puntos)
        denom = n * sxx - sx * sx
        if denom:
            pendiente = (n * sxy - sx * sy) / denom
            intercepto = (sy - pendiente * sx) / n
            idx_max = max(p[0] for p in puntos)
            y_tendencia = [
                (pendiente * i + intercepto) if i <= idx_max else None
                for i in range(len(x_dias))
            ]

    if pendiente > 0.01:
        texto_tendencia = f"▲ Tendencia al alza (+{pendiente:.2f} pp por día)"
        color_tendencia = C_VERDE
        clase_tendencia = "verde"
    elif pendiente < -0.01:
        texto_tendencia = f"▼ Tendencia a la baja ({pendiente:.2f} pp por día)"
        color_tendencia = "#DC2626"
        clase_tendencia = "rojo"
    else:
        texto_tendencia = "▬ Tendencia estable"
        color_tendencia = C_SUAVE
        clase_tendencia = ""
    # El estilo se repite en cada frame a proposito: al saltar de frame,
    # Plotly reemplaza la traza con lo que trae el frame, y si el frame solo
    # lleva los datos la serie se queda sin linea ni relleno.
    estilo_serie = dict(
        mode="lines+markers",
        # Linea recta entre dias, no curva: una curva inventaria valores
        # intermedios que no existen (solo hay un dato por dia).
        line=dict(color=C_AZUL, width=2.6),
        marker=dict(size=7, color="#FFFFFF", line=dict(color=C_AZUL, width=2.2)),
        name="Eficiencia /Línea diaria",
        fill="tozeroy", fillcolor="rgba(47,111,237,.10)",
        connectgaps=False,
        hovertemplate="%{x|%d/%m}: %{y:.2f} %<extra></extra>",
    )

    # El grafico carga con la serie completa: la pregunta numero uno del tablero
    # es si la eficiencia sube o baja, y eso tiene que verse sin pulsar nada.
    # El boton Reproducir la vuelve a dibujar dia a dia.
    #
    # La animacion NO usa frames de Plotly. Plotly.animate deja la traza sin
    # linea ni relleno (solo los puntos) al saltar de frame, asi que la
    # reproduccion se hace desde el propio tablero recortando la serie con
    # Plotly.restyle, que si redibuja bien.
    fig_timeline = go.Figure(data=[go.Scatter(x=x_dias, y=y_eficiencia, **estilo_serie)])
    if y_tendencia:
        fig_timeline.add_scatter(
            x=x_dias, y=y_tendencia, mode="lines", name="Tendencia",
            line=dict(color=color_tendencia, width=1.8, dash="dot"),
            hoverinfo="skip",
        )
    _estilo_plotly(fig_timeline, 380, margenes=dict(l=56, r=22, t=42, b=64))
    fig_timeline.update_layout(
        xaxis=dict(
            range=[x_dias[0], x_dias[-1]],
            tickmode="array",
            tickvals=x_dias,
            ticktext=[fmt_fecha_es(d) for d in x_dias],
        ),
        # Rango fijo: al recortar la serie durante la reproduccion, la escala
        # no se recalcula y la curva no da saltos de un dia a otro.
        yaxis=dict(ticksuffix=" %", range=[0, tope_eficiencia]),
    )

    # include_plotlyjs='inline' solo en el primer grafico: incrusta plotly.js
    # completo una sola vez para que el HTML abra sin internet.
    config_plotly = {"displaylogo": False, "responsive": True, "displayModeBar": False}
    html_timeline = plotly_plot(
        fig_timeline, include_plotlyjs="inline", output_type="div", config=config_plotly
    )

    # ---- Comparativa por campana (excluye 0 llamadas) ----------------------
    # Los graficos se dibujan con el nombre a mostrar, pero el contenedor
    # guarda las claves originales en data-campanas: es lo que le permite al
    # panel de Configuracion renombrar las barras sin regenerar el archivo.
    campanas_con_llamadas = [c for c in campanas if (c.get("Numero de llamadas") or 0) > 0]
    claves = [c["Campaña"] for c in campanas_con_llamadas]
    nombres = [nombre_visible(c) for c in claves]

    # El tramado diagonal de las barras es el mismo gesto que usan las barras
    # de reparto en HTML: mantiene los dos lenguajes graficos coherentes.
    def _tramado(color):
        return dict(shape="/", size=5, solidity=0.32, fgcolor="#FFFFFF", bgcolor=color)

    fig_barras = go.Figure()
    fig_barras.add_bar(
        name="Llamadas", x=nombres, y=[c["Numero de llamadas"] for c in campanas_con_llamadas],
        marker=dict(color=C_AZUL, pattern=_tramado(C_AZUL), cornerradius=6),
        hovertemplate="<b>%{x}</b><br>Llamadas: %{y:,.0f}<extra></extra>",
    )
    fig_barras.add_bar(
        name="Perdidas", x=nombres, y=[c["#Perdidas"] for c in campanas_con_llamadas],
        marker=dict(color=C_ROSA, pattern=_tramado(C_ROSA), cornerradius=6),
        hovertemplate="<b>%{x}</b><br>Perdidas: %{y:,.0f}<extra></extra>",
    )
    _estilo_plotly(fig_barras, 400, margenes=dict(l=52, r=18, t=44, b=104))
    fig_barras.update_layout(barmode="group", bargap=0.28, bargroupgap=0.12,
                             xaxis=dict(tickangle=-35))
    html_barras = plotly_plot(
        fig_barras, include_plotlyjs=False, output_type="div", config=config_plotly
    )

    # Una eficiencia imposible (>100 %) aplasta la escala y deja el resto de
    # barras invisibles. Esas campanas se dejan fuera de ESTE grafico y se
    # nombran bajo el titulo; siguen enteras en la tabla y en las notas.
    campanas_efic = [
        c for c in campanas_con_llamadas
        if (c.get("Eficiencia /Cliente") or 0) <= 1 and (c.get("Eficiencia /Linea") or 0) <= 1
    ]
    fuera_de_escala = [nombre_visible(c["Campaña"]) for c in campanas_con_llamadas
                       if c not in campanas_efic]
    claves_efic = [c["Campaña"] for c in campanas_efic]
    nombres_efic = [nombre_visible(c) for c in claves_efic]

    fig_eficiencia = go.Figure()
    fig_eficiencia.add_bar(
        name="Eficiencia /Cliente",
        x=nombres_efic, y=[(c.get("Eficiencia /Cliente") or 0) * 100 for c in campanas_efic],
        marker=dict(color=C_VERDE, pattern=_tramado(C_VERDE), cornerradius=6),
        hovertemplate="<b>%{x}</b><br>Eficiencia /Cliente: %{y:.2f} %<extra></extra>",
    )
    fig_eficiencia.add_bar(
        name="Eficiencia /Linea",
        x=nombres_efic, y=[(c.get("Eficiencia /Linea") or 0) * 100 for c in campanas_efic],
        marker=dict(color=C_VIOLETA, pattern=_tramado(C_VIOLETA), cornerradius=6),
        hovertemplate="<b>%{x}</b><br>Eficiencia /Linea: %{y:.2f} %<extra></extra>",
    )
    _estilo_plotly(fig_eficiencia, 400, margenes=dict(l=52, r=18, t=44, b=104))
    fig_eficiencia.update_layout(barmode="group", bargap=0.28, bargroupgap=0.12,
                                 xaxis=dict(tickangle=-35), yaxis=dict(ticksuffix=" %"))
    html_eficiencia = plotly_plot(
        fig_eficiencia, include_plotlyjs=False, output_type="div", config=config_plotly
    )

    # ---- Tarjetas KPI --------------------------------------------------------
    llamadas_totales = tot["Numero de llamadas"]
    llamadas_dia = llamadas_totales / datos_print["dias_habiles"] if datos_print["dias_habiles"] else None
    perdidas_totales = tot["#Perdidas"] or 0
    contestadas_totales = llamadas_totales - perdidas_totales
    efic_cliente = tot["Ventas (Cliente)"] / llamadas_totales if llamadas_totales else None
    efic_linea = tot["Ventas (Linea)"] / llamadas_totales if llamadas_totales else None

    tarjetas = [
        ("Llamadas totales", fmt_num(llamadas_totales),
         f'{fmt_num(llamadas_dia, 2)} por día hábil'),
        ("Llamadas perdidas", fmt_num(perdidas_totales),
         f'{fmt_pct(tot.get("% Llamadas Perdidas"))} del total'),
        ("Ventas (Cliente)", fmt_num(tot["Ventas (Cliente)"]),
         f'1 cada {fmt_num(tot.get("Llamadas /Cliente"), 2)} llamadas'),
        ("Ventas (Línea)", fmt_num(tot["Ventas (Linea)"]),
         f'1 cada {fmt_num(tot.get("Llamadas /Linea"), 2)} llamadas'),
        ("Monto", fmt_usd(tot["Monto"]),
         f'Total general {fmt_usd(datos_print["total_general"])}'),
        ("Eficiencia /Cliente", fmt_pct(efic_cliente), "Ventas cliente ÷ llamadas"),
        ("Eficiencia /Línea", fmt_pct(efic_linea), "Ventas línea ÷ llamadas"),
        ("Llamadas / día", fmt_num(llamadas_dia, 2),
         f'{fmt_num(datos_print["dias_habiles"])} días hábiles'),
    ]
    html_tarjetas = "".join(
        f'<div class="s3"><div class="kpi"><div class="kpi-label">{k}</div>'
        f'<div class="kpi-valor">{v}</div><div class="kpi-sub">{s}</div></div></div>'
        for k, v, s in tarjetas
    )

    # ---- Columnas: llamadas por campaña (bloque destacado) -------------------
    top_llamadas = sorted(
        campanas_con_llamadas, key=lambda c: c["Numero de llamadas"], reverse=True
    )[:5]
    tope_col, valores_eje = _escala_bonita(
        max((c["Numero de llamadas"] for c in top_llamadas), default=0)
    )
    # Se destaca la campaña con mejor eficiencia /Linea entre las cinco, dejando
    # fuera las imposibles (>100 %), que no reflejan rendimiento real.
    candidatas = [c for c in top_llamadas if (c.get("Eficiencia /Linea") or 0) <= 1]
    mejor = max(candidatas, key=lambda c: c.get("Eficiencia /Linea") or 0) if candidatas else None
    idx_foco = top_llamadas.index(mejor) if mejor else 0

    items_col = []
    for c in top_llamadas:
        pct_perd = c.get("% Llamadas Perdidas")
        items_col.append({
            "clave": c["Campaña"],
            "label": nombre_visible(c["Campaña"]),
            "valor_fmt": fmt_num(c["Numero de llamadas"]),
            "pct": c["Numero de llamadas"] / tope_col * 100 if tope_col else 0,
            "tip": (f'<b>{fmt_num(c["Numero de llamadas"])}</b> llamadas &nbsp;·&nbsp; '
                    f'Perdidas: <b>{fmt_pct(pct_perd)}</b> &nbsp;·&nbsp; '
                    f'Eficiencia /Línea: <b>{fmt_pct(c.get("Eficiencia /Linea"))}</b>'),
        })
    html_columnas = _bloque_columnas(
        items_col, idx_foco,
        [(v / tope_col * 100, fmt_num(v)) for v in valores_eje],
    )

    # ---- Reparto del monto por campaña --------------------------------------
    paleta_reparto = [C_VERDE, C_AZUL, C_ROSA, C_AMBAR, C_VIOLETA]
    con_monto = sorted(
        [c for c in campanas if (c.get("Monto") or 0) > 0],
        key=lambda c: c["Monto"], reverse=True,
    )
    monto_max = con_monto[0]["Monto"] if con_monto else 0
    html_reparto = _barras_reparto([
        (c["Campaña"], nombre_visible(c["Campaña"]), fmt_usd(c["Monto"]),
         c["Monto"] / monto_max * 100 if monto_max else 0,
         paleta_reparto[i % len(paleta_reparto)])
        for i, c in enumerate(con_monto)
    ])
    pct_del_general = (
        tot["Monto"] / datos_print["total_general"] if datos_print["total_general"] else None
    )

    # ---- Ritmo diario: área escalonada y matrices de puntos ------------------
    serie_llamadas = [(d, serie[d]["llamadas"]) for d in dias_con_datos]
    serie_perdidas = [(d, serie[d]["perdidas"]) for d in dias_con_datos]
    serie_contestadas = [(d, serie[d]["llamadas"] - serie[d]["perdidas"]) for d in dias_con_datos]

    html_area_llamadas = (
        _area_escalonada(serie_llamadas, C_ROSA, etiqueta_valor="llamadas")
        + f'<div class="eje-x"><span>{fmt_fecha_es(serie_llamadas[0][0])}</span>'
        f'<span>{fmt_fecha_es(serie_llamadas[len(serie_llamadas) // 2][0])}</span>'
        f'<span>{fmt_fecha_es(serie_llamadas[-1][0])}</span></div>'
        if serie_llamadas else ""
    )

    def _tarjeta_puntos(titulo, ayuda, total, serie_pares, color, pie_label, pie_valor,
                        etiqueta_valor="llamadas"):
        # La matriz va en su propia linea, a lo ancho de la tarjeta. Antes
        # compartia fila con la cifra y el porcentaje y le quedaban unos 180 px
        # para 31 dias: los puntos salian como rayas verticales de 3 px.
        if not serie_pares:
            return f'<div class="tarjeta">{_cabecera_tarjeta(titulo, ayuda)}</div>'
        d_pico = max(serie_pares, key=lambda p: p[1])
        return (
            f'<div class="tarjeta">{_cabecera_tarjeta(titulo, ayuda)}'
            f'<div class="tarjeta-punto">'
            f'<div class="punto-cifras">'
            f'<div class="lado">'
            f'<div class="etiqueta-pico"><span class="punto" style="background:{color}"></span>'
            f'Pico: <b>{fmt_fecha_es(d_pico[0])}</b></div>'
            f'<div class="cifra-md">{total}</div></div>'
            f'<div class="lado lado-der">{pie_label}<b>{pie_valor}</b></div>'
            f'</div>'
            f'<div class="centro">{_matriz_puntos(serie_pares, color, etiqueta_valor=etiqueta_valor)}</div>'
            f'</div></div>'
        )

    html_contestadas = _tarjeta_puntos(
        "Llamadas contestadas",
        "Llamadas totales menos perdidas, día a día. Cada columna es un día y el número de "
        "puntos es el volumen relativo; el día pico va en color pleno. Pasa el ratón por una "
        f"columna para ver su fecha y su cifra. Fuente: hoja '{HOJA_RAW_LLAMADAS}'.",
        fmt_num(contestadas_totales), serie_contestadas, C_VERDE,
        "de las llamadas",
        fmt_pct(contestadas_totales / llamadas_totales if llamadas_totales else None),
        etiqueta_valor="contestadas",
    )
    html_perdidas = _tarjeta_puntos(
        "Llamadas perdidas",
        "Llamadas que entraron a la cola y nadie atendió, día a día. Pasa el ratón por una "
        f"columna para ver su fecha y su cifra. Fuente: hoja '{HOJA_RAW_LLAMADAS}'; el total "
        f"coincide con el de la hoja '{HOJA_PRINT}'.",
        fmt_num(perdidas_totales), serie_perdidas, C_ROSA,
        "de las llamadas", fmt_pct(tot.get("% Llamadas Perdidas")),
        etiqueta_valor="perdidas",
    )

    # ---- Tarjeta de conclusión ----------------------------------------------
    dias_transcurridos = fecha_corte.day
    dias_mes = len(dias_del_mes)
    if pendiente > 0.01:
        titular = (f"La eficiencia diaria sube {fmt_num(pendiente, 2)} puntos por día "
                   f"en lo que va del mes.")
    elif pendiente < -0.01:
        titular = (f"La eficiencia diaria baja {fmt_num(abs(pendiente), 2)} puntos por día "
                   f"en lo que va del mes.")
    else:
        titular = "La eficiencia diaria se mantiene estable en lo que va del mes."
    detalle_insight = (
        f'{fmt_num(tot["Ventas (Linea)"])} ventas de línea sobre {fmt_num(llamadas_totales)} '
        f'llamadas atendidas en {fmt_num(datos_print["dias_habiles"])} días hábiles. '
        f'Se perdió {fmt_pct(tot.get("% Llamadas Perdidas"))} de las llamadas: recuperar esa '
        f'proporción al ritmo actual valdría unas '
        f'{fmt_num(round(perdidas_totales * (efic_linea or 0)))} ventas más.'
    )
    html_insight = (
        f'<div class="insight">'
        f'<div><div class="insight-pill">{_icono_bombilla()} Lectura del corte</div>'
        f'<div class="cifra" style="margin-top:18px">{fmt_pct(efic_linea)}</div>'
        f'<h3>{titular}</h3>'
        f'<p>{detalle_insight}</p></div>'
        f'<div class="insight-avance">'
        f'<div class="pista-b"><i style="width:{dias_transcurridos / dias_mes * 100:.1f}%"></i></div>'
        f'<div class="txt">Día {dias_transcurridos} de {dias_mes} del mes</div>'
        f'</div></div>'
    )

    # ---- Tabla de detalle (mismas columnas y nombres que Print) --------------
    columnas_tabla = list(campanas[0].keys()) if campanas else []

    def _fmt_celda(nombre_col, valor, sin_actividad=False):
        if nombre_col == "Campaña":
            return nombre_visible(valor) if valor else ""
        # Una campana sin llamadas no tiene nada medido: todas sus celdas van
        # como guion. Un 0 sugeriria que se midio y dio cero.
        if sin_actividad or valor is None:
            return "—"
        if "%" in nombre_col or "Eficiencia" in nombre_col:
            return fmt_pct(valor)
        if nombre_col in ("Monto", "Costo Llamada", "Costo /Cliente", "Costo /Linea"):
            return fmt_usd(valor)
        if isinstance(valor, float) and not valor.is_integer():
            return fmt_num(valor, 2)
        return fmt_num(valor)

    filas_html = []
    for c in campanas:
        sin_actividad = not (c.get("Numero de llamadas") or 0)
        celdas = "".join(
            (f'<td data-campana="{c.get(col)}">{_fmt_celda(col, c.get(col), sin_actividad)}</td>'
             if col == "Campaña" and c.get(col)
             else f"<td>{_fmt_celda(col, c.get(col), sin_actividad)}</td>")
            for col in columnas_tabla
        )
        filas_html.append(f"<tr>{celdas}</tr>")
    # En 'Print' la fila de totales deja vacias las dos eficiencias (viven aparte,
    # en las mini-tarjetas de abajo). Se recalculan aqui para que la fila TOTAL
    # de la tabla no muestre un hueco donde si hay dato.
    tot_completo = dict(tot)
    if llamadas_totales:
        tot_completo.setdefault("Eficiencia /Cliente", None)
        tot_completo.setdefault("Eficiencia /Linea", None)
        if tot_completo.get("Eficiencia /Cliente") is None:
            tot_completo["Eficiencia /Cliente"] = tot["Ventas (Cliente)"] / llamadas_totales
        if tot_completo.get("Eficiencia /Linea") is None:
            tot_completo["Eficiencia /Linea"] = tot["Ventas (Linea)"] / llamadas_totales

    fila_totales_html = "".join(
        f"<td>{_fmt_celda(col, tot_completo.get(col)) if col != 'Campaña' else 'TOTAL'}</td>"
        for col in columnas_tabla
    )
    html_tabla = f"""
    <div class="tabla-wrap">
      <table class="tabla-detalle">
        <thead><tr>{''.join(f'<th>{c}</th>' for c in columnas_tabla)}</tr></thead>
        <tbody>{''.join(filas_html)}<tr class="fila-total">{fila_totales_html}</tr></tbody>
      </table>
    </div>
    """

    # ---- Bloque de Leads -------------------------------------------------
    filas_leads_html = []
    for lead in leads:
        sin_leads = not (lead.get("Numero de leads") or 0)
        g = (lambda valor, f: "—" if sin_leads or valor is None else f(valor))
        filas_leads_html.append(
            f'<tr><td data-campana="{lead["Campaña"]}">{nombre_visible(lead["Campaña"])}</td>'
            f"<td>{g(lead.get('Numero de leads'), fmt_num)}</td>"
            f"<td>{g(lead.get('Costo Lead'), fmt_usd)}</td>"
            f"<td>{g(lead.get('Monto'), fmt_usd)}</td>"
            f"<td>{g(lead.get('Eficiencia /Cliente'), fmt_pct)}</td>"
            f"<td>{g(lead.get('Eficiencia /Linea'), fmt_pct)}</td></tr>"
        )
    html_leads = f"""
    <div class="tabla-wrap">
      <table class="tabla-detalle">
        <thead><tr><th>Campaña</th><th>Leads</th><th>Costo /Lead</th><th>Monto</th>
        <th>Eficiencia /Cliente</th><th>Eficiencia /Linea</th></tr></thead>
        <tbody>{''.join(filas_leads_html)}</tbody>
      </table>
    </div>
    """

    # ---- Avisos / notas de calidad de datos -------------------------------
    notas = []
    if campanas_excluidas_serie:
        notas.append(
            "El Excel de origen no trae desglose día a día para "
            + ", ".join(nombre_visible(c) for c in sorted(campanas_excluidas_serie))
            + " (sus ventas coinciden exactamente con las de otra campaña, sin bloque propio en "
              "'" + HOJA_RAW_VENTAS + "'). Por eso el gráfico de eficiencia diaria las deja fuera "
              "del cálculo completo, tanto en ventas como en llamadas, para que numerador y "
              "denominador cubran las mismas campañas. Las tarjetas de KPI y la tabla de detalle "
              "sí las incluyen: esas cifras vienen de 'Print' y están completas."
        )
    if costo_corte:
        dias_str = ", ".join(f"{fmt_fecha_es(f)}: {fmt_usd(v)}" for f, v in costo_corte)
        notas.append(
            f"Costo capturado a la fecha ({dias_str}). La captura de costos es manual y va con "
            f"retraso, por eso se muestra como dato del corte y no como serie diaria: los días "
            f"en cero significan «aún no cargado», no «sin inversión»."
        )
    # Una eficiencia por encima del 100 % significa mas ventas que llamadas, lo
    # que no puede pasar: delata que las ventas de esa fila no corresponden a sus
    # llamadas en el Excel de origen. Se avisa porque el numero se ve en la tabla.
    imposibles = [
        nombre_visible(c["Campaña"]) for c in campanas
        if (c.get("Numero de llamadas") or 0) > 0 and (c.get("Eficiencia /Cliente") or 0) > 1
    ]
    if imposibles:
        notas.append(
            "Eficiencia superior al 100 % en: " + ", ".join(imposibles)
            + ". Son más ventas que llamadas, algo imposible en la operación: en 'Print' esas "
              "filas tienen ventas que no corresponden a sus propias llamadas. El dato se muestra "
              "tal como está en el Excel, sin corregir, pero no debe leerse como rendimiento real."
        )

    notas.extend(avisos)
    html_notas = "".join(f'<li>{n}</li>' for n in notas)
    # El bloque es uno solo y se mueve de sitio: por defecto vive dentro de
    # Configuracion y el interruptor de alli lo trae al tablero. Asi no hay dos
    # copias que puedan acabar diciendo cosas distintas.
    bloque_notas = (
        f'<div class="notas" id="bloque-notas"><div class="notas-cab">'
        f'<span class="punto" style="background:#E8963C"></span>Notas de datos</div>'
        f'<ul>{html_notas}</ul></div>'
    ) if notas else ""

    # ---- Cabeceras de las tarjetas grandes ----------------------------------
    ayuda_columnas = (
        "Las cinco campañas con más llamadas en el corte. La barra es el volumen de llamadas y "
        "la columna destacada es la de mejor eficiencia /Línea entre esas cinco. Pasa el ratón "
        f"por encima para ver perdidas y eficiencia. Fuente: hoja '{HOJA_PRINT}'."
    )
    ayuda_monto = (
        "Monto facturado por las campañas de llamadas en el corte, repartido por campaña. Las "
        "barras están a escala de la campaña mayor. El total general incluye además los leads "
        f"de META. Fuente: hoja '{HOJA_PRINT}'."
    )
    ayuda_area = (
        "Llamadas recibidas cada día con actividad. La etiqueta marca el día pico. Los días "
        "posteriores al corte no se dibujan. Fuente: hoja "
        f"'{HOJA_RAW_LLAMADAS}'; el total coincide con el de '{HOJA_PRINT}'."
    )
    ayuda_eficiencia = (
        "Ventas de línea divididas entre las llamadas de cada día. La línea punteada es la "
        "tendencia (regresión lineal sobre los días con dato). El botón Reproducir vuelve a "
        "dibujar la serie día a día. Los días sin dato quedan como hueco, nunca como cero."
    )
    ayuda_barras = (
        "Llamadas recibidas y llamadas perdidas de cada campaña con actividad en el corte. "
        f"Fuente: hoja '{HOJA_PRINT}'."
    )
    ayuda_efic_campana = (
        "Ventas divididas entre llamadas, por campaña. Sobre cliente (una venta por cliente) y "
        f"sobre línea (una venta por línea contratada). Fuente: hoja '{HOJA_PRINT}'."
    )
    pie_efic_campana = (
        "Fuera de escala: " + ", ".join(fuera_de_escala) + " (eficiencia imposible, ver notas)"
    ) if fuera_de_escala else "Ventas ÷ llamadas, por campaña"

    insignia_tendencia = (
        f'<span class="insignia {clase_tendencia}">{texto_tendencia}</span>'
        if y_tendencia else ""
    )

    # Etiquetas y tope del control de reproduccion (solo hasta el corte: los
    # dias siguientes no tienen dato que dibujar).
    dias_hasta_corte = [d for d in x_dias if d <= ultimo_dia_con_datos]
    dias_reproductor = "|".join(fmt_fecha_es(d) for d in dias_hasta_corte)
    indice_ultimo_dia = len(dias_hasta_corte) - 1

    # ---- Panel de configuracion ---------------------------------------------
    # Las campanas de MAPEO_COLAS que no salen en el Excel de este mes tambien
    # se listan: son las que estan a la espera (una campana que vuelve, una que
    # todavia no empezo) y si no aparecieran, guardar el archivo las borraria.
    claves_llamadas = [c["Campaña"] for c in campanas]
    claves_llamadas += [k for k in MAPEO_COLAS if k not in claves_llamadas]
    claves_leads = [lead["Campaña"] for lead in leads]
    notas_en_tablero = bool(contexto.get("config", {}).get("mostrar_notas_en_dashboard"))
    html_configuracion = _panel_configuracion(
        claves_llamadas, claves_leads, bloque_notas, notas_en_tablero
    )
    notas_del_tablero = bloque_notas if notas_en_tablero else ""
    globo_notas = f'<span class="globo">{len(notas)}</span>' if notas else ""

    html = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{NOMBRE_MARCA} · {contexto['mes_nombre']}</title>
<style>{CSS_DASHBOARD}</style>
</head>
<body>

<div class="barra-sup">
  <div class="interior">
    <div class="marca">{_svg_marca()}<span>{NOMBRE_MARCA}</span></div>
    <nav class="nav">
      <a href="#resumen">Resumen</a>
      <a href="#actividad">Actividad</a>
      <a href="#ritmo">Ritmo diario</a>
      <a href="#eficiencia">Eficiencia</a>
      <a href="#campanas">Campañas</a>
      <a href="#detalle">Detalle</a>
      <a href="#leads">Leads</a>
    </nav>
    <div class="acciones-sup">
      <button type="button" class="btn-icono" id="btn-config"
              title="Nombres de campaña, colas y notas de datos">
        {_icono_engranaje()}<span>Configuración</span>{globo_notas}
      </button>
    </div>
  </div>
</div>

<div class="envoltura">

  <div class="titulo-fila">
    <div>
      <h1>{contexto['mes_nombre']}</h1>
      <div class="sub">{NOMBRE_MARCA} · datos al corte del {fecha_corte.strftime('%d/%m/%Y')}</div>
    </div>
    <div class="herramientas">
      <span class="chip">{_icono_calendario()} {contexto['mes_nombre']}</span>
      <span class="chip plano">corte al</span>
      <span class="chip">{_icono_calendario()} {fmt_fecha_es(fecha_corte)} {fecha_corte.year}</span>
      <span class="chip">{datos_print['dias_habiles']} días hábiles</span>
      <span class="chip">{_icono_hoja()} {os.path.basename(ruta_excel)}</span>
      <span class="chip">{_icono_reloj()} {hora_generacion.strftime('%d/%m/%Y %H:%M')}</span>
    </div>
  </div>

  <div id="notas-en-tablero">{notas_del_tablero}</div>

  <section id="resumen">
    <div class="rotulo">Resumen del corte</div>
    <div class="rejilla">{html_tarjetas}</div>
  </section>

  <section id="actividad">
    <div class="rotulo">Actividad por campaña</div>
    <div class="rejilla">
      <div class="s8">
        <div class="tarjeta crece">
          {_cabecera_tarjeta("Llamadas por campaña", ayuda_columnas, "Las cinco campañas con más volumen en el corte")}
          {html_columnas}
        </div>
      </div>
      <div class="s4">
        <div class="tarjeta">
          {_cabecera_tarjeta("Monto del corte", ayuda_monto)}
          <div style="display:flex; align-items:center; gap:14px; flex-wrap:wrap">
            <div class="cifra">{fmt_usd(tot['Monto'])}</div>
            <span class="insignia"><span class="punto" style="background:{C_VERDE}"></span>{fmt_pct(pct_del_general)} del total general</span>
          </div>
          {html_reparto}
        </div>
      </div>
    </div>
  </section>

  <section id="ritmo">
    <div class="rotulo">Ritmo diario</div>
    <div class="rejilla">
      <div class="s4">
        <div class="tarjeta crece">
          {_cabecera_tarjeta("Llamadas por día", ayuda_area, f"{fmt_num(llamadas_totales)} llamadas en {len(dias_con_datos)} días con actividad")}
          {html_area_llamadas}
        </div>
      </div>
      <div class="s4">
        <div class="pila">
          {html_contestadas}
          {html_perdidas}
        </div>
      </div>
      <div class="s4">{html_insight}</div>
    </div>
  </section>

  <section id="eficiencia" data-dias="{dias_reproductor}">
    <div class="rotulo">Eficiencia diaria</div>
    <div class="rejilla">
      <div class="s12">
        <div class="tarjeta">
          {_cabecera_tarjeta("Eficiencia /Línea día a día", ayuda_eficiencia, "Ventas de línea ÷ llamadas de las campañas con desglose diario", insignia_tendencia)}
          <div class="plot">{html_timeline}</div>
          <div class="repro">
            <button type="button" class="btn-repro">Reproducir</button>
            <input type="range" class="rango-dia" min="0" max="{indice_ultimo_dia}" value="{indice_ultimo_dia}"
                   step="1" aria-label="Día del mes a mostrar">
            <span class="rango-etiqueta">Día <b>{fmt_fecha_es(ultimo_dia_con_datos)}</b></span>
          </div>
        </div>
      </div>
    </div>
  </section>

  <section id="campanas">
    <div class="rotulo">Comparativa por campaña</div>
    <div class="rejilla">
      <div class="s6">
        <div class="tarjeta">
          {_cabecera_tarjeta("Llamadas y perdidas", ayuda_barras, "Campañas con actividad en el corte")}
          <div class="plot" data-campanas="{_esc('|'.join(claves))}">{html_barras}</div>
        </div>
      </div>
      <div class="s6">
        <div class="tarjeta">
          {_cabecera_tarjeta("Eficiencia por campaña", ayuda_efic_campana, pie_efic_campana)}
          <div class="plot" data-campanas="{_esc('|'.join(claves_efic))}">{html_eficiencia}</div>
        </div>
      </div>
    </div>
  </section>

  <section id="detalle">
    <div class="rotulo">Detalle</div>
    <div class="rejilla">
      <div class="s12">
        <div class="tarjeta">
          {_cabecera_tarjeta("Detalle por campaña", "Todas las columnas de la tabla de llamadas de la hoja '" + HOJA_PRINT + "', tal cual, sin recalcular. Las campañas sin llamadas van con guion en vez de cero: no se midieron.", "Mismas columnas y nombres que el Excel de origen")}
          {html_tabla}
        </div>
      </div>
    </div>
  </section>

  <section id="leads">
    <div class="rotulo">Leads</div>
    <div class="rejilla">
      <div class="s12">
        <div class="tarjeta">
          {_cabecera_tarjeta("Leads (META)", "Segunda tabla de la hoja '" + HOJA_PRINT + "': campañas que llegan por formulario de META en vez de por llamada. Su costo y sus ventas van aparte de las de llamadas.", "Campañas captadas por formulario, no por llamada")}
          {html_leads}
        </div>
      </div>
    </div>
  </section>

  <div class="pie">
    Generado el {hora_generacion.strftime('%d/%m/%Y a las %H:%M')} desde {os.path.basename(ruta_excel)}
  </div>

</div>

{html_configuracion}

<script>{JS_DASHBOARD}</script>
</body>
</html>
"""
    return html


# =============================================================================
# Main
# =============================================================================

def main():
    # Lo primero: configuracion.json puede cambiar las colas de una campana, y
    # eso tiene que estar aplicado antes de leer nada.
    config = cargar_configuracion()
    ruta_excel = archivo_mas_reciente(CARPETA_ENTRADA)
    hora_generacion = datetime.now()

    # Una sola apertura, siempre en read_only: el libro trae hojas enormes que
    # no se usan y en modo normal se cargarian enteras.
    wb = load_workbook(ruta_excel, data_only=True, read_only=True)
    try:
        hoja_print = hoja_a_memoria(wb, HOJA_PRINT)
        hoja_ventas = hoja_a_memoria(wb, HOJA_RAW_VENTAS)
        hoja_costos = hoja_a_memoria(wb, HOJA_COSTOS) if HOJA_COSTOS in wb.sheetnames else None
        # 'Print' se lee antes que las llamadas: hace falta saber que campanas
        # existen para decidir si hay que ir a buscar alguna a la hoja extra.
        datos_print = extraer_print(hoja_print)
        llamadas_agregadas, campanas_rescatadas = extraer_llamadas_diarias(
            wb, datos_print["campanas"]
        )
    finally:
        wb.close()

    bloques = extraer_bloques_raw_data_general(hoja_ventas)

    campanas_objetivo = {
        c["Campaña"]: (c.get("Ventas (Cliente)") or 0, c.get("Ventas (Linea)") or 0, c.get("Monto") or 0)
        for c in datos_print["campanas"]
    }
    bloques_asignados, avisos, campanas_excluidas_serie = asignar_bloques_a_campanas(bloques, campanas_objetivo)

    costo_corte = extraer_costo_corte(hoja_costos) if hoja_costos else None

    dias_del_mes, serie, fecha_corte = construir_serie_diaria(
        datos_print, llamadas_agregadas, bloques_asignados, datos_print["mes"], datos_print["fecha_corte"]
    )

    if campanas_rescatadas:
        por_hoja = defaultdict(list)
        for campana, hoja in campanas_rescatadas:
            por_hoja[hoja].append(campana)
        for hoja, campanas in por_hoja.items():
            avisos.append(
                "Las llamadas de "
                + ", ".join(nombre_visible(c) for c in sorted(campanas))
                + " no estaban en la hoja '" + HOJA_RAW_LLAMADAS + "' y se tomaron de '"
                + hoja + "', que este mes es su única fuente."
            )

    colas_del_raw = {cola for (_fecha, cola) in llamadas_agregadas}
    avisos = validar(datos_print, serie, avisos, colas_del_raw)

    contexto = {
        "datos_print": datos_print,
        "dias_del_mes": dias_del_mes,
        "serie": serie,
        "fecha_corte": fecha_corte,
        "ruta_excel": ruta_excel,
        "hora_generacion": hora_generacion,
        "avisos": avisos,
        "campanas_excluidas_serie": campanas_excluidas_serie,
        "costo_corte": costo_corte,
        "mes_nombre": fmt_mes_es(datos_print["mes"]),
        "config": config,
    }

    html = construir_html(contexto)

    os.makedirs(CARPETA_SALIDA, exist_ok=True)
    ruta_salida = os.path.join(CARPETA_SALIDA, NOMBRE_SALIDA)
    with open(ruta_salida, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"OK. Dashboard generado en: {ruta_salida}")
    print(f"Archivo de origen: {ruta_excel}")
    print(f"Corte: {fecha_corte} | Dias habiles: {datos_print['dias_habiles']}")
    if os.path.isfile(RUTA_CONFIG):
        print(f"Configuracion aplicada desde: {RUTA_CONFIG}")
        if config["nombres_campanas"]:
            print(f"  Nombres cambiados: {len(config['nombres_campanas'])}")
        if config["mapeo_colas"]:
            print(f"  Campanas con colas definidas ahi: {len(config['mapeo_colas'])}")
    if avisos:
        print("\nAvisos:")
        for a in avisos:
            print(f"  - {a}")

    # Lo abre el script y no el .bat porque solo aqui se conoce la ruta final:
    # si CARPETA_SALIDA apunta a otro sitio, el .bat no sabria donde buscarlo.
    if ABRIR_AL_TERMINAR and hasattr(os, "startfile"):
        try:
            os.startfile(ruta_salida)
        except OSError as e:
            print(f"\n(No se pudo abrir el navegador: {e}. El dashboard ya esta generado.)")


if __name__ == "__main__":
    try:
        main()
    except ErrorDatosExcel as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        sys.exit(1)
